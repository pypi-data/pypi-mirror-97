import openpyxl
import sys, os, requests, uuid, json
from requests.exceptions import HTTPError
from os import lseek, path

class c_Args:
    ''' 
    Argument initialization class
    Update this to provide default 
    '''
    def __init__ (self, **kwargs):
        self._theSpreadsheet = kwargs['theSpreadsheet'] if 'theSpreadsheet' in kwargs else "nothing"
        self._columnSource = kwargs['columnSource'] if 'columnSource' in kwargs else 0
        self._columnTarget = kwargs['columnTarget'] if 'columnTarget' in kwargs else 0
        self._sourceLang = kwargs['sourceLang'] if 'sourceLang' in kwargs else "en"
        self._targetLang = kwargs['targetLang'] if 'targetLang' in kwargs else "en"
        self._subKey = kwargs['subKey'] if 'subKey' in kwargs else 0 # add your default key

    def theSpreadsheet(self, t = None): # Setter and getter , None is the default value
        if t: self._theSpreadsheet = t  
        return self._theSpreadsheet 

    def columnSource(self, t = None): 
        if t: self._columnSource = t  
        return self._columnSource

    def columnTarget(self, t = None): 
        if t: self._columnTarget = t  
        return self._columnTarget   

    def sourceLang(self, t = None): 
        if t: self._sourceLang= t  
        return self._sourceLang   

    def targetLang(self, t = None): 
        if t: self._targetLang = t  
        return self._targetLang 

    def subKey(self, t = None): 
        if t: self._subKey = t  
        return self._subKey  

def parse_args(*args):
    #global the_spreasheet, ColumnSource, ColumnTarget, SourceLang, TargetLang
    numargs = len(sys.argv)-1
    if numargs < 6:
        raise TypeError(f'Expected 6 arguments, got {numargs}')
    if  not path.exists(sys.argv[1]) or  not path.isfile(sys.argv[1]):
        raise TypeError(f'Spreadsheet not found')
    ## return an object containng the arguments
    return c_Args(theSpreadsheet = sys.argv[1], columnSource=sys.argv[2], columnTarget=sys.argv[3], sourceLang=sys.argv[4], targetLang=sys.argv[5], subKey=sys.argv[6])

def main ():
    if parse() == 0:
        print ("Translation added")

def parse():
    """ This script will parse a column of a spreasheet, send the strings to Bing Translator, copy the translation into anoher column
        Usage: takes 5 arguments
        translateFeedback <spreadsheet> <Column to translate from> <Column to translate to> <language source> <language target> <subscription key>
    """
    index_of_translated=0
    try:
        args=parse_args()
    except TypeError as e:
         print (f'Error:{e}')    
    else:
        # collect variables
        #
        file=args.theSpreadsheet()
        source_column_number = int(args.columnSource()) # this correspond to the column who has the feedback to be translated
        target_column_number = int(args.columnTarget()) # this correspond to the column who has the feedback to be translated
        sourceLang = args.sourceLang()
        targetLang = args.targetLang()

        ##### open the excel file:
        excel_file = openpyxl.load_workbook(file)
        #retrieve name of first sheet:
        feedback_sheetname=excel_file.sheetnames[0]
        feedback_sheet= excel_file[feedback_sheetname]

        # validation
        # check if the column indicated exists in the spreadsheet    
        if source_column_number>feedback_sheet.max_column:
            raise TypeError(f'Column number provided out of range, column number max={feedback_sheet.max_column}')
        
        # initiate Translation API call
        subscription_key = str(args._subKey)
        # language supported : https://docs.microsoft.com/en-us/azure/cognitive-services/translator/language-support
        params = sourceLang + '&to=' + targetLang
        endpoint_translate = "https://api.cognitive.microsofttranslator.com/translate?api-version=3.0&from="
        endpoint_languages = "https://api.cognitive.microsofttranslator.com/languages?api-version=3.0&scope=translation"
        header = {
            'Ocp-Apim-Subscription-Key': subscription_key,
            'Content-type': 'application/json',
            'X-ClientTraceId': str(uuid.uuid4())
        }
        urlData = endpoint_translate + params
        settings =[endpoint_languages, endpoint_translate, sourceLang, targetLang, header, params]
        #print (validate_translation_api (settings))
        for index in range(2, feedback_sheet.max_row+1):
            entire_cell_value = feedback_sheet.cell(row=index, column=source_column_number).value
            if entire_cell_value != None and len(entire_cell_value)<5000:  #body text passed to Translator API cannot exceed 5000 char
                body  = [{
                    'Text': entire_cell_value,
                }]
                request = requests.post (urlData, headers=header, json=body)
                response_json = request.json()
                translated_text = response_json[0]["translations"][0]['text']
                feedback_sheet.cell(row=index, column=target_column_number).value=translated_text
                index_of_translated+=1 #tracks how many cells have been written
                print('.', end='', flush=True)
        index-=1 #discount the header row to report how many rows were scaned
        print (f'\nTranslated {index_of_translated} out of {index} cells in {file}')  
        #save the spreadsheet when done   
        excel_file.save(file)
        return 0

def validate_translation_api(settings):
    request = requests.get (settings[0], headers=settings[4])
    response_json = request.json() #response_json is a dict
    
    # verify validity of language pairs
    for i in range(2,4):
        if settings[i] not in response_json["translation"]:
            raise TypeError(f'{settings[i]} not a supported language')
    
    # verify subsrcription key and availabtily of service
    try:
        request = requests.post (settings[1]+settings[5], headers=settings[4], json=[{'Text': "test",}])
        request.raise_for_status()
    except HTTPError as http_err:
        print (f'HTTP error occured: {http_err}')
    except Exception as err:
        print (f'Other error occured: {err}')
            
if __name__ == '__main__': main()
