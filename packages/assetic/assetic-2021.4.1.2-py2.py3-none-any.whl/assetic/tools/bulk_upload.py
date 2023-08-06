"""Copyright 2016 Assetic."""

from __future__ import absolute_import
import os
import time
import sys
import base64
import openpyxl  # to read excel files
import six
import collections

# import assetic libraries we'll use
from assetic.api import DataExchangeJobApi
from assetic.api import DataExchangeTaskApi
from assetic.api import DocumentApi
from assetic.models.file_properties_representation import \
    FilePropertiesRepresentation
from assetic.models.document_representation import DocumentRepresentation
from assetic.models.data_exchange_job_representation import \
    DataExchangeJobRepresentation
from assetic.rest import ApiException
from assetic.api_client import ApiClient
from assetic.api import WorkOrderApi


class BulkProcesses(object):
    """Class to manage bulk processes."""

    def __init__(self, configfile=None, logfile = None, loglevelname = None
                 , api_client=None):
        # Initialise class BulkProcesses standalone from package - this
        # supports legacy implementations of package
        if configfile:
            from ..assetic_sdk import AsseticSDK
            # initialise the SDK to setup connection and logging
            asseticsdk = AsseticSDK(configfile,logfile,loglevelname)
            try:
                api_client = asseticsdk.client
            except:
                # SDK not initialised properly
                message = "Assetic SDK not initialised.  Aborting import"
                raise ValueError(message)

        if api_client is None:
            api_client = ApiClient()
        self.api_client = api_client

        self.logger = api_client.configuration.packagelogger
        self.docapi = DocumentApi(self.api_client)
        #self.docapi_get = DocumentApi(self.api_client_for_docs) #also gets response header
        self.dataexchangeapi = DataExchangeJobApi(self.api_client)
        self.dataExchangeTaskapi = DataExchangeTaskApi(self.api_client)
        self.work_order_api = WorkOrderApi(self.api_client)
        self.host = self.api_client.configuration.host
        
    def upload_document_photo(self, excelDataFileToRead,groupid=None,
                              categoryid=None,subcategoryid=None,aslink=False,
                              parent_type=None):
        """Automate document/photo upload in bulk fashion.

        Usage:
            Invoke the script: python.exe -c "import assetic; uploadapi = assetic.BulkProcesses('C:/Users/xyz/assetic.ini'); uploadapi.upload_document_photo('C:/Users/xyz/data.xlsx')"
        :param excelDataFileToRead:
            XLSX/XLSM/XLTX/XLTM file containing document/photos data to import;
            Columns required are as follows (in EXACT ORDER)
            Id | Document/Photos Complete File Path |
            Is Key-photo (TRUE/FALSE)|External ID|Label|Description
        :param groupid: The group ID to use.  If None then 7 for Photos
        otherwise 10 (Documents)
        :param categoryid: The default category to use. No default if None
        :param subcategoryid: The default subcategory to use.No default if None
        :param aslink: if True, then create documents as links, else upload
        :param parent_type: The record type to load the document against,
        i.e. asset, functional location, work order
        default is False 
        :Returns: Uploaded document/photos status is written back to excel file specified in excelDataFileToRead
        """
        self.logger.info('File for bulk document upload operation is '\
                         'specified:{0}\n'.format(excelDataFileToRead))

        validDocumentToImportExtensions = ["xlsx", "xlsm", "xltx", "xltm"]

        if groupid == 0 or groupid.strip(" ") == "":
            groupid = None

        if not os.path.exists(excelDataFileToRead) or \
                os.path.splitext(excelDataFileToRead.lower())[1][1:] not in \
                validDocumentToImportExtensions:
            self.logger.error('File: [{0}] is either not specified or not of '
                              'correct format (Supported formats are: .xlsx,'
                              '.xlsm,.xltx,.xltm)\n'.format(
                excelDataFileToRead))
            return

        try:
            wb = openpyxl.load_workbook(excelDataFileToRead)
            sheet = wb.active
            sheet['G' + str(1)] = 'Migrated Document Id'
            sheet['H' + str(1)] = 'Operation Performed On'
            sheet['I' + str(1)] = 'Status'
            wb.save(excelDataFileToRead)
        except Exception as e:
            self.logger.error('{0}\n'.format(e))
            return

        for row in range(2, sheet.max_row + 1):
            msg = ""
            doc_id = None
            parent_id = sheet['A' + str(row)].value
            if not parent_id or str(parent_id).strip() == "":
                # This should cover empty rows as well as missing id.
                msg = "No Id defined for row {0}, skipping " \
                      "row".format(str(row))
                self._log_message_to_sheet(sheet, row, None, msg)
                self.logger.warning(msg)
                continue
            # get the filename and change backslashes, json errors otherwise
            fileNameToUpload = sheet['B' + str(row)].value.replace("\\", "/")
            isKeyPhotoColumnValue = sheet['C' + str(row)].value
            isKeyPhoto = False
            if isKeyPhotoColumnValue:
                isKeyPhoto = isKeyPhotoColumnValue
            external_id = sheet['D' + str(row)].value
            if external_id:
                try:
                    external_id = str(external_id)
                except Exception:
                    external_id = None
            if external_id and len(external_id) > 200:
                external_id = external_id[0:200]
                msg = "Truncating external_id for {0}".format(parent_id)
                self.logger.warning(msg)
            label = sheet['E' + str(row)].value
            doc_description = sheet['F' + str(row)].value
            if not aslink and not os.path.exists(fileNameToUpload):
                msg += ";File [{0}] not found]".format(fileNameToUpload)
                self.logger.error(msg)
            else:
                self.logger.info(
                    'Processing {2}/{3}: Id: [{0}]; '
                    'File name: [{1}]; External ID: [{4}]'.format(
                        parent_id, fileNameToUpload, row, sheet.max_row
                        , external_id))
                result = self._process_upload_row(fileNameToUpload, parent_id
                                                  , groupid, categoryid
                                                  , subcategoryid, aslink
                                                  , isKeyPhoto, external_id
                                                  , label, doc_description
                                                  , parent_type)
                msg += ";" + result[0]
                doc_id = result[1]

            msg = msg.strip(";")
            self._log_message_to_sheet(sheet, row, doc_id, msg)

        wb.save(excelDataFileToRead)

    def _process_upload_row(self, fileNameToUpload, parent_id, groupid=None
                            ,categoryid=None,subcategoryid=None,aslink=False
                            , isKeyPhoto=False,external_id=None, label=None
                            , doc_description = None
                            , parent_type="ComplexAsset"):
        validImageExtensions = ["jpg", "png", "gif"]

        if aslink:
            # make sure there is a label for the link, else return
            if not label or label.strip() == "":
                return "Document Link must specify a label", "N/A"

        # prepare docuent object
        document = DocumentRepresentation()
        document.document_group = groupid
        document.parent_type = parent_type
        document.parent_identifier = parent_id
        document.document_category = categoryid
        document.external_id = external_id
        document.description = doc_description
        if categoryid:
            document.document_sub_category = \
                subcategoryid

        if not aslink:
            #its a file to upload
            if os.path.splitext(fileNameToUpload.lower())[1][1:] \
                    in validImageExtensions:
                if groupid is None:
                    groupid = 7
            else:
                isKeyPhoto = False
                if groupid is None:
                    groupid = 10
            fileToUploadPath = os.path.abspath(fileNameToUpload)
            if not label:
                fileToUploadName = os.path.basename(fileNameToUpload)
            else:
                fileToUploadName = label
            fileExtention = os.path.splitext(
                fileToUploadPath)[1][1:]
            filesize = os.path.getsize(fileNameToUpload)

            document.label = fileToUploadName
            document.document_size = filesize
            document.mime_type = fileExtention
            document.is_key_photo = isKeyPhoto

            # uploading document, so get it..
            with open(fileToUploadPath, "rb") as f:
                # filename = f.name
                data = f.read()

                if sys.version_info < (3, 0):
                    filecontents = data.encode("base64")
                else:
                    filecontents = base64.b64encode(data)
                    filecontents = filecontents.decode(
                        encoding="utf-8", errors="strict")
            file_properties = FilePropertiesRepresentation()
            file_properties.name = fileToUploadName
            file_properties.file_size = filesize
            file_properties.mimetype = fileExtention
            file_properties.filecontent = filecontents
            filearray = [file_properties]
            document.file_property = filearray
        else:
            # is a link
            document.document_link = fileNameToUpload
            document.label = label
            document.is_key_photo = False

        if parent_type == "WorkOrder2":
            # work around API issue.  Need to use work order GUID
            kwargs = {
                'request_params_page': 1
                , 'request_params_page_size': 1
                , 'request_params_filters': "FriendlyIdStr~eq~'{0}'".format(
                    parent_id
                )
            }
            try:
                wo = self.work_order_api.work_order_integration_api_get(
                    **kwargs)
            except ApiException as e:
                msg = "Error attempting to get work order GUID for friendly" \
                      " Id {0}.  Error status {1}, Reason: {2} {3}".format(
                    parent_id, e.status, e.body, e.reason)
                return msg, "N/A"
            else:
                if wo["TotalResults"] == 1:
                    document.parent_id = wo["ResourceList"][0]["Id"]
                    document.parent_identifier = None
                else:
                    msg = "Work Order '{0}' not found".format(parent_id)
                    return msg, "N/A"
        try:
            # create the document
            doc = self.docapi.document_post(document)
        except ApiException as e:
            msg = 'Error occurred while uploading document/photo: [{0}]; '\
                  'Error status {1}, Reason: {2} {3}\n'.format(
                fileNameToUpload, e.status, e.body, e.reason)
            self.logger.error(msg)
            return msg, "N/A"
        else:
            docid = doc[0].get('Id')
            return "SUCCESS", docid

    def _log_message_to_sheet(self, sheet, row, docid, status):
        """
        Write the result message for the upload of the particular document
        :param sheet: the spreadsheet object
        :param row: the row in the spreadsheet
        :param docid: the document ID if successfully created
        :param status: "SUCCESS" or error message
        """
        sheet['G' + str(row)] = docid
        sheet['H' + str(row)] = time.strftime("%c")
        sheet['I' + str(row)] = status

    def dataexchange_upload(self, excelDataFileToRead):
        """Request data-exchange(Bulk data import) requests to SaaS application.

        Usage:
            Invoke the script: python.exe -c "import assetic; uploadapi = assetic.BulkProcesses('C:/Users/xyz/assetic.ini'); uploadapi.dataexchange_upload('C:/Users/xyz/data.xlsx')"
        Args:
            excelDataFileToRead: XLSX/XLSM/XLTX/XLTM file containing data-exchange profile and data to import;
        Columns required are as follows (in EXACT ORDER)
        Data-exchange Profile Id | Document to Import (Complete file-path)
        Returns: Requested data-exchange task status is written back to excel file specified in excelDataFileToRead
        """
        validExcelFileExtensions = ["xlsx", "xlsm", "xltx",  "xltm"]
        validDocumentToImportExtensions = ["csv"]
        validImageExtensions = ["jpg", "png", "gif"]
        if excelDataFileToRead is not None and not excelDataFileToRead.isspace() and os.path.exists(excelDataFileToRead) and excelDataFileToRead.lower().split('.')[1] in validExcelFileExtensions:
            try:
                wb = openpyxl.load_workbook(excelDataFileToRead)
                sheet = wb.active
                sheet['C' + str(1)] = 'Imported Document Id'
                sheet['D' + str(1)] = 'Data-exchange Job Id'
                sheet['E' + str(1)] = 'Operation Performed On'
                sheet['F' + str(1)] = 'Status'
                wb.save(excelDataFileToRead)
            except Exception as e:
                self.logger.error('{0}\n'.format(e))
            else:
                totalItems = sheet.max_row - 1
                for row in range(2, sheet.max_row + 1):
                    itemCounter = row - 1
                    dataExchangeProfileId = sheet['A' + str(row)].value
                    documentFilePathToImport = sheet['B' + str(row)].value

                    if documentFilePathToImport is not None and \
                    not documentFilePathToImport.isspace() and \
                    os.path.splitext(documentFilePathToImport.lower()
                    )[1].strip(".") in validDocumentToImportExtensions and \
                    os.path.exists(documentFilePathToImport):
                        with open(documentFilePathToImport, "rb") as f:
                            data = f.read()

                        if sys.version_info < (3, 0):
                            filecontents = data.encode("base64")
                        else:
                            filecontents = base64.b64encode(data)
                            filecontents = filecontents.decode(
                                encoding="utf-8", errors="strict")

                        filename, fileExtention = os.path.splitext(
                            documentFilePathToImport)  # [1][1:]
                        fileExtention = fileExtention[1:]
                        filesize = os.path.getsize(documentFilePathToImport)

                        file_properties = FilePropertiesRepresentation()
                        file_properties.name = filename
                        file_properties.file_size = filesize
                        file_properties.mimetype = fileExtention
                        file_properties.filecontent = filecontents
                        filearray = [file_properties]

                        try:
                            document = DocumentRepresentation()
                            document.label = 'Importing data-exchange document - [' + filename + '] on ' + time.strftime("%d/%m/%Y %H:%M:%S")
                            document.file_property = filearray
                            document.document_size = filesize
                            document.mime_type = fileExtention

                            doc = self.docapi.document_post(document)
                        except ApiException as e:
                            self.logger.error('[{3}/{4}]: Error occurred while uploading data-exchange document: [{0}]; Error status {1}, Reason: {2}\n'.format(filename, e.status, e.reason, itemCounter, totalItems))
                        else:
                            docid = doc[0].get('Id')
                            self.logger.info('[{3}/{4}]: Processed document: [{0}] upload request for data-exchange profile: {1}; Response: [{2}]\n'.format(filename, dataExchangeProfileId, docid, itemCounter, totalItems))

                            isDocumentUploadSuccessful = True if docid is not None and not docid.isspace() else False
                            if isDocumentUploadSuccessful:
                                try:
                                    job = DataExchangeJobRepresentation()
                                    job.profile_id = dataExchangeProfileId
                                    job.document_id = docid
                                    task = self.dataexchangeapi.data_exchange_job_post(
                                        job)
                                except ApiException as e:
                                    sheet['C' + str(row)] = docid
                                    sheet['D' + str(row)] = 'N/A'
                                    sheet['E' + str(row)] = time.strftime("%c")
                                    sheet['F' + str(row)] = 'Error; ' + \
                                        str(e.status) + ': ' + e.reason
                                    self.logger.error('[{3}/{4}]: Error occurred while requesting data-exchange task: [{0}]; Error status {1}, Reason: {2}\n'.format(dataExchangeProfileId, e.status, e.reason, itemCounter, totalItems))
                                else:
                                    self.logger.debug('[{3}/{4}]: Requested Data-exchange task: {0} for profile: {1}/Document: {2}\n'.format(task, dataExchangeProfileId, filename, itemCounter, totalItems))

                                    sheet['C' + str(row)] = docid
                                    sheet['D' + str(row)] = task
                                    sheet['E' + str(row)] = time.strftime("%c")
                                    sheet['F' + str(row)] = 'SUCCESS'
                            else:
                                sheet['C' + str(row)] = 'N/A'
                                sheet['D' + str(row)] = 'N/A'
                                sheet['E' + str(row)] = time.strftime("%c")
                                sheet['F' + str(row)] = 'Error; Please check integration application logs for details'
                    else:
                        self.logger.warning('[{2}/{3}]: Document to import: [{1}] for data-exchange profile: [{0}] is either non-existent or of not CSV format\n'.format(
                            documentFilePathToImport, dataExchangeProfileId, itemCounter, totalItems))
                        sheet['C' + str(row)] = 'N/A'
                        sheet['D' + str(row)] = 'N/A'
                        sheet['E' + str(row)] = time.strftime("%c")
                        sheet['F' + str(row)] = 'Document to import file is either non-existent or not of CSV format'

                    wb.save(excelDataFileToRead)
        else:
            self.logger.error('File: [{0}] is either not specified or not of correct format (Supported formats are: .xlsx,.xlsm,.xltx,.xltm)\n'.format(
                excelDataFileToRead))

    def dataexchange_get_error_summary(self, excelDataFileToRead, errorFilesStorageDirectory):
        """Check status of data-exchange task(s) requested in [dataexchange_upload] function.

        Usage:
            Invoke the script: python.exe -c "import assetic; uploadapi = assetic.BulkProcesses('C:/Users/xyz/assetic.ini'); uploadapi.dataexchange_get_error_summary('C:/Users/xyz/data.xlsx', 'C:/Users/xyz/error_files/')"
        Args:
            excelDataFileToRead: XLSX/XLSM/XLTX/XLTM file containing data-exchange profile, document and task id;
                                 Columns required are as follows (in EXACT ORDER)
                                 Data-exchange ProfileId | Document to Import (Complete file-path) | Imported Document Id | Data-exchange Job Id | Operation Performed On | Status
            errorFilesStorageDirectory: Directory location where imported documents are stored

        Returns: Status of data-exchange tasks is written back to excel file specified in excelDataFileToRead and relvant imported document is saved in errorFilesStorageDirectory directory
        """
        validDocumentToImportExtensions = ["xlsx", "xlsm", "xltx",  "xltm"]
        validImageExtensions = ["jpg", "png", "gif"]
        if excelDataFileToRead is not None and not \
           excelDataFileToRead.isspace() and \
           os.path.exists(excelDataFileToRead) and \
           os.path.splitext(excelDataFileToRead.lower()
                            )[1].strip(".") in \
           validDocumentToImportExtensions:
            try:
                wb = openpyxl.load_workbook(excelDataFileToRead)
                sheet = wb.active
                sheet['G' + str(1)] = 'Error Document Id'
                sheet['H' + str(1)] = 'Summary'
                wb.save(excelDataFileToRead)
            except Exception as e:
                self.logger.error('{0}\n'.format(e))
            else:
                is_all_complete = True
                totalItems = sheet.max_row - 1
                for row in range(2, sheet.max_row + 1):
                    itemCounter = row - 1
                    importedDocumentId = sheet['C' + str(row)].value
                    dataExchangeTaskId = sheet['D' + str(row)].value

                    if dataExchangeTaskId is not None and not \
                       dataExchangeTaskId.isspace() and \
                       dataExchangeTaskId != 'N/A':
                        try:
                            task = self.dataExchangeTaskapi.data_exchange_task_get(dataExchangeTaskId)
                        except ApiException as e:
                            sheet['G' + str(row)] = e.status
                            sheet['H' + str(row)] = e.reason
                            self.logger.error('[{3}]/[{4}]: Error occurred while fetching status of data-exchange task:[{0}]; Error status {1}, Reason: {2}\n'.format(dataExchangeTaskId, e.status, e.reason, itemCounter, totalItems))
                        else:
                            status = task.get('Status')
                            statusDescription = task.get('StatusDescription')
                            self.logger.info('[{3}]/[{4}]: [{0}]: Data Exchange Job has status {1} [{2}]'.format(dataExchangeTaskId, status, statusDescription, itemCounter, totalItems))
                            # Complete with Error
                            if status == int('4'):
                                errorDocumentId = task.get('ErrorDocumentId')
                                if errorDocumentId is not None and not errorDocumentId.isspace():
                                    self.logger.info('[{3}/{4}]: Fetching error document: [{0}] for data-exchange task: [{2}] into [{1}]'.format(errorDocumentId, errorFilesStorageDirectory, dataExchangeTaskId, itemCounter, totalItems))
                                    self.document_get(errorFilesStorageDirectory, errorDocumentId)
                                    sheet['G' + str(row)] = errorDocumentId

                                try:
                                    sheet['H' + str(row)] = '[COMPLETED WITH ERROR]: ' + task.get('Summary')
                                except:
                                    sheet['H' + str(row)] = '[COMPLETED WITH ERROR]'

                            # Complete with No Error
                            elif status == int('3'):
                                sheet['G' + str(row)] = 'N/A'
                                try:
                                    sheet['H' + str(row)] = '[COMPLETED WITH NO ERROR]: ' + task.get('Summary')
                                except:
                                    sheet['H' + str(row)] = '[COMPLETED WITH NO ERROR]'

                            else:
                                is_all_complete = False
                                self.logger.info('[{3}/{4}]: Data-exchange task: [{0}] is in status: [{1}/{2}]\n'.format(dataExchangeTaskId, status, statusDescription, itemCounter, totalItems))
                                sheet['G' + str(row)] = ''
                                try:
                                    sheet['H' + str(row)] = '[{0}]: {1}'.format(statusDescription, task.get('Summary'))
                                except:
                                    self.logger.error('[{0}/{1}]: Error getting Status data. Please try again later - the upload job is likely still running.', itemCounter, totalItems)
                    else:
                        sheet['G' + str(row)] = 'N/A'
                        sheet['H' + str(row)] = 'Data-exchange task id is null, empty or not valid'

                    wb.save(excelDataFileToRead)

                return is_all_complete
        else:
            self.logger.error(
                'File: [{0}] is either not specified or not of correct format(Supported formats are: .xlsx,.xlsm,.xltx,.xltm)\n'.format(excelDataFileToRead))

    def document_get(self, fileStorageDirectoryLocation, docid):
        """Download specified document into directory.

        Usage:
            Invoke the script: python.exe -c "import assetic; uploadapi = assetic.BulkProcesses('C:/Users/xyz/assetic.ini'); uploadapi.document_get('C:/Users/xyz/', '1234-34566-23456-2334324')"
        Args:
            fileStorageDirectoryLocation: Directory location to save specified document
            docid: Document Id of document to be downloaded
        """
        if fileStorageDirectoryLocation is not None and not \
           fileStorageDirectoryLocation.isspace() and \
           os.path.exists(fileStorageDirectoryLocation) \
           and docid is not None and not docid.isspace():
            # Stop downloading a file if there's already one with the Document
            # ID stored
            for file in os.listdir(fileStorageDirectoryLocation):
                if docid in file:
                    ('A document with Document ID [{0}] already exists. Ignoring file download.'.format(docid))
                    return False
            getfile = None
            try:
                getfile = \
                    self.docapi.document_get_document_file_with_http_info(docid)
            except ApiException as e:
                self.logger.error('Error occurred while fetching Document:[{0}]; Error status {1}, Reason: {2}\n'.format(docid, e.status, e.reason))
            else:
                # getfile = self.docapi_get.document_get_document_file(docid)
                if getfile is not None and len(getfile) == 3:
                    if "Content-Disposition" in getfile[2]:
                        # get document name from response.
                        # The response is a tuple
                        if "attachment" in getfile[2]["Content-Disposition"] \
                                and "filename=" in \
                                getfile[2]["Content-Disposition"]:
                            filename = getfile[2]["Content-Disposition"].split(
                                "filename=", 1)[1]
                        else:
                            filename = docid
                        if '"' in filename or "'" in filename:
                            filename = filename[1:-1]
                        fullfilename = os.path.join(
                            fileStorageDirectoryLocation, filename)
                    else:
                        fullfilename = os.path.join(
                            fileStorageDirectoryLocation, docid)

                    data = getfile[0]
                    if sys.version_info >= (3, 0):
                        try:
                            data = data.decode('utf-8')
                        except:
                            data = data  # pass

                    # if file with name already exists then add document id at
                    # the end to make it unique
                    if fullfilename is not None and not fullfilename.isspace():
                        fileNameToModify, fileExtention = os.path.splitext(
                            fullfilename)
                        fullfilename = fileNameToModify + '_' + \
                            docid + '.' + fileExtention[1:]

                        # Need to handle binary and text files (e.g. CSV error
                        # files if Data Exchange jobs fail)
                        try:
                            with open(fullfilename, 'wb') as out_file:
                                out_file.write(data)
                                ('Created file: [{0}]\n'.format(fullfilename))
                        except:
                            try:
                                with open(fullfilename, 'w') as out_file:
                                    out_file.write(data)
                                    self.logger.debug('Created file: [{0}]\n'.format(fullfilename))
                            except:
                                self.logger.error('Failed getting file [{0}] in binary and plain text modes!'.format(fullfilename))
                else:
                    self.logger.warning('No document fetch response is received')
        else:
            self.logger.warning('Either File storage directory location or document id is not specified or is in incorrect format')

    def get_doc_group_id_prompt(self):
        """
        return a document group id having prompted user with options
        :returns: 0 if error or not selected, else group id
        """
        try:
            docgroups = self.docapi.document_get_document_groups()
        except ApiException as e:
            self.logger.error("Status {0}, Reason: {1} {2}".format(
                e.status,e.reason,e.body))
            return 0
        grouplist = list()
        print("\n*************************************")
        print("Document Groups")
        for group in docgroups:
            print("{0}. {1}".format(group["Id"],group["Label"]))
            grouplist.append(str(group["Id"]))
        print("Default Groups are 'Photos' and 'Documents' based on file"
              "type")
        msg = "Choose document group by entering the option number: "
        groupid = self.user_input(msg)

        if groupid == None or groupid.strip(" ")== "" or \
           groupid.isdigit() == False or groupid not in grouplist:
            return 0
        return groupid

    def get_parent_type_prompt(self):
        """
        return a parent record type having prompted user with options
        :returns: None if error or not selected, else parent type
        """
        print("\n*************************************")
        print("Parent Record Type")

        # Build static list of supported types
        types = collections.OrderedDict()
        types["Asset"] = "ComplexAsset"
        types["Functional Location"] = "GroupAssets"
        types["Assessment Result"] = "AsmtFormResult"
        types["Assessment Task"] = "AsmtTask"
        types["Assessment Project"] = "AsmtProject"
        types["Level 2 Assessment"] = "L2Assessment"
        types["Work Request"] = "WorkRequest"
        types["Work Order"] = "WorkOrder2"

        cnt = 0
        type_map = dict()
        for k in types:
            cnt += 1
            print("{0}. {1}".format(str(cnt), k))
            type_map[str(cnt)] = types[k]

        msg = "Choose parent record type by entering the option number\n"\
              "or hit return for the default (Asset): "
        type_id = self.user_input(msg)
        if not type_id or type_id.strip(" ") == "" or \
                not type_id.isdigit() or type_id not in type_map:
            return "ComplexAsset"
        elif type_id in type_map:
            return type_map[type_id]
        return "ComplexAsset"

    def get_doc_category_id_prompt(self):
        """
        return a document category id having prompted user with options
        :returns: 0 if error or not selected, else group id
        """
        try:
            doccategorys = self.docapi.document_get_document_categories()
        except ApiException as e:
            self.logger.error("Status {0}, Reason: {1} {2}".format(
                e.status,e.reason,e.body))
            return 0
        print("\n*************************************")
        print("Document Categories")
        categorylist = list()
        for category in doccategorys:
            print("{0}. {1}".format(category["Id"],category["Label"]))
            categorylist.append(str(category["Id"]))
        print("Document category is optional")
        msg = "Choose document category by entering the option number\n"\
              "or hit return for no category: "
        categoryid = self.user_input(msg)

        if categoryid == None or categoryid.strip(" ")== "" or \
        categoryid.isdigit() == False or categoryid not in categorylist:
            return 0
        return categoryid

    def get_doc_subcategory_id_prompt(self,categoryid):
        """
        return a document sub category id having prompted user with options
        :param categoryid: the category ID to get the subcategory for
        :returns: 0 if error or not selected, else subcategory id
        """
        try:
            docsubcategorys = self.docapi.document_get_document_sub_category(
                                categoryid)
        except ApiException as e:
            self.logger.error("Status {0}, Reason: {1} {2}".format(
                e.status,e.reason,e.body))
            return 0
        print("\n*************************************")
        print("Document Sub Categories")

        subcatlist = list()        
        for subcategory in docsubcategorys:
            print("{0}. {1}".format(subcategory["Id"],subcategory["Label"]))
            subcatlist.append(str(subcategory["Id"]))
        msg = "Choose document subcategory by entering the option number\n"\
              "or hit return if no subcategory: "
        subcategoryid = self.user_input(msg)

        if subcategoryid == None or subcategoryid.strip(" ")== "" or \
        subcategoryid.isdigit() == False or subcategoryid not in subcatlist:
            return 0
        return subcategoryid
    
    def user_input(self,message):
        """
        prompt user for input.
        :param message: message to prompt user with
        :returns: user input string
        """
        if not message or message.strip(" ") == "":
            return ""
        return six.moves.input(message)
        
    def bulk_upload_prompter(self):
        """
        prompt user for the required information for the various bulk uploads.
        :returns: success =0 else error number
        """
        self.logger.debug("Initiating Bulk Upload Prompter")
        while True:
            print("************************************************")
            print("Upload Options")
            print("1. Bulk upload Documents")
            print("2. Data Exchange Upload using saved profile id")
            print("3. Data Exchange Upload Status Check")
            print("4. Exit")
            print("************************************************")
            print(" ")
            print("Logged into: {0}".format(self.host))
            print(" ")

            msg = "Choose which upload to perform by entering the option "\
                  "number: "
            option = self.user_input(msg)
            if option == "1":
                ##Bulk upload documents
                self.prompter_upload_documents()
            elif option == "2":
                self.prompter_upload_dataexchange_via_profile()
            elif option == "3":
                self.prompter_dataexchange_status_check()
            else:
                break
            msg = "Exit Bulk Upload Prompter - last selected option "\
                  " was [{0}]".format(option)
            self.logger.debug(msg)
                
    def prompter_upload_documents(self):
        """
        prompt user for the required information for bulk document upload.
        excecute upload
        :returns: success =0 else error number
        """
        print("Please make sure the Excel import file is not open")
        print(" ")
        print("Upload required XLSX/XLSM/XLTX/XLTM file containing ")
        print("document/photos data to import;")
        print("Columns required are as follows (in EXACT ORDER):")
        print(" Id, the ID of the asset, functional location etc.")
        print(" Document/Photos Complete File Path")
        print(" Is Key-photo (TRUE/FALSE)")
        print(" External ID")
        print(" Label (mandatory if importing URL Links)")
        print(" Description")
        print(" ")

        while True:
            msg = "Enter the name and location of the Excel file or return\n"\
                  "to choose different option: "
            excelfile = self.user_input(msg)
            if len(excelfile) > 0 and excelfile.strip("") != "":
                if os.path.isfile(excelfile) == False:
                    print("Excel file not found")
                else:
                    msg = "Enter 'U' or 'L' to upload files (U) or create links"\
                          "(L) default=U:"
                    uporlink = self.user_input(msg)
                    aslink = False
                    if uporlink.upper().strip(" ") == "L":
                        aslink = True
                    parent_type = self.get_parent_type_prompt()
                    groupid = self.get_doc_group_id_prompt()
                    doccategoryid = self.get_doc_category_id_prompt()
                    docsubcategoryid = None
                    if doccategoryid != 0:
                        docsubcategoryid = self.get_doc_subcategory_id_prompt(
                            doccategoryid)
                    print("Upload results written back to Excel file")
                    self.upload_document_photo(excelfile, groupid,
                        doccategoryid, docsubcategoryid, aslink, parent_type)
                    break
            else:
                break
        return 0

    def prompter_upload_dataexchange_via_profile(self):
        """
        prompt user for the required information for data exchange upload.
        excute upload
        :returns: success =0 else error number
        """
        print("Please make sure the Excel import file is not open")
        print(" ")
        print("Upload requires XLSX/XLSM/XLTX/XLTM file containing")
        print("data-exchange profile and data to import;")
        print("Columns required are as follows (in EXACT ORDER):")
        print(" Data-exchange Profile Id")
        print(" Document to Import (Complete file-path)")
        print(" ")
      
        while True:
            msg = "Enter the name and location of the Excel file or return\n"\
                  "to choose different option: "
            excelfile = self.user_input(msg)
            if len(excelfile) > 0 and excelfile.strip("") != "":
                if os.path.isfile(excelfile) == False:
                    print("Excel file not found")
                else:
                    msg = "Requested data-exchange job status is written back"\
                          "\nto Excel file: "
                    print(msg)
                    self.dataexchange_upload(excelfile)
                    break
            else:
                break
        return 0


    def prompter_dataexchange_status_check(self):
        """
        Check status of data-exchange task(s) requested in dataexchange_upload
        prompt user for the required information for bulk document upload.
        excecute upload
        :returns: success =0 else error number
        """
        msg = "Status of data-exchange jobs is written back to excel \n"\
              "file specified in excelDataFileToRead and relvant imported\n"\
              "document is saved in errorFilesStorageDirectory directory"
        print(msg)
        print("Please make sure the Excel import file is not open")
        print(" ")
        print("Upload requires : XLSX/XLSM/XLTX/XLTM file containing ")
        print("data-exchange profile, document and task id;")
        print("Columns required are as follows (in EXACT ORDER)")
        print(" Data-exchange ProfileId")
        print(" Document to Import (Complete file-path)")
        print(" Imported Document Id")
        print(" Data-exchange Job Id")
        print(" Operation Performed On")
        print(" Status")


        while True:
            msg = "Enter the name and location of the Excel file or return\n"\
                  "to choose different option: "
            excelfile = self.user_input(msg)
            if len(excelfile) > 0 and excelfile.strip(" ") != "":
                if os.path.isfile(excelfile) == False:
                    print("Excel file not found")
                else:
                    msg = "Enter directory location to write error file "\
                              "(if error), \nenter nothing to use same "\
                              "directory as import file: "
                    errorfiledir = self.user_input(msg)
                    if len(errorfiledir) == 0 or errorfiledir.strip(" ") == "":
                        errorfiledir = os.path.dirname(excelfile)
                    if os.path.isdir(errorfiledir) == False:
                            print("Directory not found")
                    else:
                        msg = "Requested data-exchange job status written back"\
                              "\nto Excel file: "
                        print(msg)
                        self.dataexchange_get_error_summary(
                            excelfile, errorfiledir)
                        break
            else:
                break
        return 0

