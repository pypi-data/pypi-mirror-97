"""Copyright 2016 Assetic."""

from __future__ import absolute_import
import os
import time
import sys
import base64
import openpyxl  # to read excel files
#import six
try:
    import configparser as configparser
except:
    # V2 python?
    import ConfigParser as configparser

# import assetic libraries we'll use
from assetic.api_client import ApiClient
from assetic.configuration import Configuration
from assetic.api import DocumentApi
from assetic.models.file_properties_representation import \
    FilePropertiesRepresentation
from assetic.models.document_representation import DocumentRepresentation
from assetic.rest import ApiException


class MigrateMyDataToSaaS(object):
    """Class to manage bulk processes especially for myData to SaaS application."""

    def __init__(self, configfile=None):
        """Initialize class with configuration file."""
        if configfile is None:
            configfile = os.path.join(os.getcwd(), 'assetic.ini')

        if configfile is not None and not configfile.isspace() and not os.path.isfile(configfile):
            message = "Configuration file not found.  Expecting assetic.ini in current folder, or set via module init"
            print(message)
            raise ValueError(message)
        else:
            self.configfile = configfile

##        if six.PY3 == True:
##            if sys.stdout.encoding != 'cp850':
##                sys.stdout = codecs.getwriter('cp850')(sys.stdout.buffer, 'strict')
##            if sys.stderr.encoding != 'cp850':
##                sys.stderr = codecs.getwriter('cp850')(sys.stderr.buffer, 'strict')
##        else:
##            if sys.stdout.encoding != 'cp850':
##                sys.stdout = codecs.getwriter('cp850')(sys.stdout, 'strict')
##            if sys.stderr.encoding != 'cp850':
##                sys.stderr = codecs.getwriter('cp850')(sys.stderr, 'strict')

    def bulk_upload_document_photo(self, excelDataFileToRead, myDataSharedDirectoryFilePath):
        """Automate document/photo upload in bulk fashion from myData shared directory to SaaS application.

        Usage:
            Invoke the script: python.exe -c "import assetic; uploadapi = assetic.MigrateMyDataToSaaS('C:/Users/xyz/assetic.ini'); uploadapi.bulk_upload_document_photo('C:/Users/xyz/data.xlsx', 'C:/Users/myDataSharedDirectory/')"
        Args:
            excelDataFileToRead: XLSX/XLSM/XLTX/XLTM file containing document/photos data to import;
                                 Columns required are as follows (in EXACT ORDER)
                                 Asset Id | Document/Photos File name to search in [myDataSharedDirectoryFilePath]  | Is Key-photo (TRUE/FALSE)
            myDataSharedDirectoryFilePath: Directory location where myData shared contents are stored eg. document,photos, profiles etc
        Returns: Uploaded document/photos status is written back to excel file specified in excelDataFileToRead
        """
        # Read ini
        settings = configparser.ConfigParser()
        configfile = self.configfile
        settings.read(configfile)
        # Apply config settings
        conf = Configuration()
        conf._default.host = settings.get('environment', 'url')
        conf._default.username = settings.get('auth', 'username')
        conf._default.password = settings.get('auth', 'api_key')
        auth = conf.get_basic_auth_token()

        client = ApiClient(conf._default.host, "Authorization", auth)
        docapi = DocumentApi(client)
        validDocumentToImportExtensions = ["xlsx", "xlsm", "xltx",  "xltm"]
        validImageExtensions = ["jpg", "png", "gif"]

        isValidToProceed = True
        if myDataSharedDirectoryFilePath is None or myDataSharedDirectoryFilePath.isspace():
            print('myData Shared directory path is either null or empty')
            isValidToProceed = False
        elif not os.path.isdir(myDataSharedDirectoryFilePath):
            print('myData Shared directory: [{0}] is non existing'.format(myDataSharedDirectoryFilePath))
            isValidToProceed = False

        if isValidToProceed:
            if excelDataFileToRead is not None and not excelDataFileToRead.isspace() and os.path.exists(excelDataFileToRead) and excelDataFileToRead.lower().split('.')[1] in validDocumentToImportExtensions:
                print('File for bulk document upload operation is specified: {0}\n'.format(
                    excelDataFileToRead))
                try:
                    wb = openpyxl.load_workbook(excelDataFileToRead)
                    sheet = wb.active
                    sheet['D' + str(1)] = 'Migrated Document Id'
                    sheet['E' + str(1)] = 'Operation Performed On'
                    sheet['F' + str(1)] = 'Status'
                    wb.save(excelDataFileToRead)
                except Exception as e:
                    print('{0}\n'.format(e))
                else:
                    for row in range(2, sheet.max_row + 1):
                        assetId = sheet['A' + str(row)].value
                        fileNameToUpload = sheet['B' + str(row)].value
                        isKeyPhotoColumnValue = sheet['C' + str(row)].value
                        isKeyPhoto = False
                        if isKeyPhotoColumnValue:
                            isKeyPhoto = isKeyPhotoColumnValue

                        if fileNameToUpload is None or fileNameToUpload.isspace():
                            print('Processing {0}/{1}: No File name is specified for processing'.format(row, sheet.max_row))
                            continue

                        print('Processing {2}/{3}: Complex Asset Id: [{0}]; File name: [{1}]'.format(assetId, fileNameToUpload, row, sheet.max_row))
                        myDataSharedDirectoryDocumentsFilePath = os.path.join(myDataSharedDirectoryFilePath, 'Documents')
                        for path, subdirs, files in os.walk(myDataSharedDirectoryDocumentsFilePath):
                            for fileItem in files:
                                fullPath = os.path.join(path, fileItem)
                                fetchCorrectFile = True if fileItem == fileNameToUpload else False
                                if (fetchCorrectFile):
                                    fileToUploadPath = fullPath
                                    break

                        if fileToUploadPath is None or fileToUploadPath.isspace():
                            myDataSharedDirectoryPhotosFilePath = os.path.join(myDataSharedDirectoryFilePath, 'Photos')
                            for path, subdirs, files in os.walk(myDataSharedDirectoryPhotosFilePath):
                                for fileItem in files:
                                    fullPath = os.path.join(path, fileItem)
                                    fetchCorrectFile = True if fileItem == fileNameToUpload else False
                                    if (fetchCorrectFile):
                                        fileToUploadPath = fullPath
                                        break

                    if (fileToUploadPath is not None and not fileToUploadPath.isspace() and os.path.exists(fileToUploadPath)):
                        fileToUploadPath = os.path.abspath(fileToUploadPath)
                        fileToUploadName = os.path.basename(fileToUploadPath)
                        fileExtention = os.path.splitext(
                            fileToUploadPath)[1][1:]
                        with open(fileToUploadPath, "rb") as f:
                            filename = f.name
                            data = f.read()

                            if sys.version_info < (3, 0):
                                filecontents = data.encode("base64")
                            else:
                                filecontents = base64.b64encode(data)
                                filecontents = filecontents.decode(encoding="utf-8", errors="strict")

                            filesize = os.path.getsize(filename)

                            file_properties = FilePropertiesRepresentation()
                            file_properties.name = fileToUploadName
                            file_properties.file_size = filesize
                            file_properties.mimetype = fileExtention
                            file_properties.filecontent = filecontents
                            filearray = [file_properties]

                            if fileToUploadName.lower().split('.')[1] in validImageExtensions:
                                if isKeyPhoto:
                                    isKeyPhoto = True
                                else:
                                    isKeyPhoto = False
                                documentGroupId = 7
                            else:
                                isKeyPhoto = False
                                documentGroupId = 10
							
                            docid = ' '
                            errorMessage = ' '
                            try:
                                document = DocumentRepresentation()
                                document.is_key_photo = isKeyPhoto
                                document.document_group = documentGroupId
                                document.parent_type = 'ComplexAsset'
                                document.parent_identifier = str(assetId)
                                document.label = 'Importing complex asset document/photo - [' + str(fileToUploadName) + '] on ' + time.strftime("%d/%m/%Y %H:%M:%S")
                                document.file_property = filearray
                                document.document_size = filesize
                                document.mime_type = fileExtention

                                doc = docapi.document_post(document)
                            except ApiException as e:
                                print('Error occurred while uploading document/photo: [{0}]; Error status {1}, Reason: {2}\n'.format(fileToUploadName, e.status, e.reason))
                                errorMessage = 'Error: Status '+ str(e.status) + '; Reason: '+ str(e.reason) + '; Please check integration application logs for actual error details'
                            except Exception as e:                            
                                print('Error occurred while uploading document/photo: [{0}]; Error status {1}, Reason: {2}\n'.format(filename, e.status, e.reason))								
                                errorMessage = 'Error: Status '+ str(e.status) + '; Reason: '+ str(e.reason) + '; Please check integration application logs for actual error details'
                            else:
                                docid = doc[0].get('Id')
                                print('UPLOADED document/photo: [{1}]; Response: [{0}]\n'.format(docid, fileToUploadName))

                            isDocumentUploadSuccessful = True if docid is not None and not doci.isspace() else False
                            errorMessage = errorMessage if errorMessage is not None and not errorMessage.isspace() else 'Error; Please check integration application logs for details'							
                            if isDocumentUploadSuccessful:
                                sheet['D' + str(row)] = docid
                                sheet['E' + str(row)] = time.strftime("%c")
                                sheet['F' + str(row)] = 'SUCCESS'
                            else:
                                sheet['D' + str(row)] = 'N/A'
                                sheet['E' + str(row)] = time.strftime("%c")
                                sheet['F' + str(row)] = errorMessage
                    else:
                        print('File: [{0}] is non-existing\n'.format(fileToUploadPath))
                        sheet['C' + str(row)] = 'N/A'
                        sheet['D' + str(row)] = 'N/A'
                        sheet['E' + str(row)] = time.strftime("%c")
                        sheet['F' + str(row)] = 'Document/Photos to import is non-existing'

                    wb.save(excelDataFileToRead)
        else:
            print('File: [{0}] is either not specified or not of correct format (Supported formats are: .xlsx,.xlsm,.xltx,.xltm)\n'.format(
                excelDataFileToRead))

    def bulk_upload_keyphotos(self, keyPhotosStorageLocation, debug=False):
        """Automate uploads and assignment of key-photos(s) to SaaS application.

        Usage:
            Invoke the script: python.exe -c "import assetic; uploadapi = assetic.MigrateMyDataToSaaS('C:/Users/xyz/assetic.ini'); uploadapi.bulk_upload_keyphotos('C:/Users/myDataSharedDirectory/KeyPhotos/')"
        Args:
            keyPhotosStorageLocation: Directory location where all photo images are stored having naming convention [{SaaS AssetID}_keyphoto.jpg]
        Returns: Uploaded document/photos status is written back to excel file specified in excelDataFileToRead
        """
        # Read ini
        settings = configparser.ConfigParser()
        configfile = self.configfile
        settings.read(configfile)
        # Apply config settings
        conf = Configuration()
        conf._default.host = settings.get('environment', 'url')
        conf._default.username = settings.get('auth', 'username')
        conf._default.password = settings.get('auth', 'api_key')
        auth = conf.get_basic_auth_token()
        # Debug settings
        conf._default.debug = debug

        client = ApiClient(conf._default.host, "Authorization", auth)
        docapi = DocumentApi(client)
        validImageExtensions = ["jpg", "png", "gif"]

        isValidToProceed = True
        if keyPhotosStorageLocation is None or keyPhotosStorageLocation.isspace():
            print('Key-photo directory location is either null or empty')
            isValidToProceed = False
        elif not os.path.isdir(keyPhotosStorageLocation):
            print('Key-photo directory location: [{0}] is non existing'.format(keyPhotosStorageLocation))
            isValidToProceed = False

        if isValidToProceed:
            print('Key-photo storage location specified: {0}'.format(keyPhotosStorageLocation))
            # Key-photo file naming format: <SaaS AssetID>_keyphoto.jpg            
            curDir = os.path.abspath(os.path.join(keyPhotosStorageLocation, os.pardir))
            parentDir = os.path.abspath(os.path.join(curDir,os.pardir))            
            excelSaveFilePath = os.path.join(curDir, 'KeyPhotosMigrationStatus.xlsx')
##            print('Current: {0}\nParent: {1}\nJoined Path: {2}'.format(curDir, parentDir, excelSaveFilePath))
            keyPhotosFiles = os.listdir(keyPhotosStorageLocation)
            try:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "Key-photo Migration status"
                sheet['A' + str(1)] = 'Operation Performed On'
                sheet['B' + str(1)] = 'File Name'
                sheet['C' + str(1)] = 'Migrated Document Id'
                sheet['D' + str(1)] = 'Status'

                wb.save(excelSaveFilePath)
            except Exception as e:
                print('{0}\n'.format(e))
            else:
                noOfKeyPhotosToUpload = len([keyPhotoFileItem for keyPhotoFileItem in keyPhotosFiles if (keyPhotoFileItem.lower().split('.')[1] in validImageExtensions)])
                itemCounter = 0
                for keyPhotoFile in keyPhotosFiles:
                    fullPath = os.path.join(keyPhotosStorageLocation, keyPhotoFile)
                    splitedFileName, fileExtention = os.path.splitext(keyPhotoFile)  # [1][1:]
                    fileExtention = fileExtention[1:]
                    assetId = keyPhotoFile.rsplit('_keyphoto', 1)[0]

                    if keyPhotoFile.lower().rsplit('.', 1)[1] in validImageExtensions:
                        itemCounter = itemCounter + 1
                        if (os.path.exists(fullPath)):
                            with open(fullPath, "rb") as f:
                                filename = os.path.basename(f.name)
                                data = f.read()

                            if sys.version_info < (3, 0):
                                filecontents = data.encode("base64")
                            else:
                                filecontents = base64.b64encode(data)
                                filecontents = filecontents.decode(encoding="utf-8", errors="strict")

                            filesize = os.path.getsize(fullPath)

                            file_properties = FilePropertiesRepresentation()
                            file_properties.name = filename
                            file_properties.file_size = filesize
                            file_properties.mimetype = fileExtention
                            file_properties.filecontent = filecontents
                            filearray = [file_properties]

                            docid = ' '
                            errorMessage = ' '
                            try:
                                document = DocumentRepresentation()
                                document.is_key_photo = True
                                document.document_group = 7  # Photos items Id from DOCCBGroup table
                                document.parent_type = 'ComplexAsset'
                                document.parent_identifier = str(assetId)
                                document.label = 'Importing complex asset key-photo - [' + str(filename) + '] on ' + time.strftime("%d/%m/%Y %H:%M:%S")
                                document.file_property = filearray
                                document.document_size = filesize
                                document.mime_type = fileExtention

                                print('[{1}/{2}]: Processing: [{0}]'.format(str(filename), itemCounter, noOfKeyPhotosToUpload))
                                doc = docapi.document_post(document)
                            except ApiException as e:
                                print('[{3}/{4}]: Error occurred while uploading key-photos: [{0}]; Error status {1}, Reason: {2}\n'.format(str(filename), e.status, e.reason, itemCounter, noOfKeyPhotosToUpload))								
                                errorMessage = 'Error: Status '+ str(e.status) + '; Reason: '+ str(e.reason) + '; Please check integration application logs for actual error details'
                            except Exception as e:                            
                                print('[{3}/{4}]: Error occurred while uploading key-photos: [{0}]; Error status {1}, Reason: {2}\n'.format(str(filename), e.status, e.reason, itemCounter, noOfKeyPhotosToUpload))								
                                errorMessage = 'Error: Status '+ str(e.status) + '; Reason: '+ str(e.reason) + '; Please check integration application logs for actual error details'
                            else:
                                docid = doc[0].get('Id')
                                print('[{3}/{4}]: Uploaded file: [{1}] associated with assetId: [{2}]; Response: [{0}]\n'.format(docid, str(filename), assetId, itemCounter, noOfKeyPhotosToUpload))														
                            
                            isDocumentUploadSuccessful = True if docid is not None and not docid.isspace() else False
                            errorMessage = errorMessage if errorMessage is not None and not errorMessage.isspace() else 'Error; Please check integration application logs for details'
##                            print('docid: {0}; error message:{1}; Upload successful: {2}'.format(docid, errorMessage, isDocumentUploadSuccessful))
                            for row in range(sheet.max_row + 1, sheet.max_row + 2):
                                sheet['A' + str(row)] = time.strftime("%c")
                                sheet['B' + str(row)] = str(filename)
                                if isDocumentUploadSuccessful:
                                    sheet['C' + str(row)] = docid
                                    sheet['D' + str(row)] = 'SUCCESS'
                                else:
                                    sheet['C' + str(row)] = 'N/A'
                                    sheet['D' + str(row)] = errorMessage
                        else:
                            print('File: [{0}] is non-existing\n'.format(str(keyPhotoFile)))
                            for row in range(sheet.max_row + 1, sheet.max_row + 2):
                                sheet['A' + str(row)] = time.strftime("%c")
                                sheet['B' + str(row)] = str(keyPhotoFile)
                                sheet['C' + str(row)] = 'N/A'
                                sheet['D' + str(row)] = 'Photo file import file is non-existing'
                    else:
                        print('File: [{0}] is either not specified or not of correct format (Supported formats are: jpg, png, gif)\n'.format(str(keyPhotoFile)))
                        for row in range(sheet.max_row + 1, sheet.max_row + 2):
                            sheet['A' + str(row)] = time.strftime("%c")
                            sheet['B' + str(row)] = str(keyPhotoFile)
                            sheet['C' + str(row)] = 'N/A'
                            sheet['D' + str(row)] = 'Photo file import file is either non-existing or not of correct format (Supported formats are: jpg, png, gif)'

                    wb.save(excelSaveFilePath)
