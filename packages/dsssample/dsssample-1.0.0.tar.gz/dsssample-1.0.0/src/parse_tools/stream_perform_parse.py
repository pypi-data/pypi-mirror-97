#!/usr/bin/ python
# -*- coding: utf-8 -*-
########################################################################
#
# Copyright (c) 2020 Baidu.com, Inc. All Rights Reserved
#
########################################################################

"""
    File: stream_perform_parse.py
    Author: dushuangshuang(dushuangshuang@baidu.com)
    Date:   2020/08/24 下午14:50
"""
import os
import subprocess
import openpyxl
import json
import logging
import time
import signal
import copy
import chardet
import re
import difflib
from openpyxl import Workbook, load_workbook
import importlib, sys
importlib.reload(sys)

import globalVal
import db_manager
import xls
import speech_parse
import read_message

if globalVal._DEBUG_MODE:
    logging.basicConfig(level=logging.DEBUG)


def checkNetError(asr_list, first_flag=True):
    """检查当前错误码是否为网络引起的错误，用于统计单次识别字准、句准
    """
    if first_flag:
        if asr_list['asr_result'] is not None and str(asr_list['asr_result']) != '':
            return False
        if 'asrCancel' in asr_list and asr_list['asrCancel'] is not None:
            return True

    if asr_list['errCode'] in globalVal._NET_ERROR_CODE_LIST:
        return True
    else:
        return False


def getAverage(timeList, flags=1):
    """计算均值，flags=1时将负值也纳入计算，flags=0时仅统计正值
    """
    totalCount = 0
    totalIndex = 0
    for line in timeList:
        if flags or line > 0:
            totalCount += line
            totalIndex += 1
    if totalIndex > 0:
        return float(totalCount) / float(totalIndex)
    else:
        return 0


def getRoleColor(role):
    """计算角色对应写入excel的颜色
    """
    return int(role) + 20


def initWorkbook(audio_scene):
    """
        [表格初始化] 以测试场景区分，详细记录每个测试版本、每条query请求情况
    Args:
        audio_scene : [string] 测试集场景
    Returns:
        workBook : workbook
    """
    workbook_path = os.path.join(globalVal._WORK_DIRECTORY, audio_scene + '.xlsx')
    if workbook_path in globalVal._WORKBOOK_DICT:
        wb = globalVal._WORKBOOK_DICT[workbook_path]
    elif os.path.exists(workbook_path):
        wb = load_workbook(workbook_path)
        globalVal._WORKBOOK_DICT.update({workbook_path:wb})
    else:
        wb = Workbook()
        wb.remove(wb['Sheet'])
        globalVal._WORKBOOK_DICT.update({workbook_path:wb})
    return wb


class SpeechParseThread():
    """
        灌测性能数据对齐和分析
    """

    def __init__(self, argv_item):
        """
            参数初始化
        """
        # [0] 输入解析
        self.__sdkVersion = argv_item[0]
        self.__initSceneName = argv_item[1]
        if len(argv_item) == 6:
            vp_id = argv_item[5]
        else:
            vp_id = ''
        # [1] 读取测试音频的人工标注答案
        self.__wpStatusDict = {}            # [唤醒指标]唤醒指标:包括标注唤醒词、总条数，唤醒对齐次数、正确次数、词误串、音区误串、忽略次数
        self.__wpFalseDict = {}             # [唤醒指标]误唤醒指标:包括每个唤醒词的误报次数
        self.__asrQueryCount = 0            # [识别指标]标注的识别query个数（多轮:仅统计2~N次的情况）
        # [2] 读取灌测日志
        self.__totalWpCount = 0             # [唤醒指标]唤醒总次数（不区分唤醒词，包括正常唤醒和误唤醒）
        self.__totalFirstAsrCount = 0       # [错误码指标]首次识别总请求次数（asr ready次数）
        self.__totalMultiAsrCount = 0       # [错误码指标]2~N次识别总请求次数（有屏:end次数）
        self.__totalFirstAsrError = []      # [错误码指标]全部日志里的，首次识别错误码
        self.__totalMultiAsrError = []      # [错误码指标]全部日志里的，2~N次识别错误码
        # [3] 将灌测日志和标注答案，根据时间顺序进行对齐
        self.__wpDetectCount = 0            # [唤醒指标]唤醒请求对齐次数
        self.__asrDetectCount = 0           # [识别指标]识别请求对齐次数（单次:对齐的首次请求，多次:对齐的2~N次请求），未对齐的记为no wakeup、no asr
        self.__asr_reject_count = 0         # [拒识指标]多轮拒识总次数
        self.__ignore_asr = 0               # [小度指令词指标]识别忽略次数(唤醒正确的基础上，当前仍在识别状态中、且被“小度”唤醒，则此次识别请求被忽略，不会启动)
        self.__falseAsrCount = 0            # [误识别指标]（单次）误识别次数: 所有标注答案外asr_reject为0或为空的次数
        self.__falseMutiplyasrCount = 0     # [误识别指标]（多次）误识别次数: 所有标注答案外asr_reject为0的次数
        self.__falseMutiplyVadCount = 0     # [误识别指标]多轮误交互次数: 所有标注答案外请求的次数，不区分是否有识别结果（不区分拒识）或error
        self.__firstAsrErrorList = []       # [错误码指标]与标注对齐的，首次识别错误码
        self.__multiAsrErrorList = []       # [错误码指标]与标注对齐的，2~N次识别错误码
        # [4] query有效性过滤 & 识别成功率、召回率指标计算
        # [5] 速度及其他指标计算
        self.__vpTotalCount = 0             # [声纹指标]对齐的声纹认证总次数
        self.__vpRightCount = 0             # [声纹指标]正确认证次数（集内：判定为正确的人，集外：判定为不认识）
        self.__vpWrongPersonCount = 0       # [声纹指标]错误认证次数（集内：判断为集内其他人，集外：判定为集内某一人）
        self.__asrResponseTimeList = []     # [速度指标]识别首包响应
        self.__asrFinishTimeList = []       # [速度指标]硬延迟
        self.__asrChunkTtsTimeList = []     # [速度指标]TTS首包
        self.__asrVadBeginTimeList = []     # [速度指标]VAD起点检测时间
        self.__asrVadEndTimeList = []       # [速度指标]VAD尾点检测时间

        self.__audioScene, self.__audioName, self.__device_distance, self.__device_md5 = \
            globalVal.get_scene_info(self.__initSceneName)
        self.__samplesetName = self.__audioScene + '_' + self.__audioName
        self.__speech_type = globalVal.get_speech_type(self.__samplesetName)
        self.__workBook = initWorkbook(self.__audioScene)
        print(self.__sdkVersion, self.__samplesetName)

        self.__wp_mode = 0                  # 标志:唤醒统计类别（0:1个唤醒词；1:carlife；2:小度+指令词；3:其他）
        self.__mutiply_asr_mode = False     # 标志:是否多次识别
        self.__pure_asr_mode = False        # 标志:是否纯识别模式（无唤醒）
        self.__role_asr_mode = False        # 标志:是否双人交互模式（听清识别）
        self.__split_asr_mode = False       # 标志:是否是切分单条识别pcm
        self.__offline_asr_mode = False     # 标志:是否是离线识别
        self.__semantic_asr_mode = False    # 标志:是否语义识别模式

        self.__use_vad = False              # 标志:是否有vad
        self.__angle_in_range = False       # 标志:纯识别模式，是否在角度内（无唤醒）
        self.__support_info = ''            # 标志:多次识别是否参与统计（不参与统计的情况:未唤醒、误唤醒、唤醒后首次识别无识别结果或回调网络错误）
        self.__curAnswerQueryIndex = 0      # 标注答案的index
        self.__MutiplyasrWpAnswerIndex = 0  # 多轮下，唤醒标注的index
        self.__AnswerList = []              # 识别答案列表
        self.__logResultList = []           # 识别打点日志列表
        self.__queryResultDict = {}         # 识别详情记录dict --- 是否将识别的所有信息合入
        self.__wpLocationAnswerDict = {}    # 声源定位标注答案
        self.__wpLocationResultDict = {}    # 声源定位结果
        self.__vp_id = ''                   # 声纹注册期望的usr_id
        self.__vp_list = []                 # 声纹注册集内的usr_id列表
        self.__asr_status = {'acoustics_reject': 0, 'semantics_reject': 0, 'other_reject': 0, 'voiceprint_rejcet': 0}
        if globalVal._PARAM['AUDIO_FOR_SERVER_FLAG'] and not 'offline' in self.__audioScene:
            self.__record_for_sever = True
        else:
            self.__record_for_sever = False

        if self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
            if vp_id == '':
                logging.error('声纹性能统计，缺少vpID!')
                globalVal.record_err(argv_item[0:2], 1102, '声纹性能统计，缺少vpID!')
                return
            self.__vp_id = vp_id.split('&')[0]
            self.__vp_list = re.split('[()]', vp_id.split('&')[1])[1].split(',')
            if self.__vp_id == 'NULL' or self.__vp_id == 'null':
                self.__is_jinei = '集外'
            else:
                self.__is_jinei = '集内'

        self.__savePath = globalVal.get_save_path(self.__audioScene, self.__audioName)
        if not os.path.exists(self.__savePath):
            os.makedirs(self.__savePath)
        if 'asr' in self.__audioScene or self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
            self.__asr_answer_file_name = os.path.join(self.__savePath, self.__samplesetName + '_answer_result.txt')
            self.__rec_result_file_name = os.path.join(self.__savePath, \
                self.__sdkVersion + '_' + self.__samplesetName + '_recognize_result.txt')
            if self.__record_for_sever:
                self.__severInfoPath = os.path.join(globalVal._WORK_DIRECTORY, 'server')
                if not os.path.exists(self.__severInfoPath):
                    os.makedirs(self.__severInfoPath + '/info')
                self.__server_answer_file_name = os.path.join(self.__severInfoPath,
                    self.__sdkVersion + '_' + self.__audioScene + '.txt')
                self.__server_info_file_name = os.path.join(self.__severInfoPath, 'info',
                    self.__sdkVersion + '_' + self.__audioScene + '.txt')
                self.__server_answer_file = open(self.__server_answer_file_name, 'a+', encoding='utf-8')
                self.__server_info_file = open(self.__server_info_file_name, 'a+', encoding='utf-8')
        if self.__speech_type == globalVal.SpeechType.DCI:
            self.__dci_data_file_name = os.path.join(self.__savePath,
                self.__sdkVersion + '_' + self.__samplesetName + '_' + self.__device_distance + '_wp_dci.txt')
            self.__dci_data_file = open(self.__dci_data_file_name, 'w', encoding='utf-8')
        
        if self.__device_md5 is not None:
            self.__sdkVersion += '_' + self.__device_distance

        if self.__sdkVersion in self.__workBook.sheetnames:
            self.__worksheet = self.__workBook[self.__sdkVersion]
            self.__excelCount = self.__worksheet.max_row + 5
        else:
            self.__worksheet = self.__workBook.create_sheet(self.__sdkVersion)
            self.__excelCount = 1

    def read_messages(self, wp_answer_path, asr_answer_path, log_path):
        """
        读取标注答案，从log中筛选关键信息
        """
        #【0】获取测试类型：根据场景名中是否包含mutiplyasr，判断是一次唤醒、多次识别场景
        if self.__speech_type == globalVal.SpeechType.MULTI_ASR:
            self.__mutiply_asr_mode = True
        elif self.__speech_type == globalVal.SpeechType.PUER_ASR:
            self.__pure_asr_mode = True
            angle = int(self.__samplesetName.split('_')[4].split('-')[1].split('du')[0])
            if globalVal._PARAM['PUREASR_ANGLE_RANGE'][0] <= angle <= globalVal._PARAM['PUREASR_ANGLE_RANGE'][1]:
                self.__angle_in_range = True
        elif self.__speech_type == globalVal.SpeechType.ROLE_ASR:
            self.__pure_asr_mode = True
            self.__role_asr_mode = True
        elif self.__speech_type == globalVal.SpeechType.SEMANTIC_ASR:
            self.__semantic_asr_mode = True
            self.__offline_asr_mode = True
        elif self.__speech_type == globalVal.SpeechType.OFFLINE_ASR:
            self.__offline_asr_mode = True
        elif self.__speech_type == globalVal.SpeechType.SPLIT_ASR:
            self.__split_asr_mode = True
        #【1】读取标注答案
        answer_parse_thread = read_message.AnswerParseThread(
            self.__initSceneName, self.__savePath, wp_answer_path, asr_answer_path)
        self.__AnswerList, self.__wpStatusDict, self.__asrQueryCount = answer_parse_thread.readAnswerMessages()
        if self.__speech_type == globalVal.SpeechType.WP_LOCATION:
            self.__wpLocationAnswerDict = answer_parse_thread.getLocationAnswerList()
        elif self.__speech_type == globalVal.SpeechType.DCI:
            self.__dci_data_file.write('total_wakeup_count:' + str(len(self.__AnswerList)) \
                + '\tdevice_distance:' + str(self.__device_distance) + '\tdevice_md5:' + str(self.__device_md5) + '\n')
        #【2】读取SDK关键回调打点文件
        log_parse_thread = read_message.LogParseThread(log_path, self.__initSceneName)
        list_info, count_info, [self.__use_vad, self.__wp_mode] = log_parse_thread.readLogMessages()
        self.__logResultList, self.__totalFirstAsrError, self.__totalMultiAsrError = list_info
        self.__totalWpCount, self.__totalFirstAsrCount, self.__totalMultiAsrCount = count_info
        if self.__wp_mode == 0 and len(self.__wpStatusDict) > 2:
            self.__wp_mode = 2
        if self.__wp_mode == 1 or 'kws' in self.__samplesetName:
            globalVal._DISPLAY_BY_COL = True
        #【3】初始化excel
        title_list = copy.deepcopy(globalVal._DETAIL_TITILE_LIST)
        if self.__wp_mode:
            title_list.insert(2, '唤醒结果')
        if self.__split_asr_mode:
            title_list[0] = '文件名'
        if self.__role_asr_mode:
            title_list[0] = '角色答案'
            title_list.insert(2, '角色结果')
        if not 'asr' in self.__audioScene and self.__speech_type != globalVal.SpeechType.VOICE_PRINT:
            title_list[1] = '唤醒Index'
        if self.__speech_type == globalVal.SpeechType.WP_LOCATION:
            title_list.extend(['', '标记声源位置', 'soundLocation', '是否正确定位(1:定位正确，2:相邻灯位正确，0:错误)', '定位角度差'])
        elif self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
            title_list.extend(['', '声纹模型ID', '声纹验证结果', '是否集内', '是否正确判定'])
        xls.write_content(self.__worksheet, self.__excelCount, 1, self.__samplesetName, 3)
        # 主要耗时1
        xls.write_title(self.__worksheet, title_list, self.__excelCount + 1)

    def process_data(self):
        """
        对齐答案和打点数据
        """
        if len(self.__logResultList) or len(self.__AnswerList):
            if self.__mutiply_asr_mode or self.__pure_asr_mode:
                self.process_mutiplyasr()
            elif self.__speech_type == globalVal.SpeechType.SPLIT_ASR:
                self.process_splitasr()
            else:
                self.process_wp_asr()
        if 'asr' in self.__audioScene:
            if self.__record_for_sever:
                self.__server_answer_file.close()
                self.__server_info_file.close()
        if self.__speech_type == globalVal.SpeechType.DCI:
            self.__dci_data_file.close()
        self.create_performance()

    def process_wp_asr(self):
        """
        对齐模块:纯唤醒、单次识别、dci、声源定位、声纹
        """
        for query_list in self.__logResultList:
            if len(query_list) == 0:
                continue
            if not 'wpTriggered' in query_list:
                logging.error('缺少唤醒回调: %s %s', self.__sdkVersion, self.__samplesetName)
                globalVal.record_err(str(self.__sdkVersion) + '/' + str(self.__samplesetName), 2103, \
                    str(query_list) + ',缺少唤醒回调')
                continue
            wpTime = int(query_list['wpTriggered'])
            if len(self.__AnswerList) > self.__curAnswerQueryIndex:
                if self.is_right_wakeup(wpTime):
                    cur_wakeupTime = wpTime - int(self.__AnswerList[self.__curAnswerQueryIndex]['wp_end'])
                    record_flag = self.process_wp_result(cur_wakeupTime, query_list)
                    if self.__speech_type == globalVal.SpeechType.DCI:
                        # DCI
                        self.process_wp_dci(cur_wakeupTime, query_list, record_flag)
                    elif self.__speech_type == globalVal.SpeechType.WP_LOCATION:
                        # 声源定位(唤醒词错误时，目前也参与统计声源定位准确率)
                        self.process_wp_location(cur_wakeupTime, query_list)
                    elif 'asr' in self.__audioScene or self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
                        if 'asr_end' in self.__AnswerList[self.__curAnswerQueryIndex] \
                            and self.__AnswerList[self.__curAnswerQueryIndex]['asr_end'] != 0:
                            # 单次识别/声纹
                            if 'asr_sn' in query_list:
                                self.process_oneround_asr(wpTime, cur_wakeupTime, query_list, record_flag)
                            else:
                                self.record_to_database(cur_wakeupTime, 'no asr', query_list)
                        else:
                            # 唤醒后，无识别query
                            self.record_to_database(cur_wakeupTime, 'false asr', query_list)
                    else:
                        # 纯唤醒
                        self.record_to_database(cur_wakeupTime, 'false asr', query_list)
                    self.__curAnswerQueryIndex += 1
                else:
                    # 误唤醒
                    self.process_wp_result(None, query_list)
                    self.record_to_database(-9999, 'false wp', query_list)
                    self.record_false_asr(query_list)
            else:
                # 误唤醒&识别
                self.process_wp_result(None, query_list)
                self.record_to_database(-9999, 'false wp', query_list)
                self.record_false_asr(query_list)
        self.process_remain_answer()

    def process_mutiplyasr(self):
        """
        对齐模块:一次唤醒多次识别、纯识别
        """
        for query_list in self.__logResultList:
            if len(query_list) == 0 or not 'vadBegin' in query_list:
                if not 'vadBegin' in query_list:
                    logging.error('缺少vad回调: %s %s', self.__sdkVersion, self.__samplesetName)
                    globalVal.record_err(str(self.__sdkVersion) + '/' + str(self.__samplesetName), 2103, \
                        str(query_list) + ',缺少vad回调')
                continue
            if 'wpTriggered' in query_list and self.__mutiply_asr_mode:
                wp_time = query_list['wpTriggered']
                if self.is_right_wakeup(wp_time):
                    self.__support_info = ''
                    cur_wakeupTime = wp_time - int(self.__AnswerList[self.__curAnswerQueryIndex]['wp_end'])
                    record_flag = self.process_wp_result(cur_wakeupTime, query_list)
                    description = 'first asr'
                    if query_list['asr_result'] is None or query_list['asr_result'] == '':
                        if globalVal._PARAM['SUPPORT_RESTRICT'][1]:
                            # 首次无识别结果，不参与统计
                            self.__support_info = '&unsupported'
                        self.__firstAsrErrorList.append(self.get_recognize_result(query_list, 1))
                    elif checkNetError(query_list, False):
                        if globalVal._PARAM['SUPPORT_RESTRICT'][2]:
                            # 首次有识别结果、但是回调了网络错误码，不参与统计
                            self.__support_info = '&unsupported'
                        description += '&net_err:' + str(query_list['errCode'])
                    self.record_to_database(cur_wakeupTime, description, query_list, 
                        self.record_speed_data(query_list, False))
                    self.__MutiplyasrWpAnswerIndex += 1
                    self.__curAnswerQueryIndex += 1
                else:
                    self.process_wp_result(None, query_list)
                    if globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                        # 误唤醒，之后的2~N次识别不参与统计
                        self.__support_info = '&unsupported'
                    self.record_to_database(-9999, 'false wp', query_list)
            else:
                recgResult = ''
                if query_list['asr_result'] is not None and str(query_list['asr_result']) != '':
                    recgResult = query_list['asr_result']
                if self.__curAnswerQueryIndex < len(self.__AnswerList) and \
                    'wp_answer' in self.__AnswerList[self.__curAnswerQueryIndex]:
                    if int(query_list['vadBegin']) > self.__AnswerList[self.__curAnswerQueryIndex]['wp_end']:
                        if globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                            # 没唤醒，之后的2~N次识别不参与统计
                            self.__support_info = '&unsupported'
                if query_list['asrFinalResult'] is not None \
                    or query_list['errCode'] is not None or query_list['asrCancel'] is not None:
                    rightAsrFlag = self.is_right_mutiplyasr(query_list)
                    if rightAsrFlag == 1 or rightAsrFlag == 2 or rightAsrFlag == 6:
                        if rightAsrFlag == 2:
                            self.record_to_database(-5555, 'asr reject', \
                                query_list, self.record_speed_data(query_list, False))
                        else:
                            if rightAsrFlag == 1 and self.__support_info == '':
                                recordflags = True
                            else:
                                recordflags = False
                            self.record_to_database(-5555, '', query_list, \
                                self.record_speed_data(query_list, recordflags))
                        if rightAsrFlag == 6:
                            self.__multiAsrErrorList.append(self.get_recognize_result(query_list, 1))
                        elif self.__record_for_sever:
                            self.record_for_server(
                                str(recgResult), 
                                query_list['asr_sn'], 
                                query_list['asr_corpus_id']
                            )
                        self.__curAnswerQueryIndex += 1
                    elif rightAsrFlag == 3:
                        # print("!!!!!!!!!!!!!!!!!!!递归路径出现!!!!!!!!!!!!!!!!!!")
                        if 'wpTriggered' in query_list and query_list['wpTriggered'] is not None:
                            self.__support_info = ''
                            cur_wakeupTime = query_list['wpTriggered'] - \
                                int(self.__AnswerList[self.__curAnswerQueryIndex]['wp_end'])
                            self.record_to_database(cur_wakeupTime, 'first asr', query_list)
                        else:
                            if globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                                self.__support_info = '&unsupported'
                            self.record_to_database(-8888, 'no wakeup', query_list)
                            logging.debug('[no wakeup] version:%s sample:%s DOT_ASR_VAD_BEGIN:%s', \
                                self.__sdkVersion, self.__samplesetName, str(query_list['vadBegin']))
                        self.__MutiplyasrWpAnswerIndex += 1
                        self.__curAnswerQueryIndex += 1
                else:
                    logging.debug('[no mutiplyasr callback] version:%s sample:%s DOT_ASR_VAD_BEGIN:%s', \
                        self.__sdkVersion, self.__samplesetName, str(query_list['vadBegin']))
        self.process_remain_answer()

    def process_splitasr(self):
        """
        codriver 识别单条音频
        :return:
        """
        for query_list in self.__logResultList:
            self.record_to_database(-6666, '', query_list)
            self.__curAnswerQueryIndex += 1

    def process_wp_result(self, wpTime, wp_list):
        """
        唤醒结果统计
        """
        if globalVal._PARAM['WRONG_WP_WORDS_RESTRICT']:
            record_flag = False
        else:
            record_flag = True
        curr_wp = wp_list['wpWords']
        if wpTime is not None:  # 正常唤醒
            wp_words = self.__AnswerList[self.__curAnswerQueryIndex]['wp_words']
            wp_index = self.__AnswerList[self.__curAnswerQueryIndex]['wp_index']
            wp_whisper = self.__AnswerList[self.__curAnswerQueryIndex]['wp_whisper']
            self.__wpStatusDict[wp_words]['detectCount'] += 1
            self.__wpStatusDict[wp_words]['totalTime'] += wpTime
            self.__wpStatusDict[wp_words]['wp_speed_list'].append(wpTime)
            if self.__wpStatusDict[wp_words]['wpMaxTime'] < wpTime:
                self.__wpStatusDict[wp_words]['wpMaxTime'] = wpTime
            if curr_wp == wp_words:
                record_flag = True
                if self.__speech_type == globalVal.SpeechType.WHISPER:
                    self.__wpStatusDict[wp_words]['whisperInfo'][wp_whisper]['detectCount'] += 1
                    self.__wpStatusDict[wp_words]['whisperInfo'][wp_whisper]['totalTime'] += wpTime
                    if wp_list['wpWhisper'] > 0:
                        self.__wpStatusDict[wp_words]['whisperInfo'][wp_whisper]['whisperCount'] += 1
                    elif wp_list['wpWhisper'] == 0:
                        self.__wpStatusDict[wp_words]['whisperInfo'][wp_whisper]['normalCount'] += 1
                    if wp_whisper == wp_list['wpWhisper'] or \
                        (wp_whisper > 0 and wp_list['wpWhisper'] > 0):
                        self.__wpStatusDict[wp_words]['whisperInfo'][wp_whisper]['rightCount'] += 1
                        self.__wpStatusDict[wp_words]['rightCount'] += 1
                else:
                    self.__wpStatusDict[wp_words]['rightWordCount'] += 1
                    if wp_list['wpIndex'] == wp_index:
                        self.__wpStatusDict[wp_words]['rightIndexCount'] += 1
                        self.__wpStatusDict[wp_words]['rightCount'] += 1
            elif wp_list['wpIndex'] == wp_index:
                self.__wpStatusDict[wp_words]['rightIndexCount'] += 1
        else:                   # 误唤醒
            if curr_wp in self.__wpFalseDict:
                self.__wpFalseDict[curr_wp] += 1
            else:
                self.__wpFalseDict.update({curr_wp: 1})
        return record_flag

    def process_oneround_asr(self, wpTime, cur_wakeupTime, asr_list, record_flag):
        """
        唤醒后识别对齐
        """
        rightAsrFlag = -1
        if asr_list['asr_result'] is None or str(asr_list['asr_result']) == '':
            recgResult = ''
        else:
            recgResult = asr_list['asr_result']
        if 'jike-true' in self.__audioScene and globalVal.is_jikemode_sn(str(asr_list['asr_sn'])):
            self.record_false_asr(asr_list)
            return
        if asr_list['asrFinalResult'] is not None or asr_list['asrError'] is not None \
            or asr_list['asrCancel'] is not None:
            try:
                max_delay_time = \
                    int(self.__AnswerList[self.__curAnswerQueryIndex + 1]['wp_end']) + 1000
            except:
                max_delay_time = \
                    int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_end']) + 20000
            if self.__offline_asr_mode:
                rightAsrFlag = self.is_right_offline_asr(asr_list, max_delay_time, wpTime)
            else:
                rightAsrFlag = self.is_right_asr(asr_list, max_delay_time)
            if rightAsrFlag == 0: # 对齐，有识别结果、无识别结果
                # 唤醒词正确，记录识别
                if record_flag:
                    if recgResult != '':
                        if self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
                            if recgResult == "我是谁":
                                if asr_list['vpResult']:
                                    asr_list['vpAccuracy'] = self.is_right_voiceprint(asr_list['vpResult'])
                            else:
                                globalVal.record_err(
                                    self.__sdkVersion + '/' + self.__samplesetName, 
                                    5001, 
                                    '声纹识别结果错误 sn:' + asr_list['asr_sn'] \
                                    + ' 实际识别结果:' + recgResult \
                                    + ' tts:' + str(asr_list['vpResult'])
                                )
                        self.record_to_database(cur_wakeupTime, '', asr_list, self.record_speed_data(asr_list, True))
                        if self.__record_for_sever:
                            self.record_for_server(
                                recgResult, 
                                asr_list['asr_sn'], 
                                asr_list['asr_corpus_id']
                            )
                    else:
                        self.__firstAsrErrorList.append(self.get_recognize_result(asr_list, 1))
                        self.record_to_database(cur_wakeupTime, '', 
                            asr_list, self.record_speed_data(asr_list, False))
                # 唤醒词错误，识别不参与记录
                else:
                    self.record_to_database(cur_wakeupTime, 'ignore asr', asr_list)
            elif rightAsrFlag == 1: # 未对齐、原句无识别
                self.record_to_database(cur_wakeupTime, 'no asr', asr_list)
                self.record_to_database(-7777, 'false asr', asr_list)
                self.record_false_asr(asr_list)
            elif rightAsrFlag == 2: # 未对齐，无标注答案
                self.record_to_database(-7777, 'false asr', asr_list)
                self.record_false_asr(asr_list)
        else:
            # 0423 如果缺少识别关键回调、无法对齐，则记录识别结果为空
            if asr_list['asrIgnore'] is not None:
                self.__ignore_asr += 1
                self.record_to_database(cur_wakeupTime, 'ignore asr', asr_list)
            else:
                self.record_to_database(cur_wakeupTime, 'no asr', asr_list)
                logging.debug('[no asr callback] version:%s sample:%s DOT_WP_SUCCEESS:%s', 
                    self.__sdkVersion, self.__samplesetName, str(wpTime))

    def process_wp_dci(self, cur_wakeupTime, wp_list, record_flag):
        """
        写入dci特征文件
        """
        if record_flag:
            if wp_list['dci_data'] is not None:
                self.__dci_data_file.write(self.__samplesetName + '_' + str(self.__curAnswerQueryIndex) + '.pcm:' \
                    + str(cur_wakeupTime) + ':' + wp_list['dci_data'] + '\n')
            else:
                logging.error('找不到dci特征:' + str(wp_list))
        self.write_to_excel(cur_wakeupTime, str(self.__curAnswerQueryIndex + 1), wp_list)

    def process_wp_location(self, cur_wakeupTime, wp_list):
        """
        记录声源定位准确率
        """
        if wp_list['soundLocation'] is not None:
            wp_list['locationAccuracy'] = self.is_right_location(
                self.__AnswerList[self.__curAnswerQueryIndex]['wp_location'], 
                wp_list['soundLocation'])
        else:
            wp_list['locationAccuracy'] = 'False:None'
            logging.debug('[no soundLocation] version:%s sample:%s DOT_WP_SUCCEESS:%s', 
                self.__sdkVersion, self.__samplesetName, str(cur_wakeupTime))
        self.write_to_excel(cur_wakeupTime, self.get_recognize_result(wp_list), wp_list)

    def process_remain_answer(self):
        """
        对齐结束后，判断标注答案是否还没对齐完，补充剩余的答案展示
        """
        for answer_index in range(self.__curAnswerQueryIndex, len(self.__AnswerList)):
            if self.__mutiply_asr_mode or self.__pure_asr_mode:
                if 'wp_begin' in self.__AnswerList[answer_index]:
                    if globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                        self.__support_info = '&unsupported'
                    self.record_to_database(-8888, 'no wakeup', [])
                    self.__MutiplyasrWpAnswerIndex += 1
                else:
                    self.record_to_database(-8888, 'no asr multi', [])
            else:
                self.record_to_database(-8888, 'no wakeup', [])
            self.__curAnswerQueryIndex += 1

    def is_right_wakeup(self, wpTime):
        """
        判断此次唤醒，是正常唤醒/误唤醒
        """
        if self.__curAnswerQueryIndex >= len(self.__AnswerList):
            return 0
        if not 'wp_begin' in self.__AnswerList[self.__curAnswerQueryIndex]:
            if self.__mutiply_asr_mode and int(wpTime) > int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_end']):
                logging.debug('no asr multi')
                self.record_to_database(-8888, 'no asr multi', [])
                self.__curAnswerQueryIndex += 1
                return self.is_right_wakeup(wpTime)
            else:
                return 0

        wpBegin = int(self.__AnswerList[self.__curAnswerQueryIndex]['wp_begin'])
        wpEnd = int(self.__AnswerList[self.__curAnswerQueryIndex]['wp_end'])
        logging.debug('[isRightWakeup] wpBegin:' + str(wpBegin) + ', wpTime:' + str(wpTime))
        if wpTime < wpBegin:
            logging.debug('false wakeup')
            return 0
        # 认为在唤醒词说完800ms内能唤醒，否则算作误唤醒
        elif wpTime > wpBegin and wpTime <= wpEnd + 800:
            logging.debug('right wakeup')
            self.__wpDetectCount += 1
            return 1
        else:
            self.record_to_database(-8888, 'no wakeup', [])
            logging.debug('[isRightWakeup] no wakeup')
            self.__MutiplyasrWpAnswerIndex += 1
            self.__curAnswerQueryIndex += 1
            if self.__mutiply_asr_mode and globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                # 没唤醒，不参与统计
                self.__support_info = '&unsupported'
            return self.is_right_wakeup(wpTime)

    def is_right_asr(self, list_value, max_delay_time):
        """
        唤醒后识别:判断此次识别，是正常识别/误识别
        0:对齐,有识别结果、无识别结果&非网络错误、无识别结果&网络错误
        1:未对齐、原句无识别，2:未对齐，无标注答案
        """
        if self.__curAnswerQueryIndex >= len(self.__AnswerList):
            logging.debug('wrong asr after all answer')
            return 2

        finalResultTime = 0
        endTime = 0
        firstPartialTime = 0
        if list_value['asrFinalResult'] is not None:
            finalResultTime = int(list_value['asrFinalResult'])
        elif list_value['asrFirstPartial'] is not None:
            finalResultTime = int(list_value['asrFirstPartial'])
        if list_value['asrError'] is not None:
            endTime = int(list_value['asrError'])
        elif list_value['asrCancel'] is not None:
            endTime = int(list_value['asrCancel'])
        # if endTime > max_delay_time: // 1202
        #     # logging.error('endTime error, sdk ' + str(self.__sdkVersion) + ', audio ' + str(self.__samplesetName) + \
        #     #     ', query at ' + str(self.__AnswerList[self.__curAnswerQueryIndex]['asr_query']))
        if list_value['asrFirstPartial'] is not None:
            firstPartialTime = int(list_value['asrFirstPartial'])
        elif finalResultTime != 0:
            firstPartialTime = finalResultTime
        else:
            firstPartialTime = endTime
        
        try:
            vadBeginTime = int(list_value['vadBegin'])
            vadEndTime = int(list_value['vadEnd'])
        except:
            if self.__use_vad:
                logging.error('缺少vad回调: %s %s %s ms处', self.__sdkVersion, self.__samplesetName, finalResultTime)
                globalVal.record_err(str(self.__sdkVersion) + '/' + str(self.__samplesetName), 2101, \
                    str(finalResultTime) + 'ms附近缺少vad回调')
            vadBeginTime = 0
            vadEndTime = 0
        
        asrBegin, asrEnd, threshold_time = self.get_curr_answer_time()
        # 假设1秒硬延迟、到下一次的唤醒+1秒超时
        if asrBegin == 0 or firstPartialTime < asrBegin - 2000:
            logging.debug('wrong asr before speak')
            return 1
        elif (asrBegin + 600 + threshold_time < finalResultTime < max_delay_time) \
            or (vadBeginTime < asrEnd - threshold_time and vadEndTime > asrBegin + 600 + threshold_time) \
            or (asrBegin - 1000 < endTime < max_delay_time):
            logging.debug('right asr')
            self.__asrDetectCount += 1
            return 0
        else:
            logging.debug('wrong asr after speak')
            return 1

    def is_right_offline_asr(self, list_value, max_delay_time, wakeupTime):
        """
        唤醒后识别:判断此次识别，是正常识别/误识别
        0:对齐,有识别结果、无识别结果&非网络错误、无识别结果&网络错误
        1:未对齐、原句无识别，2:未对齐，无标注答案
        """
        if self.__curAnswerQueryIndex >= len(self.__AnswerList):
            logging.debug('wrong asr after all answer')
            return 2

        finalResultTime = 0
        endTime = 0
        if list_value['asrFinalResult'] is not None:
            finalResultTime = int(list_value['asrFinalResult'])
        if list_value['asrError'] is not None:
            endTime = int(list_value['asrError'])

        try:
            vadBeginTime = int(list_value['vadBegin'])
            vadEndTime = int(list_value['vadEnd'])
        except:
            if self.__use_vad:
                logging.error('缺少vad回调: %s %s %s ms处', self.__sdkVersion, self.__samplesetName, finalResultTime)
                globalVal.record_err(str(self.__sdkVersion) + '/' + str(self.__samplesetName), 2101, \
                    str(finalResultTime) + 'ms附近缺少vad回调')
            vadBeginTime = 0
            vadEndTime = 0
        
        asrBegin, asrEnd, threshold_time = self.get_curr_answer_time()
        # 只要finalResult在 [asrBegin, 下一次唤醒时刻] 范围内，即认为识别对齐
        if asrBegin == 0 or (finalResultTime != 0 and finalResultTime < asrBegin):
            logging.debug('wrong asr before speak')
            return 1
        elif (asrBegin < finalResultTime < max_delay_time) or \
            (vadBeginTime < asrEnd - threshold_time and vadEndTime > asrBegin + 600 + threshold_time) or \
            (wakeupTime < endTime < max_delay_time):
            logging.debug('right asr')
            self.__asrDetectCount += 1
            return 0
        else:
            logging.debug('wrong asr after speak')
            return 1

    def is_right_mutiplyasr(self, list_value):
        """
        一次唤醒多次识别:判断此次识别，是正常识别/误识别
        1:有识别&多次识别&asr_reject为0， 2:有识别&多次识别&asr_reject为1， 3:有识别&唤醒后识别，
        4:误识别&asr_reject为0， 5:误识别&asr_reject为1 6:对齐但报错（无识别结果）
        """
        reject_flag = -1
        if list_value['asrFinalResult'] is not None:
            finalResultTime = int(list_value['asrFinalResult'])
            reject_flag = self.record_asr_reject_status(list_value)
        elif list_value['asrError'] is not None:
            finalResultTime = int(list_value['asrError'])
        elif list_value['asrCancel'] is not None:
            finalResultTime = int(list_value['asrCancel'])
        else:
            finalResultTime = 0
        vadBeginTime = int(list_value['vadBegin'])
        try:
            vadEndTime = int(list_value['vadEnd'])
        except:
            vadEndTime = finalResultTime

        if self.__curAnswerQueryIndex >= len(self.__AnswerList) \
            or not 'asr_begin' in self.__AnswerList[self.__curAnswerQueryIndex]:
            if self.__curAnswerQueryIndex < len(self.__AnswerList) and \
                'wp_begin' in self.__AnswerList[self.__curAnswerQueryIndex] and \
                vadBeginTime > int(self.__AnswerList[self.__curAnswerQueryIndex]['wp_end']):
                    if globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                        self.__support_info = '&unsupported'
                    self.record_to_database(-8888, 'no wakeup&无识别标注', list_value) # 非首次的未唤醒, 只有唤醒标注、无识别标注
                    self.__MutiplyasrWpAnswerIndex += 1
                    self.__curAnswerQueryIndex += 1
                    logging.debug('[is_right_mutiplyasr] no wakeup 1')
                    return self.is_right_mutiplyasr(list_value)
            self.__falseMutiplyVadCount += 1
            logging.debug('wrong asr, reject:%s', str(reject_flag))
            if reject_flag == 1:
                self.record_to_database(-7777, 'false asr&asr reject', list_value)
                return 5
            else:
                if list_value['asrError'] is None and list_value['asrCancel'] is None \
                    and self.__support_info == '':
                    self.__falseMutiplyasrCount += 1
                self.record_to_database(-7777, 'false asr', list_value)
                return 4

        asrBegin, asrEnd, threshold_time = self.get_curr_answer_time()
        asr_answer = self.__AnswerList[self.__curAnswerQueryIndex]['asr_query']
        asr_result = list_value['asr_result']
        ratio_threshold = 0.5
        if asr_answer and asr_result:
            if asr_answer.isalpha() and asr_result.isalpha():
                asr_answer = asr_answer.upper()
                asr_result = asr_result.upper()
            seq = difflib.SequenceMatcher(None, asr_answer, asr_result)
            ratio = seq.ratio()
            if min(len(asr_answer), len(asr_result)) >= 10:
                ratio_threshold = 1
        else:
            ratio = 0

        logging.debug('[is_right_mutiplyasr] asr_begin:' + str(asrBegin) + ',asr_end:' + str(asrEnd) + ',vad_begin:' \
            + str(vadBeginTime) + ',vad_end:' + str(vadEndTime) + ",threshold_time:" + str(threshold_time) \
            + ",asr_answer:" + str(asr_answer) + ",asr_result:" + str(asr_result) \
            + ",ratio:" + str(ratio) + ',ratio_threshold:' + str(ratio_threshold))
        # 认为至少需要识别1.25秒的正常语音，才不算误识别(如何排除网络因素)
        if (vadEndTime < asrBegin + threshold_time and not self.__pure_asr_mode) \
            or (self.__pure_asr_mode and vadEndTime < asrBegin + threshold_time and ratio <= ratio_threshold):  # wrong asr
            self.__falseMutiplyVadCount += 1
            logging.debug('wrong asr, reject:%s', str(reject_flag))
            if reject_flag == 1:
                self.record_to_database(-7777, 'false asr&asr reject', list_value)
                return 5
            else:
                if list_value['asrError'] is None and list_value['asrCancel'] is None:
                    self.__falseMutiplyasrCount += 1
                self.record_to_database(-7777, 'false asr', list_value)
                return 4
        elif (asrEnd < vadBeginTime + threshold_time and not self.__pure_asr_mode) \
            or (self.__pure_asr_mode and asrEnd < vadBeginTime + threshold_time and ratio <= ratio_threshold):
            if 'wp_begin' in self.__AnswerList[self.__curAnswerQueryIndex]:
                if globalVal._PARAM['SUPPORT_RESTRICT'][0]:
                    self.__support_info = '&unsupported'
                self.record_to_database(-8888, 'no wakeup', list_value) # 非首次的未唤醒，有唤醒+识别标注
                logging.debug('[is_right_mutiplyasr] no wakeup 2')
                self.__MutiplyasrWpAnswerIndex += 1
            else:
                self.record_to_database(-8888, 'no asr multi', list_value)
            self.__curAnswerQueryIndex += 1
            logging.debug('no asr multi')
            return self.is_right_mutiplyasr(list_value)
        else:
            if 'wp_begin' in self.__AnswerList[self.__curAnswerQueryIndex]:
                return 3
            self.__asrDetectCount += 1
            logging.debug('right asr, reject:%s', str(reject_flag))
            if list_value['asrFinalResult']:
                if reject_flag == 1:
                    return 2
                else:
                    return 1
            else:
                return 6

    def is_right_location(self, answer, result):
        """
        声源定位:判断定位结果是否正确
        """
        scene = answer.rsplit('_', 1)[0]
        location = answer.rsplit('_', 1)[1]
        is_right_location, location_diff = speech_parse.cal_wp_location_diff(int(location), int(result))
        if scene in self.__wpLocationResultDict:
            if location in self.__wpLocationResultDict[scene]:
                self.__wpLocationResultDict[scene][location].update(
                    {self.__curAnswerQueryIndex:[is_right_location, location_diff]})
            else:
                self.__wpLocationResultDict[scene].update(
                    {location:{self.__curAnswerQueryIndex:[is_right_location, location_diff]}})
        else:
            self.__wpLocationResultDict.update(
                {scene:{location:{self.__curAnswerQueryIndex:[is_right_location, location_diff]}}})
        self.__wpLocationAnswerDict[scene][location]['wpCount'] += 1
        return str(is_right_location) + ':' + str(location_diff)

    def is_right_voiceprint(self, result):
        """
        声纹判别结果是否正确
        [集内判定] 正确:判定为正确的人, 集内错误:判定为集内其他人, 未识别:不认识
        [集外判定] 正确:不认识, 集内错误:判定为集内某一人
        """
        self.__vpTotalCount += 1
        if self.__is_jinei == '集内':
            if str(self.__vp_id) in str(result):
                self.__vpRightCount += 1
                is_judge_right = '正确'
            else:
                for vp_id in self.__vp_list:
                    if str(vp_id) in str(result):
                        self.__vpWrongPersonCount += 1
                        return '错误:' + str(vp_id)
                is_judge_right = '未识别'
        else:
            for vp_id in self.__vp_list:
                if str(vp_id) in str(result):
                    self.__vpWrongPersonCount += 1
                    return '错误:' + str(vp_id)
            self.__vpRightCount += 1
            is_judge_right = '正确'
        return is_judge_right

    def get_recognize_result(self, asr_list, mode=0):
        """
        获取识别结果: finalresult/cancel/error
        """
        if not 'asr_sn' in asr_list:
            return ''
        if 'file_name' in asr_list or self.__split_asr_mode:
            if asr_list['asr_result'] and asr_list['asr_result'] != '' and asr_list['asr_result'] != 'NULL':
                result = asr_list['asr_result']
            else:
                result = 'err:' + str(asr_list['errCode'])
        elif asr_list['asr_result']:
            if self.__semantic_asr_mode:
                result = str(asr_list['semantics_index']) + '#' + str(asr_list['asr_result'])
            else:
                result = str(asr_list['asr_result'])
        elif asr_list['errCode'] is not None and asr_list['errCode'] != 0:
            if mode:
                result = str(asr_list['errCode'])
            elif checkNetError(asr_list, False):
                result = 'net_err:' + str(asr_list['errCode'])
            else:
                result = 'err:' + str(asr_list['errCode'])
        elif asr_list['asrCancel']:
            if mode:
                result = 'cancel'
            else:
                result = 'net_err:cancel'
        else:
            result = ''
        return result

    def record_for_server(self, recognize_result, sn, corpus_id):
        """
        写标注答案和识别结果文件
        """
        currAnswer = self.__AnswerList[self.__curAnswerQueryIndex]
        if 'asr_query' in currAnswer:
            if sn != '' and str(sn) != 'None':
                if corpus_id != '' and str(corpus_id) != 'None':
                    self.__server_answer_file.write(str(corpus_id) + ':' \
                        + str(currAnswer['asr_query']) + '\n')
                self.__server_info_file.write('asr_answer:' + str(currAnswer['asr_query']) + '\tasr_result:' \
                    + recognize_result + '\tasr_sn:' + str(sn) + '\tcorpus_no:' + str(corpus_id) + '\n')

    def record_false_asr(self, asr_list):
        """
        判断当前识别结果，是否符合误识别标准(没对齐&有识别结果&未拒识)
        """
        if 'asr_result' in asr_list:
            if asr_list['asr_result'] is not None and str(asr_list['asr_result']) != '':
                if str(asr_list['asr_reject']) == '0':
                    self.__falseAsrCount += 1

    def record_asr_reject_status(self, list_value):
        """
        统计极客模式下的拒识情况
        """
        if self.__pure_asr_mode:
            if self.__role_asr_mode:
                return 0
            else:
                if self.__angle_in_range and list_value['direction_invalid'] == '0':
                    return 1
                elif not self.__angle_in_range and list_value['direction_invalid'] == '1':
                    return 1
                else:
                    return 0
        elif list_value['asr_reject'] == '1':
            self.__asr_reject_count += 1
            if list_value['state'] == '00000001':       # 正常接收
                self.__asr_status['normal_accept'] = self.__asr_status['normal_accept'] + 1
            elif list_value['state'] == '00000010':     # 语义拒识
                self.__asr_status['semantics_reject'] = self.__asr_status['semantics_reject'] + 1
            elif list_value['state'] == '00000100':     # 声学拒识
                self.__asr_status['acoustics_reject'] = self.__asr_status['acoustics_reject'] + 1
            elif list_value['state'] == '00001000':     # 其他拒识
                self.__asr_status['other_reject'] = self.__asr_status['other_reject'] + 1
            elif list_value['state'] == '00010000':     # 声纹拒识
                self.__asr_status['voiceprint_rejcet'] = self.__asr_status['voiceprint_rejcet'] + 1
            return 1
        else:
            return 0

    def record_speed_data(self, list_value, record_lag=True):
        """
        速度指标记录:全部识别结果用来计算字准、句准并写入表格，仅符合预期的时间用来计算速度数据
        """
        if list_value['asrFinalResult'] is None or self.__curAnswerQueryIndex >= len(self.__AnswerList) \
            or not 'asr_end' in self.__AnswerList[self.__curAnswerQueryIndex]:
            return [0, 0, 0, 0, 0]
        finishTime = int(list_value['asrFinalResult']) - int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_end'])
        try:
            responseTime = int(list_value['asrFirstPartial']) \
                - int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_begin'])
        except:
            # 如果没有首包信息，则以final result时间作为拿到首包的时间
            responseTime = int(list_value['asrFinalResult']) \
                - int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_begin'])
        try:
            ttsTime = int(list_value['asrFirstTTS']) - int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_end'])
        except:
            ttsTime = 0
        if self.__semantic_asr_mode or self.__offline_asr_mode:
            record_lag = True
        elif (globalVal._PARAM['PARTIAL_TIME_RANGE'][0] is not None and \
            responseTime <= globalVal._PARAM['PARTIAL_TIME_RANGE'][0]) or \
            (globalVal._PARAM['PARTIAL_TIME_RANGE'][1] is not None and \
            responseTime > globalVal._PARAM['PARTIAL_TIME_RANGE'][1]) or \
            (globalVal._PARAM['FINAL_TIME_RANGE'][0] is not None and \
            finishTime <= globalVal._PARAM['FINAL_TIME_RANGE'][0]) or \
            (globalVal._PARAM['FINAL_TIME_RANGE'][1] is not None and \
            finishTime > globalVal._PARAM['FINAL_TIME_RANGE'][1]) or \
            (globalVal._PARAM['TTS_TIME_RANGE'][0] is not None and \
            ttsTime <= globalVal._PARAM['TTS_TIME_RANGE'][0]) or \
            (globalVal._PARAM['TTS_TIME_RANGE'][1] is not None and \
            ttsTime > globalVal._PARAM['TTS_TIME_RANGE'][1]):
            record_lag = False
        if self.__use_vad:
            try:
                vadBeginTime = int(list_value['vadBegin']) \
                    - int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_begin'])
            except:
                vadBeginTime = 0
            try:
                vadEndTime = int(list_value['vadEnd']) - int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_end'])
            except:
                vadEndTime = 0
        if record_lag:
            self.__asrResponseTimeList.append(responseTime)
            self.__asrFinishTimeList.append(finishTime)
            self.__asrChunkTtsTimeList.append(ttsTime)
            if self.__use_vad:
                self.__asrVadBeginTimeList.append(vadBeginTime)
                self.__asrVadEndTimeList.append(vadEndTime)
        if self.__use_vad:
            return [responseTime, finishTime, ttsTime, vadBeginTime, vadEndTime]
        else:
            return [responseTime, finishTime, ttsTime, 0, 0]

    def get_curr_answer_time(self):
        """根据当前标注query时长，动态调整判定阈值"""
        asrBegin = int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_begin'])
        asrEnd = int(self.__AnswerList[self.__curAnswerQueryIndex]['asr_end'])
        if asrEnd - asrBegin > 1500:
            threshold_time = 500
        else:
            threshold_time = (asrEnd - asrBegin) / 4
        return asrBegin, asrEnd, threshold_time

    def write_to_excel(self, wpTime, asr_result, asr_list, time_list=[0, 0, 0, 0, 0]):
        """
        将详细有效数据计入表格
        -8888: 未唤醒（标注答案里，该时刻喊了唤醒词，但是没有被唤醒）
        -9999: 误唤醒（标注答案里，该时刻没有标记唤醒）
        -6666: 单条识别，codriver
        -7777: 误识别（标注答案里，该时刻没有标记query，但是有识别结果）
        """
        responseTime, finishTime, ttsTime, vadBeginTime, vadEndTime = time_list
        self.__excelCount += 1
        row_count = self.__excelCount + 1
        if self.__wp_mode:
            ins = 1
            if wpTime != -8888 and wpTime != -7777:
                if self.__speech_type == globalVal.SpeechType.WHISPER:
                    xls.write_content(self.__worksheet, row_count, 3, 
                        asr_list['wpWords'] + '&' + str(asr_list['wpWhisper']))
                elif 'wpWords' in asr_list:
                    xls.write_content(self.__worksheet, row_count, 3, asr_list['wpWords'])
        elif self.__role_asr_mode:
            ins = 1
        else:
            ins = 0

        if wpTime == -8888:
            xls.write_content(self.__worksheet, row_count, 3, asr_result)
        elif 'no asr' in asr_result:
            xls.write_content(self.__worksheet, row_count, 3 + ins, asr_result, 4)
        elif 'false asr' in asr_result:
            xls.write_content(self.__worksheet, row_count, 3 + ins, asr_result, 2)
        else:
            xls.write_content(self.__worksheet, row_count, 3 + ins, asr_result)

        if wpTime == -9999:                                                      # false wakeup
            xls.write_content(self.__worksheet, row_count, 1, '-', 1)
            xls.write_content(self.__worksheet, row_count, 2, '-', 1)
            xls.write_content(self.__worksheet, row_count, 11 + ins, asr_list['wpTriggered'])
            if self.__speech_type == globalVal.SpeechType.WP_LOCATION:
                xls.write_content(self.__worksheet, row_count, 26 + ins, asr_list['soundLocation'])
            if not 'asr_sn' in asr_list:
                return
        elif wpTime == -6666:
            curr_filename = asr_list['file_name']
            for answer_item in self.__AnswerList:
                if curr_filename == answer_item['file_name']:
                    xls.write_content(self.__worksheet, row_count, 1, curr_filename)
                    xls.write_content(self.__worksheet, row_count, 2, answer_item['asr_query'])
                    xls.write_content(self.__worksheet, row_count, 4 + ins, asr_list['asr_sn'])
                    return
        else:
            if wpTime != -7777:
                if self.__curAnswerQueryIndex < len(self.__AnswerList):
                    curr_answer = self.__AnswerList[self.__curAnswerQueryIndex]
                else:
                    curr_answer = []
                if 'wp_words' in curr_answer:
                    if self.__speech_type == globalVal.SpeechType.WHISPER:
                        xls.write_content(self.__worksheet, row_count, 1, 
                            curr_answer['wp_words'] + '&' + str(curr_answer['wp_whisper']))
                    else:
                        xls.write_content(self.__worksheet, row_count, 1, curr_answer['wp_words'])
                    xls.write_content(self.__worksheet, row_count, 6 + ins, curr_answer['wp_begin'])
                    xls.write_content(self.__worksheet, row_count, 7 + ins, curr_answer['wp_end'])
                    if 'wp_location' in self.__audioScene:
                        xls.write_content(self.__worksheet, row_count, 
                            25 + ins, curr_answer['wp_location'])
                        if wpTime == -8888:
                            return
                        xls.write_content(self.__worksheet, 
                            row_count, 26 + ins, asr_list['soundLocation'])
                        xls.write_content(self.__worksheet, 
                            row_count, 27 + ins, asr_list['locationAccuracy'].split(':')[0])
                        xls.write_content(self.__worksheet, 
                            row_count, 28 + ins, asr_list['locationAccuracy'].split(':')[1])
                if 'asr_query' in curr_answer:
                    if self.__role_asr_mode:
                        xls.write_content(self.__worksheet, row_count, 1, curr_answer['asr_role'], 
                            getRoleColor(curr_answer['asr_role']))
                    xls.write_content(self.__worksheet, row_count, 2, curr_answer['asr_query'])
                    xls.write_content(self.__worksheet, row_count, 8 + ins, curr_answer['asr_begin'])
                    xls.write_content(self.__worksheet, row_count, 9 + ins, curr_answer['asr_end'])
                    if wpTime != -8888 and 'voiceprint' in self.__audioScene \
                        and 'vpResult' in asr_list and asr_list['vpResult'] is not None:
                        xls.write_content(self.__worksheet, row_count, 25 + ins, self.__vp_id)
                        xls.write_content(self.__worksheet, row_count, 26 + ins, asr_list['vpResult'])
                        xls.write_content(self.__worksheet, row_count, 27 + ins, self.__is_jinei)
                        xls.write_content(self.__worksheet, row_count, 28 + ins, asr_list['vpAccuracy'])
                elif 'asr' in self.__audioScene or self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
                    xls.write_content(self.__worksheet, row_count, 2, '-', 1)
                else:
                    xls.write_content(self.__worksheet, row_count, 2, self.__curAnswerQueryIndex + 1)
            if wpTime == -8888:
                return
            if 'wpTriggered' in asr_list and wpTime != -7777:
                xls.write_content(self.__worksheet, row_count, 11 + ins, asr_list['wpTriggered'])
                xls.write_content(self.__worksheet, row_count, 18 + ins, wpTime)
            if not 'asrReady' in asr_list or 'no asr' in asr_result or not 'asr_sn' in asr_list:
                return
            xls.write_content(self.__worksheet, row_count, 12 + ins, asr_list['asrFirstPartial'])
            xls.write_content(self.__worksheet, row_count, 19 + ins, responseTime)
            xls.write_content(self.__worksheet, row_count, 14 + ins, asr_list['asrFirstTTS'])
            xls.write_content(self.__worksheet, row_count, 20 + ins, finishTime)
            xls.write_content(self.__worksheet, row_count, 21 + ins, ttsTime)
            if self.__use_vad:
                xls.write_content(self.__worksheet, row_count, 22 + ins, vadBeginTime)
                xls.write_content(self.__worksheet, row_count, 23 + ins, vadEndTime)
        if self.__use_vad:
            xls.write_content(self.__worksheet, row_count, 15 + ins, asr_list['vadBegin'])
            xls.write_content(self.__worksheet, row_count, 16 + ins, asr_list['vadEnd'])
        if globalVal._PARAM['ASR_SN_WRITE_FLAG']:
            xls.write_content(self.__worksheet, row_count, 4 + ins, asr_list['asr_sn'])
            xls.write_content(self.__worksheet, row_count, 5 + ins, asr_list['asr_corpus_id'])
        if self.__role_asr_mode and asr_list['asrFinalResult'] is not None:
            xls.write_content(self.__worksheet, row_count, 3, asr_list['asr_role'],
                getRoleColor(asr_list['asr_role']))
        if self.__split_asr_mode:
            return
        xls.write_content(self.__worksheet, row_count, 13 + ins, asr_list['asrFinalResult'])

    def write_to_wp_excel(self):
        """
            针对车载多唤醒词、车载指令词、小度指令词功能，单独进行唤醒详情统计
        """
        wp_excel_name = self.__sdkVersion + '_wp'
        if wp_excel_name in self.__workBook.sheetnames:
            wp_excel = self.__workBook[wp_excel_name]
            row = wp_excel.max_row + 5
        else:
            wp_excel = self.__workBook.create_sheet(wp_excel_name)
            row = 1
        if self.__wp_mode == 1:
            wak_title = copy.deepcopy(globalVal._CODRIVER_WAK_PERFORM_LIST)
            wak_list = ['queryCount', 'detectCount', 'rightCount', 'rightRate', 
                'wrongWordCount', 'wrongWordRate', 'wrongIndexCount', 'wrongIndexRate', 
                'wpAvgTime', 'wpMaxTime']
        elif self.__wp_mode == 2:
            wak_title = copy.deepcopy(globalVal._XD_WAK_PERFORM_LIST)
            wak_list = ['queryCount', 'detectCount', 'rightWordCount', 'rightWordRate', 
                'wrongWordCount', 'wrongWordRate', 'wrongIndexCount', 'wpAvgTime']
        elif self.__wp_mode == 3:
            wak_title = copy.deepcopy(globalVal._QWQD_WAK_PERFORM_LIST)
            wak_list = ['queryCount', 'detectCount', 'normalCount', 
                'whisperCount', 'wpRate', 'rightRate', 'wpAvgTime']
        
        xls.write_content(wp_excel, row, 1, self.__samplesetName, 3)
        xls.write_title(wp_excel, wak_title, row + 1)
        add_row = 1
        # 写标注答案里，每个唤醒词的唤醒详情
        for wpWords in self.__wpStatusDict:
            if wpWords == 'summary':
                continue
            add_row += 1
            xls.write_content(wp_excel, row + add_row, 1, wpWords)
            if self.__wp_mode == 3:
                whisperInfo = self.__wpStatusDict[wpWords]['whisperInfo']
                for mode in whisperInfo:
                    col = 2
                    if mode == 1:
                        xls.write_content(wp_excel, row + add_row, col, '气声')
                    elif mode == 2:
                        xls.write_content(wp_excel, row + add_row, col, '轻声')
                    else:
                        xls.write_content(wp_excel, row + add_row, col, '正常')
                    col += 1
                    for key in wak_list:
                        xls.write_content(wp_excel, row + add_row, col, whisperInfo[mode][key])
                        col += 1
                    add_row += 1
            else:
                col = 2
                for key in wak_list:
                    xls.write_content(wp_excel, row + add_row, col, self.__wpStatusDict[wpWords][key])
                    col += 1
        # 写汇总唤醒详情
        if self.__wp_mode != 3:
            add_row += 1
            xls.write_content(wp_excel, row + add_row, 1, '汇总', 7)
            col = 2
            for key in wak_list:
                xls.write_content(wp_excel, row + add_row, col, self.__wpStatusDict['summary'][key], 7)
                col += 1
        # 写标注答案外，每个唤醒词的误报次数
        col = 1
        xls.write_content(wp_excel, row + add_row + 1, col, '误报次数:总计', 0)
        xls.write_content(wp_excel, row + add_row + 2, col, self.__wpFalseDict['summary'])
        for wpWords in self.__wpFalseDict:
            if wpWords == 'summary':
                continue
            col += 1
            xls.write_content(wp_excel, row + add_row + 1, col, str(wpWords), 0)
            xls.write_content(wp_excel, row + add_row + 2, col, self.__wpFalseDict[wpWords])

    def write_to_role_excel(self, row):
        """
            针对听清盒子双人识别交互功能，单独进行分角色的准确率统计
        """
        # [1] 计算角色判定准确率、识别率等指标
        roleThread = speech_parse.InteractAccuracy(
            [self.__audioScene, self.__audioName, self.__sdkVersion], 
            self.__queryResultDict
        )
        role_accuracy = roleThread.get_interact_accuracy()
        # [2] A、B角色，及汇总准确率分别写入excel
        role_title = copy.deepcopy(globalVal._ROLE_PERFORM_LIST)
        role_list = ['total_round', 'right_round', 'round_rate', 'role_wer', 'total_wer']
        i = 0
        for role in list(role_accuracy['role_info'].keys()) + ['sum_info']:
            col = i * len(role_title) + 1
            xls.write_content(self.__worksheet, row, [col, col + len(role_title) - 1], '角色' + str(role), 20 + i)
            xls.write_title(self.__worksheet, role_title, row + 1, col, 0, 0, 20 + i)
            if role == 'sum_info':
                info = role_accuracy['sum_info']
            else:
                info = role_accuracy['role_info'][role]
            k = 0
            for key in role_list:
                xls.write_content(self.__worksheet, row + 2, col + k, info[key])
                k += 1
            i += 1
        return role_accuracy

    def create_performance(self):
        """
            性能数据统计，并写入对应的场景表格和.db文件
        """
        wp_query_count = 0
        asr_ready_count = 0         # [识别指标]识别请求对齐次数（单次:正常唤醒且有识别标注，多次:supported的query）
        asr_invalid_count = 0       # [识别指标]无效识别请求次数（对齐且因为网络原因引起的失败）
        asr_valid_count = 0         # [识别指标]有效识别请求次数（识别成功率的分母）
        asr_success_count = 0       # [识别指标]识别成功次数（对齐且有识别结果、不区分是否拒识）
        first_success_count = 0     # [识别指标]多次识别首次请求成功次数
        asr_recall_count = 0        # [识别指标]多次召回次数、纯识别准召次数

        wp_rate = 0                 # [唤醒指标]唤醒成功率
        success_rate = 0            # [识别指标]识别成功率（识别成功次数/有效识别请求次数）
        recall_rate = 0             # [识别指标]多次召回率、纯识别准召率（召回次数/有效识别请求次数）
        acu_wer = 0                 # [识别指标]字准
        acu_ser = 0                 # [识别指标]句准
        err = []                    # [识别指标]错误率
        row_count = self.__excelCount + 3

        # [1.1]计算唤醒指标
        speech_parse.cal_wp_performance(self.__wpStatusDict, self.__wpFalseDict)
        wp_query_count = self.__wpStatusDict['summary']['queryCount']
        wp_detect_count = self.__wpStatusDict['summary']['detectCount']
        wpTime = self.__wpStatusDict['summary']['wpAvgTime']
        # [1.2]计算识别指标，并记录每一条query请求的详细结果到db
        query_class = speech_parse.QueryFilter(self.__initSceneName, self.__sdkVersion)
        recog_query, count_list = query_class.parse_asr_query(self.__queryResultDict)
        asr_ready_count, asr_invalid_count, asr_valid_count = count_list[0:3]
        asr_success_count, asr_recall_count, first_success_count = count_list[3:]
        asr_valid_count = asr_ready_count - asr_invalid_count
        if asr_valid_count:
            success_rate = round(float((asr_success_count) * 100) / float(asr_valid_count), 3)
            recall_rate = round(float((asr_recall_count) * 100) / float(asr_valid_count), 3)
        if 'asr' in self.__audioScene:
            if self.__semantic_asr_mode:
                acu_wer, acu_ser = speech_parse.cal_asr_wer(
                    self.__asr_answer_file_name, self.__rec_result_file_name, asr_ready_count)
            else:
                acu_wer, acu_ser, err = speech_parse.cal_asr_wer(
                    self.__asr_answer_file_name, self.__rec_result_file_name)
        # [1.3]计算识别速度均值
        responseTime = getAverage(self.__asrResponseTimeList, 0)
        finishTime = getAverage(self.__asrFinishTimeList, 0)
        ttsTime = getAverage(self.__asrChunkTtsTimeList, 0)
        if self.__use_vad:
            vadBeginTime = getAverage(self.__asrVadBeginTimeList, 1)
            vadEndTime = getAverage(self.__asrVadEndTimeList, 1)
        # [2]统计表格写标题、版本、测试集合名称
        work_tile = copy.deepcopy(globalVal._WORK_TITLE_LIST[2:])
        if self.__semantic_asr_mode:
            work_tile[11] = '识别准确率'
            work_tile[12] = '识别错误条数'
        elif self.__split_asr_mode:
            work_tile[8] = '识别成功次数'
            work_tile[9] = '识别成功率'
        elif self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
            work_tile[19] = '集合类型'
            work_tile[20] = '声纹判定总条数'
            work_tile[21] = '正确判定次数'
            work_tile[22] = '错误判定次数'
            if self.__is_jinei == '集内':
                work_tile[23] = '未识别次数'
            else:
                work_tile[23] = ''
        elif self.__wp_mode == 2 and 'asr' in self.__audioScene:
            work_tile.append('识别忽略次数')
        xls.write_title(self.__worksheet, work_tile, row_count - 1)
        if not self.__pure_asr_mode and not self.__split_asr_mode:
            # [2.1]记录唤醒指标
            if wp_query_count:
                wp_rate = round(float(wp_detect_count * 100) / float(wp_query_count), 3)
                xls.write_content(self.__worksheet, row_count, 1, wp_query_count)
                xls.write_content(self.__worksheet, row_count, 3, str(wp_rate) + '%')
            elif self.__speech_type in globalVal.SPEECH_MODE['wp+asr']:
                globalVal.record_err(self.__sdkVersion + '/' + self.__samplesetName, 
                    3100, '统计数据异常，对齐唤醒次数为0')
            xls.write_content(self.__worksheet, row_count, 2, wp_detect_count)
            xls.write_content(self.__worksheet, row_count, 4, self.__totalWpCount - wp_detect_count)
            xls.write_content(self.__worksheet, row_count, 5, wpTime)
            # 如果存在多个唤醒词，再分别记录每个唤醒词的性能
            if self.__wp_mode:
                self.write_to_wp_excel()
        if self.__mutiply_asr_mode:
            # [2.2]极客模式:记录拒识情况
            xls.write_content(self.__worksheet, row_count, 20, self.__asr_reject_count)
            xls.write_content(self.__worksheet, row_count, 21, self.__asr_status['acoustics_reject'])
            xls.write_content(self.__worksheet, row_count, 22, self.__asr_status['semantics_reject'])
            xls.write_content(self.__worksheet, row_count, 23, self.__asr_status['other_reject'])
            xls.write_content(self.__worksheet, row_count, 24, self.__asr_status['voiceprint_rejcet'])
        elif self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
            # [2.3]声纹模式:记录集内、集外指标
            xls.write_content(self.__worksheet, row_count, 20, self.__is_jinei)
            xls.write_content(self.__worksheet, row_count, 21, self.__vpTotalCount)
            xls.write_content(self.__worksheet, row_count, 22, self.__vpRightCount)
            xls.write_content(self.__worksheet, row_count, 23, self.__vpWrongPersonCount)
            if self.__is_jinei == '集内':
                xls.write_content(self.__worksheet, 
                    row_count, 24, self.__vpTotalCount - self.__vpRightCount - self.__vpWrongPersonCount)
        # [2.4]记录识别成功率、召回率、误识别
        xls.write_content(self.__worksheet, row_count, 6, self.__asrQueryCount)
        xls.write_content(self.__worksheet, row_count, 7, asr_ready_count)
        xls.write_content(self.__worksheet, row_count, 8,
            str(asr_invalid_count) + '/' + str(asr_valid_count))
        if self.__mutiply_asr_mode or self.__pure_asr_mode:
            xls.write_content(self.__worksheet, row_count, 9, 
                str(asr_recall_count) + '/' + str(asr_success_count))
            xls.write_content(self.__worksheet, row_count, 10, 
                str(recall_rate) + '%/' + str(success_rate) + '%')
            xls.write_content(self.__worksheet, row_count, 11, 
                str(self.__falseMutiplyasrCount) + '/' + str(self.__falseMutiplyVadCount))
            xls.write_content(self.__worksheet, row_count, 26, first_success_count)
        else:
            xls.write_content(self.__worksheet, row_count, 9, asr_success_count)
            xls.write_content(self.__worksheet, row_count, 10, str(success_rate) + '%')
            xls.write_content(self.__worksheet, row_count, 11, self.__falseAsrCount)
        if 'asr' in self.__audioScene:
            xls.write_content(self.__worksheet, row_count, 12, acu_wer)
            xls.write_content(self.__worksheet, row_count, 13, acu_ser)
        # [2.5]记录速度指标
        xls.write_content(self.__worksheet, row_count, 14, responseTime)
        xls.write_content(self.__worksheet, row_count, 15, finishTime)
        xls.write_content(self.__worksheet, row_count, 16, ttsTime)
        if self.__use_vad:
            xls.write_content(self.__worksheet, row_count, 17, vadBeginTime)
            xls.write_content(self.__worksheet, row_count, 18, vadEndTime)
        xls.write_content(self.__worksheet, row_count, 19, len(self.__asrFinishTimeList))
        # [2.6]记录错误码信息 -- 20210302新版本
        errorList = ''
        for error_item in set(self.__firstAsrErrorList):
            if error_item != '':
                errorList = str(error_item) + ':' + str(self.__firstAsrErrorList.count(error_item)) + '\n'
        xls.write_content(self.__worksheet, row_count, 25, errorList)
        if self.__totalFirstAsrCount == 0:
            self.__totalFirstAsrCount = asr_ready_count
        if self.__speech_type == globalVal.SpeechType.MULTI_ASR:
            error_info = {
                '与标注对齐的首次': speech_parse.parse_asr_err(self.__firstAsrErrorList, self.__wpDetectCount),
                '与标注对齐的2~N次': speech_parse.parse_asr_err(self.__multiAsrErrorList, self.__asrDetectCount)
            }
        else:
            error_info = {
                '与标注对齐的首次': speech_parse.parse_asr_err(self.__firstAsrErrorList, self.__asrDetectCount),
                '与标注对齐的2~N次': speech_parse.parse_asr_err(self.__multiAsrErrorList, 0)
            }
        error_info.update({
            '全部首次': speech_parse.parse_asr_err(self.__totalFirstAsrError, self.__totalFirstAsrCount),
            '全部2~N次': speech_parse.parse_asr_err(self.__totalMultiAsrError, self.__totalMultiAsrCount)
        })
        # [2.7]记录声源定位指标
        if self.__speech_type == globalVal.SpeechType.WP_LOCATION:
            speech_parse.record_wp_location_accuracy(
                self.__wpLocationAnswerDict, self.__wpLocationResultDict,
                self.__sdkVersion, self.__samplesetName
            )
        # [2.8]双人交互测试:记录每个人的性能
        if self.__role_asr_mode:
            role_accuracy = self.write_to_role_excel(row_count + 1)

        # [3]错误码模块，记录统计问题到error_info.txt
        if 'asr' in self.__audioScene and asr_valid_count == 0:
            logging.error('%s %s 统计异常，对齐识别次数为0', self.__sdkVersion, self.__samplesetName)
            globalVal.record_err(self.__sdkVersion + '/' + self.__samplesetName, 3101, '统计数据异常，对齐识别次数为0')
        
        # [4]测试数据归档，记录入db
        query_result = {
            'wp_mode': self.__wp_mode,
            'wp_query_count': wp_query_count,
            'wp_count': wp_detect_count,
            'wp_rate': str(wp_rate) + '%',
            'false_wp_count': self.__totalWpCount - wp_detect_count,
            'wp_time': wpTime,
            'wp_status': self.__wpStatusDict,
            'wp_false': self.__wpFalseDict,
            'asr_query_count': self.__asrQueryCount,
            'asr_count': asr_ready_count,
            'invalid_asr_request_count': asr_invalid_count,
            'ignore_asr_count': self.__ignore_asr,
            'valid_asr_request_count': asr_valid_count,
            'success_count': asr_success_count,
            'success_rate': str(success_rate) + '%',
            'recall_count': asr_recall_count,
            'recall_rate': str(recall_rate) + '%',
            'false_asr_count': self.__falseAsrCount,
            'false_interact_count': self.__falseMutiplyVadCount,
            'asr_error_code_info': error_info,
            'wer': acu_wer,
            'ser': acu_ser,
            'err': err,
            'first_success_count': first_success_count
        }
        query_result.update({
            'partial_time': responseTime,
            'finish_time': finishTime,
            'tts_time': ttsTime,
            'speed_query_num': len(self.__asrFinishTimeList),
            'asr_speed_list': {
                'partial_speed': self.__asrResponseTimeList,
                'tts_speed': self.__asrChunkTtsTimeList,
                'finish_speed': self.__asrFinishTimeList
            }
        })
        if self.__use_vad:
            query_result.update({'vad_start_time': vadBeginTime, 'vad_end_time': vadEndTime})
            query_result['asr_speed_list'].update({
                'vad_start_speed': self.__asrVadBeginTimeList,
                'vad_end_speed': self.__asrVadEndTimeList
            })
        if 'voiceprint' in self.__audioScene:
            query_result.update({
                'vp_total': self.__vpTotalCount,
                'vp_right': self.__vpRightCount,
                'vp_wrong': self.__vpWrongPersonCount
            })
        if self.__role_asr_mode:
            query_result.update({'role_accuracy': role_accuracy})
        db_manager.updateDataToDB(self.__initSceneName, self.__sdkVersion, 'query_result', self.__queryResultDict)
        db_manager.updateDataToDB(self.__initSceneName, self.__sdkVersion, 'recognize_result', recog_query)
        db_manager.updateDataToDB(self.__initSceneName, self.__sdkVersion, 'conclusion', query_result)

    def record_to_database(self, wp_time, recog_decs, asr_list, time_list=[0,0,0,0,0]):
        """
            数据记录：每条query的识别情况
        Args:
            wp_time: [int] 唤醒时刻，或者-8888未唤醒、-7777误识别、-9999误唤醒
            recog_decs: [string] 对齐结论，识别结果的description
            asr_list: [dict] 每条query的识别详情
            time_list: [list] 速度指标，依次为识别首包、TTS首包、识别硬延迟、VAD起点检测时间、VAD尾点检测时间
        Returns:
            null
        """
        # [1] 写入excel, 按照『recg_resullt(desc)』写入
        if self.__mutiply_asr_mode:
            recog_decs += self.__support_info
        if recog_decs == '':
            write_recg_result = self.get_recognize_result(asr_list)
        elif 'false asr' in recog_decs and 'asr_sn' not in asr_list:
            write_recg_result = ''
        else:
            if 'no asr' in recog_decs or 'ignore asr' in recog_decs:
                write_recg_result = '(' + recog_decs + ')'
            else:
                write_recg_result = self.get_recognize_result(asr_list) + '(' + recog_decs + ')'
        self.write_to_excel(wp_time, write_recg_result, asr_list, time_list)
        # [2] 除误唤醒、误识别外，记入resultdict，用于统计识别指标
        if not self.__speech_type in globalVal.SPEECH_MODE['wp'] and not 'false' in recog_decs:
            if self.__mutiply_asr_mode: # pcm_index:同一条整轨音频，识别query的index
                if 'first asr' in recog_decs or 'no wakeup' in recog_decs:
                    pcm_index = str(self.__MutiplyasrWpAnswerIndex + 1) + '_first'
                else:
                    pcm_index = self.__curAnswerQueryIndex - self.__MutiplyasrWpAnswerIndex + 1
            else:
                pcm_index = self.__curAnswerQueryIndex + 1
            reg_result = ''
            if self.__speech_type == globalVal.SpeechType.VOICE_PRINT:
                vpAccuracy = recog_decs
                vpResult = None
                if 'asr_sn' in asr_list:
                    vpResult = asr_list['asr_result']
                    if asr_list['vpAccuracy']:
                        vpAccuracy = self.__is_jinei + ':' + asr_list['vpAccuracy']
                asr_json = {
                    self.__samplesetName + '_' + str(pcm_index) + '.pcm': {
                        'desc': vpAccuracy,
                        'recg_result': vpResult
                    }
                }
                self.__queryResultDict.update(asr_json)
            else:
                asr_sn = None
                asr_corpus_id = None
                asr_role = None
                if not 'no ' in recog_decs and not 'ignore ' in recog_decs:
                    if 'asr_sn' in asr_list and (asr_list['asr_result'] or asr_list['asrFinalResult']):
                        if asr_list['asr_result']:
                            if self.__semantic_asr_mode:
                                reg_result = str(asr_list['semantics_index']) + '#' + str(asr_list['asr_result'])
                            else:
                                reg_result = asr_list['asr_result']
                        asr_sn = asr_list['asr_sn']
                        if 'asr_corpus_id' in asr_list:
                            asr_corpus_id = asr_list['asr_corpus_id']
                    elif recog_decs == '':
                        recog_decs = self.get_recognize_result(asr_list)
                    else:
                        recog_decs = self.get_recognize_result(asr_list) + '(' + recog_decs + ')'
                    if self.__role_asr_mode and asr_list['asrFinalResult'] is not None:
                        asr_role = asr_list['asr_role']
                flag = True  # 表示是否记录本条结果，针对车机split测试集没有结果
                if self.__speech_type == globalVal.SpeechType.SPLIT_ASR:
                    if asr_list['asr_result'] == 'NULL':
                        reg_result = ''
                    query_name = asr_list['file_name']
                    flag = False
                    for answer in self.__AnswerList:
                        if query_name == answer['file_name']:
                            flag = True
                    query_name = self.__samplesetName + '_' + query_name
                else:
                    query_name = self.__samplesetName + '_' + str(pcm_index) + '.pcm'
                if not flag:
                    return
                asr_json = {
                    query_name: {
                        'desc': recog_decs,
                        'recg_result': reg_result,
                        'sn': asr_sn,
                        'corpus_id': asr_corpus_id,
                        'asr_role': asr_role
                    }
                }
                self.__queryResultDict.update(asr_json)

