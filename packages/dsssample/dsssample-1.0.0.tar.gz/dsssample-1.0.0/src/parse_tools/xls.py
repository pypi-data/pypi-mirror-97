#!/usr/bin/ python
# -*- coding: utf-8 -*-
########################################################################
#
# Copyright (c) 2020 Baidu.com, Inc. All Rights Reserved
#
########################################################################

"""
    File: xls.py
    Author: dushuangshuang(dushuangshuang@baidu.com)
    Date: 2020/12/04 下午17:05
"""

import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import importlib, sys
importlib.reload(sys)
import globalVal


_RED_FONT = Font(bold=True, color='FF0000')
_GREEN_FONT = Font(bold=True, color='008000')
_BLUE_FONT = Font(bold=True, color='0000FF')
# _GRAY_FONT = Font(bold=True, color='BEBEBE')
_GRAY_FONT = Font(bold=True, color='778899')
_GRAY_FILL = PatternFill(fill_type='solid', fgColor='DCDCDC')
_GREEN_FILL = PatternFill(fill_type='solid', fgColor='00FF00')
_YELLOW_FILL = PatternFill(fill_type='solid', fgColor='FFFF00')
_ORANGE_FILL = PatternFill(fill_type='solid', fgColor='FFA500')
_LAKEBLUE_FILL = PatternFill(fill_type='solid', fgColor='48D1CC')
_PURPULE_FILL = PatternFill(fill_type='solid', fgColor='DDA0DD')
_BABYBLUE_FILL = PatternFill(fill_type='solid', fgColor='87CEFA')
_LIGHTWOOD_FILL = PatternFill(fill_type='solid', fgColor='E9C2A6')
_LIGHTBLUE_FILL = PatternFill(fill_type='solid', fgColor='C0D9D9')
_LIGHTPINK_FILL = PatternFill(fill_type='solid', fgColor='BC8F8F')
_MEDIUM_BORDER = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
_CENTER_ALIGNMENT = Alignment(horizontal='center',vertical='center',wrap_text=True)
_RIGHT_ALIGNMENT = Alignment(horizontal='right',vertical='center',wrap_text=True)
    

def write_content(sheet, rows, cols, content, fontMat=-1):
    """excel写入数据
    rows/cols 行数/列数，可以是int或数组（数组表示需要合并表格）
    content 要写入的值
    fontMat (不居中) 0:底色飘黄，2:底色飘灰，3:底色飘绿，4:字体飘红，5:字体飘绿，6:字体飘蓝，31:字体飘灰
            (居中显示) 1:底色飘黄，7:底色土橘，8:底色湖蓝，9:底色飘紫，10:底色淡蓝，11:底色浅木，12:底色飘灰，13:底色浅蓝，14:底色暗粉
            (居中显示) 20:底色浅木，21:底色浅蓝，22:底色暗粉
    """
    if sheet is None:
        return
    merge_flag, start_rows, end_rows, start_cols, end_cols = get_merge_info(rows, cols)
    # 给表格写入值
    if merge_flag:
        sheet.merge_cells(start_row=start_rows, start_column=start_cols, end_row=end_rows, end_column=end_cols)
    cell_name = sheet.cell(row=start_rows, column=start_cols, value=content)
        
    # 设置表格格式
    if fontMat != -1:
        if fontMat == 0:
            cell_name.fill = _YELLOW_FILL
        elif fontMat == 4:
            cell_name.font = _RED_FONT
        elif fontMat == 5:
            cell_name.font = _GREEN_FONT
        elif fontMat == 6:
            cell_name.font = _BLUE_FONT
        elif fontMat == 31:
            cell_name.font = _GRAY_FONT
        else:
            if fontMat == 2:
                cell_name.fill = _GRAY_FILL
            elif fontMat == 3:
                cell_name.fill = _GREEN_FILL
            else:
                # cell_name.border = _MEDIUM_BORDER
                cell_name.alignment = _CENTER_ALIGNMENT
                if fontMat == 1:
                    cell_name.fill = _YELLOW_FILL
                elif fontMat == 7:
                    cell_name.fill = _ORANGE_FILL
                elif fontMat == 8:
                    cell_name.fill = _LAKEBLUE_FILL
                elif fontMat == 9:
                    cell_name.fill = _PURPULE_FILL
                elif fontMat == 10:
                    cell_name.fill = _BABYBLUE_FILL
                elif fontMat == 12:
                    cell_name.fill = _GRAY_FILL
                elif fontMat == 20 or fontMat == 11:
                    cell_name.fill = _LIGHTWOOD_FILL
                elif fontMat == 21 or fontMat == 13:
                    cell_name.fill = _LIGHTBLUE_FILL
                elif fontMat == 22 or fontMat == 14:
                    cell_name.fill = _LIGHTPINK_FILL


def write_content_with_color(sheet, row, col, content, reverse=False):
    """写数据
    值小于0时字体飘红，大于0时字体飘绿
    """
    if sheet is None:
        return
    if '%' in str(content):
        gap = float(content.split('%')[0])
    elif 'Nan' in str(content):
        gap = 'Nan'
    else:
        gap = int(content)
    formats = 31
    if gap != 'Nan':
        # if gap > 0:
        #     content = '+' + str(content)
        if (reverse and gap > 0) or (not reverse and gap < 0):
            formats = 4
        elif (reverse and gap < 0) or (not reverse and gap > 0):
            formats = 5
    write_content(sheet, row, col, content, formats)


def write_title(sheet, title_list, row_count, start_col=1, merge_row=0, merge_col=0, font=1):
    """写表头
    在指定row_count，向sheet写入title_list
    """
    if sheet is None:
        return
    col = start_col
    for i in range(0, len(title_list)):
        write_content(sheet, [row_count, row_count + merge_row], [col, col + merge_col], title_list[i], font)
        col = col + merge_col + 1

# def set_border(ws, cell_range):
#     rows = ws.range(cell_range)
#     for row in rows:
#         row[0].style.borders.left.border_style = Border.BORDER_THIN
#         row[-1].style.borders.right.border_style = Border.BORDER_THIN
#     for c in rows[0]:
#         c.style.borders.top.border_style = Border.BORDER_THIN
#     for c in rows[-1]:
#         c.style.borders.bottom.border_style = Border.BORDER_THIN


def get_merge_info(rows, cols):
    """表格信息读取，是否merge写入
    """
    merge_flag = False
    if isinstance(rows, list):
        start_rows = rows[0]
        end_rows = rows[1]
        merge_flag = True
    else:
        start_rows = rows
        end_rows = rows
    if isinstance(cols, list):
        start_cols = cols[0]
        end_cols = cols[1]
        merge_flag = True
    else:
        start_cols = cols
        end_cols = cols

    return merge_flag, start_rows, end_rows, start_cols, end_cols


def set_fill(sheet, rows, cols, fontMat):
    """表格设置
    从direction.txt读取测试指标说明，并写入 统计结果.xlsx
    """
    if sheet is None:
        return
    if fontMat == 2:
        fill = _GRAY_FILL
    else:
        return
    merge_flag, start_rows, end_rows, start_cols, end_cols = get_merge_info(rows, cols)
    for r in range(start_rows, end_rows + 1):
        for c in range(start_cols, end_cols + 1):
            sheet.cell(row=r, column=c).fill = fill
            # cell_name = sheet.cell(row=r, column=c)
            # cell_name.fill = fill
            # cell_name.alignment = _RIGHT_ALIGNMENT


def streamReportDirections(result_workbook):
    """测试指标说明
    从direction.txt读取测试指标说明，并写入 统计结果.xlsx
    """
    directions = result_workbook.create_sheet('统计结果&指标说明')
    file_path = os.path.join(globalVal._TOOL_DIR, 'direction.txt')
    if os.path.exists(file_path):
        row_count = 0
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                row_count += 1
                if '[' in line:
                    item = line.strip().split(']', 1)
                    write_content(directions, row_count, 1, str(item[0]) + ']', 3)
                    if len(item) > 1:
                        cell_name = directions.cell(row=row_count, column=2, value=item[1])
                        cell_name.hyperlink = str(item[1])
                elif line.strip() != '':
                    item = line.split(': ', 1)
                    write_content(directions, row_count, 1, item[0])
                    if len(item) > 1:
                        write_content(directions, row_count, 2, item[1])
        f.close()
    result_workbook.save(globalVal._RESULT_WORKBOOK_NAME)
