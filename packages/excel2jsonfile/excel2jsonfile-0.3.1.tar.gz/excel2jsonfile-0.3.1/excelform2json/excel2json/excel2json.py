import csv
import json
from os import path, makedirs
from os.path import basename
from datetime import datetime

from openpyxl.worksheet.cell_range import CellRange

from excelform2json.utils import messages, excel


class Excel2JSON:
    start_time = datetime.now()

    def __init__(self, configuration_file="resources/config/excelform2json_config.json"):
        self.excel_location = "input/"
        self.json_directory = "out/"
        self.json_file_prefix = "generated_"
        self.result = messages.message["ok"]
        self.work_book = None
        self.excel_sheets = []
        self.excel_utils = excel.Excel()

        try:
            with open(configuration_file) as config:
                data = json.load(config)
                self.excel_location = data["excel_location"]
                if "sheets" in data:
                    self.excel_sheets = data["sheets"]
                if "json_directory" in data:
                    self.json_directory = data["json_directory"]
                if "json_file_prefix" in data:
                    self.json_file_prefix = data["json_file_prefix"]
                if self.json_file_prefix == "TIMESTAMP":
                    self.json_file_prefix = self.start_time.isoformat(timespec="microseconds").replace(':', '-')
                self.json_file_prefix += "--"
        except FileNotFoundError:
            print("Configuration file >" + configuration_file + "< not found.")
            self.result = messages.message["main_config_not_found"]

    def create_output_files(self):
        overall_result = messages.message["ok"]

        return overall_result

    def search_value_in_row_index(self, search_string, row=1):
        return self.excel_utils.search_value_in_row_index(work_sheet=self.work_sheet, search_string=search_string,
                                                          row=row)

    def search_value_in_column_index(self, search_string, column, give_column, give_column2=None,
                                     name_column1=None, name_column2=None, give_list=False):
        # TODO: Make more generic and flexible
        # Note: give_column2 is only valid in combination with give_list=True

        # print("Searching for", search_string, "in column", column)
        values = []
        for row in self.work_sheet.iter_rows(column):
            for cell in row:
                # print(cell.row, cell.column, cell.value)
                if cell.value == search_string:
                    # print(cell.value)
                    if not give_list:
                        # we found it and we should not provide an entire list, so return to caller
                        return str(self.work_sheet.cell(row=cell.row, column=give_column).value).replace('"',
                                                                                                         '\"').replace(
                            '\\n', ' ')
                    else:
                        if give_column2 is None:
                            value = self.work_sheet.cell(row=cell.row, column=give_column).value
                            if value is None:
                                value = ""
                            else:
                                value = value.replace('"', '\"')
                            values.append(value)
                        else:
                            value1 = self.work_sheet.cell(row=cell.row, column=give_column).value
                            if value1 is None:
                                value1 = ""
                            else:
                                value1 = value1.replace('"', '\"')
                            value2 = self.work_sheet.cell(row=cell.row, column=give_column2).value
                            if value2 is None:
                                value2 = ""
                            else:
                                value2 = value2.replace('"', '\"')
                            col_dict = {name_column1: value1
                                , name_column2: value2}
                            values.append(col_dict)
        if give_list:
            nr = 0
            out_list = []
            for value in values:
                already_exists = False
                for out in out_list:
                    if value["attribute_name"].lower() == out["attribute_name"].lower():
                        already_exists = True
                        break
                if already_exists:
                    pass
                else:
                    out_list.append(value)
            return out_list
        #    return list(set(values))
        else:
            return None

    def write_dict_to_file(self, out_file, result_list):
        result = messages.message["ok"]
        # json
        file_name = path.join(self.json_directory, self.json_file_prefix + out_file + ".json")
        try:
            with open(file_name, "w", encoding="utf8", newline='') as json_file:
                json_file.write(json.dumps(result_list, indent=4))
        except IOError:
            result = messages.message["write_error"]
            result["info"] = "Error creating or writing dictionary to json output file >" + file_name + "<."
            print(result["info"])

        return result

    def process_sheet(self,work_book, work_sheet, sheet):
        result, cells = self.get_named_cells(work_book, work_sheet, sheet)
        if result["code"] != "OK":
            return result

        self.write_dict_to_file(sheet["name"], cells)

        return result

    def get_named_cells(self, work_book, work_sheet, sheet):
        sheet_content = []
        result = messages.message["ok"]

        # get all defined names
        defined_names = work_book.defined_names.definedName
        for defined_name in defined_names[:]:
            cells = []
            parameters = {}
            # print(defined_name)
            # print(defined_name.name, defined_name.attr_text)
            sheet_name, cell_range = list(defined_name.destinations)[0]
            if cell_range is None or len(cell_range) == 0:
                cells.append(dict(cell="#ERR", value="#UNKNOWN"))
                print("There is something wrong with >%s<. Probably an error in the Excel, e.g. invalid REF" % defined_name.name)
            else:
                cr = CellRange(cell_range)
                width = cr.max_col - cr.min_col
                length = cr.max_row - cr.min_row
                if width > 0 or length >0:
                    # cells = [cell.value for cell in [row[0] for row in work_sheet[cell_range]]]
                    for row in work_sheet[cell_range]:
                        for c in row:
                            cells.append(dict(cell=c.coordinate, value=c.value))
                        parameters[defined_name] = cells
                else:
                    cells.append(dict(cell=cell_range, value=work_sheet[cell_range].value))
            sheet_content.append(dict(sheet=sheet["name"],name=defined_name.name, content=cells))
        # q1_1 = work_sheet[list(work_book.defined_names['Question1.1'].destinations)[0][1]].value
        # print("q1_1 is", q1_1)
        print("#named cells:", len(sheet_content))

        return result, sheet_content

    def main(self, excel_file=None, excel_sheet=None, structure=None):
        main_result = messages.message["ok"]
        if self.result["code"] != "OK":
            return self.result
        if excel_sheet is not None:
            self.excel_sheets.add(dict(name=excel_sheet, structure=structure))
        makedirs(self.json_directory, exist_ok=True)
        print(self.json_directory)

        for sheet in self.excel_sheets:
            main_result, work_book, work_sheet = self.excel_utils.read_excel(file=excel_file, sheet=sheet["name"])
            if main_result["code"] != "OK":
                return main_result
            sheet_result = self.process_sheet(work_book, work_sheet, sheet=sheet)
            if sheet_result["code"] != "OK":
                main_result = sheet_result

        return main_result


if __name__ == "__main__":
    excel2json = Excel2JSON(configuration_file="resources/config/excelform2json_config.json")
    result = excel2json.main(excel_file="input/excels/Test1 questionnaire privacy.xlsx")
    print("result:", result)
    # if result["code"] != "OK":
    #    exit(1)
    # else:
    #    exit(0)
