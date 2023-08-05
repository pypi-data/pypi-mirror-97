# -------------------------------------------------------------------------------
# Licence:
# Copyright (c) 2012-2017 Luzzi Valerio 
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES
# OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
# HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
# WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
#
#
# Name:        excel.py
# Purpose:
#
# Author:      Luzzi Valerio
#
# Created:     14/08/2017
# -------------------------------------------------------------------------------
import openpyxl
from .datatypes import *

TYPE_STRING = 's'
TYPE_FORMULA = 'f'
TYPE_NUMERIC = 'n'
TYPE_BOOL = 'b'
TYPE_NULL = 'n'
TYPE_INLINE = 'inlineStr'
TYPE_ERROR = 'e'
TYPE_FORMULA_CACHE_STRING = 'str'

XLRD = 0
OPENPYXL = 1

DATATYPES = {
    xlrd.XL_CELL_EMPTY: TYPE_NULL,
    xlrd.XL_CELL_TEXT: TYPE_STRING,
    xlrd.XL_CELL_NUMBER: TYPE_NUMERIC,
    xlrd.XL_CELL_DATE: TYPE_STRING,
    xlrd.XL_CELL_BOOLEAN: TYPE_BOOL,
    xlrd.XL_CELL_ERROR: TYPE_ERROR,
    xlrd.XL_CELL_BLANK: TYPE_NULL,

    TYPE_NULL: xlrd.XL_CELL_EMPTY,
    TYPE_STRING: xlrd.XL_CELL_TEXT,
    TYPE_NUMERIC: xlrd.XL_CELL_NUMBER,
    TYPE_BOOL: xlrd.XL_CELL_BOOLEAN,
    TYPE_ERROR: xlrd.XL_CELL_ERROR
}


class Workbook:

    def __init__(self, filename='', type='xls', mode='r'):
        """
        constructor
        """
        self.lib = OPENPYXL if type == "xlsx" else XLRD

        if not filename:
            if self.lib == OPENPYXL:
                self.wb = openpyxl.Workbook()
                sheets = self.wb.worksheets
                for sheet in sheets:
                    self.wb.remove_sheet(sheet)

            elif self.lib == XLRD:
                self.wb = xlwt.Workbook()

        elif mode == "r":
            if self.lib == OPENPYXL:
                self.wb = openpyxl.load_workbook(filename, read_only=True)
            elif self.lib == XLRD:
                self.wb = xlrd.open_workbook(filename, on_demand=True)

        elif mode == "rw":
            if self.lib == OPENPYXL:
                self.wb = openpyxl.load_workbook(filename)
            elif self.lib == XLRD:
                rb = xlrd.open_workbook(filename, on_demand=True)
                self.wb = copy(rb)

    def open_workbook(self, filename):
        pass

    def safename(self, sheetname):
        return re.sub(r'[\:\\/\?\*\[\]]', "_", sheetname[:31])

    def add_sheet(self, sheet_name):
        """
        add_sheet
        :param sheet_name:
        :return:
        """
        if self.lib == OPENPYXL:
            sheet = self.wb.create_sheet(title=sheet_name)
        elif self.lib == XLRD:
            sheet = self.wb.add_sheet(sheet_name)
        return WorkSheet(sheet)

    def create_sheet(self, sheet_name):
        return self.add_sheet(self.safename(sheet_name))


    def sheets(self):
        """
        sheets
        :return:
        """
        if self.lib == OPENPYXL:
            return [WorkSheet(sheet) for sheet in self.wb]
        elif self.lib == XLRD:
            return [WorkSheet(sheet) for sheet in self.wb.sheets()]
        return []

    def sheet_names(self):
        """
        sheet_names
        :return:
        """
        if self.lib == OPENPYXL:
            return self.wb.get_sheet_names()
        elif self.lib == XLRD:
            return self.wb.sheet_names()
        return []

    def get_sheets_by_names(self):
        return self.sheet_names()

    def sheet_by_index(self, index):
        """
        sheet_by_index
        :param index:
        :return:
        """
        if self.lib == OPENPYXL:
            ws = self.wb.sheetnames[index]
        elif self.lib == XLRD:
            ws = self.wb.sheet_by_index(index)
        else:
            ws = None
        return WorkSheet(ws) if ws else None

    def get_sheet(self, index):
        return self.sheet_by_index(index)

    def sheet_by_name(self, sheet_name):
        """
        get_sheet_by_name
        :param sheet_name:
        :return:
        """
        if self.lib == OPENPYXL:
            ws = self.wb.get_sheet_by_name(sheet_name)
        elif self.lib == XLRD:
            ws = self.wb.sheet_by_name(sheet_name)
        else:
            ws = None
        return WorkSheet(ws) if ws else None

    def get_sheet_by_name(self, sheet_name):
        return self.sheet_by_name(sheet_name)

    def save(self, filename, attempts=3, interval=10):
        """
        save
        :param filename:
        :return:
        """
        success = False
        count = 0
        while not success and count < attempts:
            try:
                if self.lib == OPENPYXL:
                    self.wb.save(filename)
                elif self.lib == XLRD:
                    self.wb.save(filename)
                success = True
            except Exception as ex:
                count += 1
                print(ex)
                print(
                            "May be the file:%s already exists and it is open in Excel or you have not the right to write." % (
                        filename))
                print("In the first case close the file. I will retry saving in %s seconds from now." % (interval))
                time.sleep(interval)

        if not success:
            print("Unable to save %s." % filename)
            raise Exception(ex.message)

class WorkSheet:
    def __init__(self, sheet):
        if isinstance(sheet, (xlrd.sheet.Sheet, xlwt.Worksheet)):
            self.lib = XLRD
        else:
            self.lib = OPENPYXL
        self.ws = sheet
        self._cached_easyxf = {}

    def cell(self, i, j, value=None, stylecss=None):
        if self.lib == OPENPYXL:
            if value == None:
                cl = self.ws.cell(row=i + 1, column=j + 1)
            else:
                # write mode
                cl = self.ws.cell(row=i + 1, column=j + 1, value=value)
                if stylecss:
                    cl.style = self.parseStyleCSS(stylecss)

        elif self.lib == XLRD:
            if value == None:
                cl = self.ws.cell(rowx=i, colx=j)
            else:
                if stylecss:
                    cl = self.ws.write(i, j, value, style=self.parseStyleCSS(stylecss))
                else:
                    cl = self.ws.write(i, j, value)
        else:
            cl = None

        return cl

    def cell_value(self, i, j):
        cl = self.cell(i, j)
        return cl.value if cl else None

    def cell_type(self, i, j):
        cl = self.cell(i, j)
        return cl.ctype if cl else None

    def write(self, i, j, value):
        if self.lib == OPENPYXL:
            self.ws.cell(i, j, value)
        elif self.lib == XLRD:
            self.ws.write(i, j, value)

    def merged_cells(self):
        if self.lib == OPENPYXL:
            return self.ws.merged_cell_ranges
        elif self.lib == XLRD:
            return self.ws.merged_cells
        return []

    def get_rows(self):
        if self.lib == OPENPYXL:
            return self.ws.rows
        elif self.lib == XLRD:
            return self.ws.get_rows()
        return []

    def get_columns(self):
        if self.lib == OPENPYXL:
            return self.ws.columns
        elif self.lib == XLRD:
            return self.ws.get_columns()
        return []

    def parseColor(self, text):
        """
        parseColor
        """
        if isstring(text):
            if text in xlwt.Style.colour_map:
                color_idx = xlwt.Style.colour_map[text] - 8
                color = ("%08x" % xlwt.Style.excel_default_palette_b8[color_idx])[:6]
        else:
            color = "000000"

        return color

    def parseStyleCSS(self, text):

        # text = "font: bold on; align: wrap on, vert centre, horiz center"

        if text in self._cached_easyxf:
            return self._cached_easyxf[text]

        if self.lib == OPENPYXL:
            style = NamedStyle(name=text)

            css = mapify(text.lower(), ";", ":")

            for key in css:
                css[key] = mapify(css[key], ",", " ")

            if "font" in css:
                props = css["font"]
                style.font = Font()
                if "name" in props:
                    style.font.name = props["name"]
                if "size" in props:
                    style.font.size = val(props["size"])
                if "bold" in props:
                    style.font.bold = parseBool(props["bold"])
                if "italic" in props:
                    style.font.italic = parseBool(props["italic"])
                if "vertalign" in props:
                    style.font.vertAlign = props["vertalign"]
                if "underline" in props:
                    style.font.underline = props["underline"]
                if "strike" in props:
                    style.font.strike = props["strike"]
                if "color" in props:
                    style.font.color = self.parseColor(props["color"])

            if "pattern" in css:
                props = css["pattern"]
                style.fill = PatternFill("solid")
                if "pattern" in props:
                    style.fill.fill_type = props["pattern"]
                if "fore_colour" in props:
                    style.fill.fgColor = self.parseColor(props["fore_colour"])

            if "alignment" in css:
                props = css["alignment"]
                style.alignment = Alignment()
                if "horizontal" in props:
                    style.alignment.horizontal = props["horizontal"]
                if "vertical" in props:
                    style.alignment.vertical = props["vertical"]
                if "wrap" in props:
                    style.alignment.wrap_text = parseBool(props["wrap"])

        elif self.lib == XLRD:
            style = xlwt.easyxf(text)

        self._cached_easyxf[text] = style
        return style


if __name__ == "__main__":
    filename = "./tests/xls/Dati_stazione_superficiale_test.xlsx"

    from gecosistema_lite import *

    db = SqliteDB.QuickTest()
    db.execute("""ALTER TABLE test ADD COLUMN [style-descr];""")
    db.ExportTo("./tests/xls/test.xlsx")
    db.close()
