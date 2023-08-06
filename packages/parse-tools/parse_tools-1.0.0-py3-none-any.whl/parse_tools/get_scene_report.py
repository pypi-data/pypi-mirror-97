#!/usr/bin/ python
# -*- coding: utf-8 -*-
########################################################################
#
# Copyright (c) 2020 Baidu.com, Inc. All Rights Reserved
#
########################################################################

"""
    File: get_scene_report.py
    Author: dushuangshuang(dushuangshuang@baidu.com)
    Date: 2020/01/19 上午12:05
"""

import os
import time
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import json 
import subprocess
import logging
import copy
import importlib, sys
importlib.reload(sys)

import globalVal
import speech_parse
import get_contrast
import xls
import get_conclusion as conclusion


if globalVal._DEBUG_MODE:
    logging.basicConfig(level=logging.DEBUG)


def updateInfo(version_nums):
    """
        测试集、SDK版本去重
    """
    tmplist = []
    for audio_type in globalVal._SAMPLE_LIST:
        tmplist += sorted(globalVal._SAMPLE_LIST[audio_type])
    globalVal._SAMPLE_LIST = tmplist
    if version_nums == 1:
        globalVal._PARAM['CONTRAST_WITH_ALL'] = True
    print('[Scene List]', globalVal._SAMPLE_LIST)
    print('[Version List]', globalVal._VERSION_LIST)


def get_audio_type(scene_name):
    """
        得到测试类型
    """
    if 'jike' in scene_name:
        audio_type = scene_name.split('_')[1]
    else:
        audio_type = scene_name.split('_')[0]
    if 'split' in scene_name:
        audio_type = 'splitasr'
    if 'offlineasr' in scene_name:
        audio_type = 'asr'
    return audio_type


def get_substraction(value_a, value_b):
    """
        [差值计算] 得到两个值的差
    Args:
        value_a : [string] 减数
        value_b : [string] 被减数
    Returns:
        value : [string] 差值
    """
    if '%' in str(value_a) and '%' in str(value_b):
        value = '%.3f%%' % (float(value_a.strip('%')) - float(value_b.strip('%')))
        return str(value)
    elif str(value_a) == 'Nan' or str(value_b) == 'Nan':
        return 'Nan'
    else:
        try:
            if isinstance(value_a, float):
                return round(float(value_a) - float(value_b), 3)
            else:
                return int(value_a) - int(value_b)
        except TypeError:
            return 'Nan'


def write_wak_detail_info_by_col(sheet_name, row, col, wak_list, curr_version):
    """写入数据：每个唤醒词的详细指标
    """
    # [1] 写入唤醒详情
    for wpWords in curr_version['wp_status']:
        if wpWords == 'summary':
            continue
        c_idx = 1
        xls.write_content(sheet_name, row, col, wpWords)
        for key in wak_list:
            xls.write_content(sheet_name, row, col + c_idx, curr_version['wp_status'][wpWords][key])
            c_idx += 1
        row += 1
    # [2] 写入唤醒汇总
    xls.write_content(sheet_name, row, col, '汇总', 7)
    c_idx = 1
    for key in wak_list:
        xls.write_content(sheet_name, row, col + c_idx, curr_version['wp_status']['summary'][key], 7)
        c_idx += 1
    row += 1

    if len(curr_version['wp_false']) > 1:
        # [3] 写入误唤醒详情
        xls.write_content(sheet_name, row, col, '唤醒词', 9)
        xls.write_content(sheet_name, row, col + 1, '误报次数', 9)
        row += 1
        for wpWords in curr_version['wp_false']:
            if wpWords == 'summary':
                continue
            xls.write_content(sheet_name, row, col, wpWords)
            xls.write_content(sheet_name, row, col + 1, curr_version['wp_false'][wpWords])
            row += 1
        # [4] 写入误唤醒汇总
        xls.write_content(sheet_name, row, col, '汇总', 7)
        xls.write_content(sheet_name, row, col + 1, curr_version['wp_false']['summary'], 7)
        row += 1

    return row


def write_wp_report_title_by_col(sheet_name, wak_words):
    """写入统计结果.xlsx的title
    """
    words_num = len(wak_words)
    if words_num == 0:
        words_num = 1
    wak_col = [0, 0]
    for scene_name in globalVal._SAMPLE_LIST:
        if (not 'wp' in scene_name and not 'asr' in scene_name) or 'wp_dci' in scene_name:
            continue
        if 'kws' in scene_name:
            wak_col[1] += 1
        else:
            wak_col[0] += 1

    v_idx = 0
    for version in globalVal._VERSION_LIST:
        xls.write_content(sheet_name, [3 + v_idx * words_num, 2 + (v_idx + 1) * words_num], 1, version)
        xls.write_content(sheet_name, [3 + v_idx * words_num, 2 + (v_idx + 1) * words_num], 3 + wak_col[0], version)
        col_1 = 3
        col_2 = col_1 + wak_col[0] + 2
        for scene_name in globalVal._SAMPLE_LIST:
            if (not 'wp' in scene_name and not 'asr' in scene_name) or 'wp_dci' in scene_name:
                continue
            if 'kws' in scene_name:
                xls.write_content(sheet_name, 2, col_2, scene_name.split('_')[3], 9)
                col_2 += 1
            else:
                xls.write_content(sheet_name, 2, col_1, scene_name.split('_')[3], 9)
                col_1 += 1
        r_idx = 0
        for wp_word in wak_words:
            xls.write_content(sheet_name, 3 + v_idx * words_num + r_idx, 2, wp_word)
            r_idx += 1
        xls.write_content(sheet_name, 3 + v_idx * words_num, 4 + wak_col[0], '总唤醒率')
        v_idx += 1
    xls.write_content(sheet_name, 2, 1, '版本')
    xls.write_content(sheet_name, 2, 2, '唤醒词', 9)
    xls.write_content(sheet_name, 2, 3 + wak_col[0], '版本')
    xls.write_content(sheet_name, 2, 4 + wak_col[0], '指令词', 9)
    # 对比差值
    xls.write_content(sheet_name, [3 + v_idx * words_num, 2 + (v_idx + 1) * words_num], 1, '差值')
    xls.write_content(sheet_name, [3 + v_idx * words_num, 2 + (v_idx + 1) * words_num], 3 + wak_col[0], '差值')
    r_idx = 0
    for wp_word in wak_words:
        xls.write_content(sheet_name, 3 + v_idx * words_num + r_idx, 2, wp_word)
        r_idx += 1
    xls.write_content(sheet_name, 3 + v_idx * words_num, 4 + wak_col[0], '总唤醒率')
    return wak_col


def write_wak_detail_info_by_row(sheet_name, row, col, wak_col, wak_words, wak_list, curr_version, wak_mode):
    """写入数据：每个唤醒词的详细指标
    """
    wak_info_col, false_wak_info_col = wak_col
    true_wak_words, false_wak_words = wak_words
    if wak_mode == 3:
        wid_2 = len(globalVal._QWQD_WAK_PERFORM_LIST)
    else:
        wid_2 = len(globalVal._R_XD_WAK_LIST)
    wid_3 = len(globalVal._R_FALSE_WP_LIST)
    if wak_info_col:
        w_idx = 0
        for wpWords in true_wak_words:
            if wpWords == 'summary':
                continue
            r_idx = 0
            if wak_mode == 3:
                whisperInfo = curr_version['wp_status'][wpWords]['whisperInfo']
                for mode in whisperInfo:
                    c_idx = 1
                    if mode == 0:
                        xls.write_content(sheet_name, row + r_idx, col + w_idx * wid_2 + c_idx, '正常')
                    elif mode == 1:
                        xls.write_content(sheet_name, row + r_idx, col + w_idx * wid_2 + c_idx, '气声')
                    else:
                        xls.write_content(sheet_name, row + r_idx, col + w_idx * wid_2 + c_idx, '轻声')
                    c_idx += 1
                    for key in wak_list:
                        xls.write_content(sheet_name, row + r_idx, col + w_idx * wid_2 + c_idx, whisperInfo[mode][key])
                        c_idx += 1
                    r_idx += 1
            else:
                c_idx = 1
                for key in wak_list:
                    if wpWords in curr_version['wp_status']:
                        xls.write_content(sheet_name, row, col + w_idx * wid_2 + c_idx, 
                            curr_version['wp_status'][wpWords][key])
                    else:
                        xls.write_content(sheet_name, row, col + w_idx * wid_2 + c_idx, 0)
                    c_idx += 1
                r_idx += 1
            xls.write_content(sheet_name, [row, row + r_idx - 1], col + w_idx * wid_2, wpWords)
            w_idx += 1
    if false_wak_info_col:
        w_idx = 0
        for wpWords in false_wak_words:
            if wpWords == 'summary':
                continue
            xls.write_content(sheet_name, row, col + wak_info_col + w_idx * wid_3, wpWords)
            if wpWords in curr_version['wp_false']:
                xls.write_content(sheet_name, row, col + wak_info_col + w_idx * wid_3 + 1, 
                    curr_version['wp_false'][wpWords])
            else:
                xls.write_content(sheet_name, row, col + wak_info_col + w_idx * wid_3 + 1, 0)
            w_idx += 1

def write_wp_report_title_by_row(audio_type, sheet_name, row, wak_words, wak_mode):
    """写入统计结果.xlsx的title
    """
    col = 3
    wak_info_col = 0
    false_wak_info_col = 0
    true_wak_words, false_wak_words = wak_words

    if true_wak_words:
        if wak_mode == 2:
            wak_tile = globalVal._R_XD_WAK_LIST
        else:
            wak_tile = globalVal._QWQD_WAK_PERFORM_LIST
        wak_info_col = len(wak_tile) * len(true_wak_words)
    if wak_mode > 0 and len(false_wak_words) > 0:
        false_wak_tile = globalVal._R_FALSE_WP_LIST
        false_wak_info_col = len(false_wak_tile) * len(false_wak_words)

    xls.write_content(sheet_name, [row, row], 
        [1, col + wak_info_col + false_wak_info_col - 1], str(audio_type) + '（唤醒效果详情）', 3)
    xls.write_content(sheet_name, [row + 1, row + 2], 1, '测试集', 1)
    xls.write_content(sheet_name, [row + 1, row + 2], 2, '测试版本', 1)
    if wak_info_col:
        xls.write_content(sheet_name, [row + 1, row + 1], 
            [col, col + len(true_wak_words) * len(wak_tile) - 1], 
            '唤醒详细信息', 7)
        for i in range(len(true_wak_words)):
            xls.write_title(sheet_name, wak_tile,
                row + 2, col + i * len(wak_tile), 0, 0, 9)
    if false_wak_info_col:
        xls.write_content(sheet_name, [row + 1, row + 1], 
            [col + wak_info_col, \
            col + wak_info_col + len(false_wak_words) * len(false_wak_tile) - 1], 
            '误唤醒详细信息', 8)
        for i in range(len(false_wak_words)):
            xls.write_title(sheet_name, false_wak_tile, 
                row + 2, col + wak_info_col + i * len(false_wak_tile), 0, 0, 10)

    return [wak_info_col, false_wak_info_col]


def get_speed_list(audio_type):
    """根据测试类型、得到相应的速度列表
    """
    wak_title = ['唤醒速度统计条数', '唤醒时间']
    if 'pureasr' == audio_type or 'splitasr' == audio_type:
        speed_list = globalVal._SPEED_TYPE['asr']
        title_list = copy.deepcopy(globalVal._R_SPEED_LIST[audio_type])
    elif 'asr' in audio_type:
        speed_list = copy.deepcopy(globalVal._SPEED_TYPE['wp'])
        speed_list.update(globalVal._SPEED_TYPE['asr'])
        title_list = wak_title + copy.deepcopy(globalVal._R_SPEED_LIST[audio_type])
    else:
        speed_list = globalVal._SPEED_TYPE['wp']
        title_list = wak_title
    return title_list, speed_list


def write_speed_analyse_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：速度专项
    """
    title_list, speed_list = get_speed_list(get_audio_type(sample))
    # 写汇总
    if '唤醒速度统计条数' in title_list:
        xls.write_content(sheet_name, row, col, curr_version['wp_count'])
        xls.write_content(sheet_name, row, col + 1, curr_version['wp_time'])
        if len(base_version):
            xls.write_content_with_color(sheet_name, row + 1, col + 1, 
                get_substraction(curr_version['wp_time'], base_version['wp_time']), True)
        col += 2
    if '识别速度统计条数' in title_list:
        write_asr_speed_data(row, col, curr_version, sheet_name, sample, vers, base_version)
        col += len(globalVal._R_SPEED_LIST['asr'])
    # 写详情：中位数、90分位数、标准差、最小值、最大值
    idx = 0
    for key in speed_list:
        if 'wp_speed_list' == key:
            if 'wp_speed_list' in curr_version:
                speed_list = curr_version['wp_speed_list']
            else:
                speed_list = curr_version['wp_status']['summary']['wp_speed_list']
        else:
            if not key in curr_version['asr_speed_list']:
                logging.error('sdk:%s, audio:%s, could not find key:%s', vers, sample, key)
                continue
            speed_list = curr_version['asr_speed_list'][key]
        if len(speed_list):
            speed_data = speech_parse.cal_speed_performance(speed_list)
            for k in range(len(globalVal._SPEED_ANALYSE_LIST)):
                xls.write_content(sheet_name, row, col + idx, speed_data[k])
                idx += 1
        else:
            idx += len(globalVal._SPEED_ANALYSE_LIST)
    if len(base_version):
        xls.set_fill(sheet_name, row + 1, [6, col + idx - 1], 2)


def write_acu_analyse_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：识别准确率专项
    """
    idx = 0
    for key in ['wer', 'ser', 'sub_e', 'del_e', 'in_e']:
        if not key in curr_version:
            continue
        xls.write_content(sheet_name, row, col + idx, curr_version[key])
        if len(base_version):
            xls.write_content_with_color(sheet_name, row + 1, col + idx, 
                get_substraction(curr_version[key], base_version[key]))
        idx += 1
    if len(base_version):
        xls.set_fill(sheet_name, row + 1, [6, col + idx - 1], 2)


def get_err_flag(k, audio_type):
    """根据测试类型audio_type、当前错误码类型k，判断该错误码是否用于准出
    """
    if '全部' in k:
        return False
    else:
        if '首次' in k:
            flag_1 = True
        else:
            flag_1 = False
        if 'mutiplyasr' in audio_type or 'puerasr' in audio_type:
            flag_2 = False
        else:
            flag_2 = True
    return flag_1 == flag_2


def write_err_analyse_data(row, col, curr_version, sheet_name, sample, vers, err_list, base_version={}):
    """写入数据：识别准确率专项
    """
    idx = 0
    for k in err_list:
        curr_data = curr_version['asr_error_code_info'][k]
        xls.write_content(sheet_name, row, col + idx, curr_data['denominator'])
        for err_code in err_list[k]:
            if err_code in curr_data['info']:
                curr_count = curr_data['info'][err_code]
                xls.write_content(sheet_name, row, col + idx + 1, curr_count)
            else:
                curr_count = 0
            if len(base_version):
                curr_err_rate = "%.3f%%" % float(curr_count * 100 / curr_data['denominator'])
                if err_code in base_version['asr_error_code_info'][k]['info']:
                    base_err_rate = "%.3f%%" % float(base_version['asr_error_code_info'][k]['info'][err_code] \
                        * 100 / base_version['asr_error_code_info'][k]['denominator'])
                    if curr_count == 0:
                        xls.write_content(sheet_name, row, col + idx + 1, curr_count)
                elif curr_count != 0:
                    # 基线版本没有出现err_code时记次数为0，并记录两个版本的diff
                    base_err_rate = '0.0%'
                    xls.write_content(sheet_name, row - 1, col + idx + 1, 0)
                err_diff = get_substraction(curr_err_rate, base_err_rate)
                xls.write_content_with_color(sheet_name, row + 1, col + idx + 1, err_diff, True)
                curr_version['asr_error_code_info'][k]['diff'].update({err_code: err_diff})
            idx += 1
        idx += 1
    if len(base_version):
        xls.set_fill(sheet_name, row + 1, [6, col + idx - 1], 2)


def write_wp_result_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：唤醒指标
    """
    xls.write_content(sheet_name, row, col, curr_version['wp_query_count'])
    xls.write_content(sheet_name, row, col + 1, curr_version['wp_count'])
    xls.write_content(sheet_name, row, col + 2, curr_version['wp_rate'])
    xls.write_content(sheet_name, row, col + 3, curr_version['false_wp_count'])
    if not globalVal._MANUAL_TEST_MODE:
        xls.write_content(sheet_name, row, col + 4, curr_version['wp_time'])
    if len(base_version):
        xls.write_content_with_color(sheet_name, row + 1, col + 2, 
            get_substraction(curr_version['wp_rate'], base_version['wp_rate']))
        xls.write_content_with_color(sheet_name, row + 1, col + 3, 
            get_substraction(curr_version['false_wp_count'], base_version['false_wp_count']), True)
        if not globalVal._MANUAL_TEST_MODE:
            xls.write_content_with_color(sheet_name, row + 1, col + 4, 
                get_substraction(curr_version['wp_time'], base_version['wp_time']), True)


def write_acu_result_data(row, col, acu_list, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：识别准确率
    """
    idx = 0
    for key in acu_list:
        if not key in curr_version:
            logging.debug('sdk:%s, audio:%s, could not find key:%s', vers, sample, key)
        else:
            xls.write_content(sheet_name, row, col + idx, curr_version[key])
            if ('wer' in key or 'ser' in key) and len(base_version):
                xls.write_content_with_color(sheet_name, row + 1, col + idx, 
                    get_substraction(curr_version[key], base_version[key]))
        idx += 1


def write_asr_speed_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：识别速度指标
    """
    idx = 0
    xls.write_content(sheet_name, row, col, curr_version['speed_query_num'])
    if int(curr_version['speed_query_num']):
        for key in ['partial_time', 'finish_time', 'tts_time', 'vad_start_time', 'vad_end_time']:
            idx += 1
            if key in curr_version:
                xls.write_content(sheet_name, row, col + idx, curr_version[key])
                if len(base_version):
                    xls.write_content_with_color(sheet_name, row + 1, col + idx, 
                        get_substraction(curr_version[key], base_version[key]), True)
    else:
        logging.error('sdk:%s, audio:%s, speed_query_num is 0', vers, sample)
        globalVal.record_err(vers + '/' + sample, 3103, '速度统计条数为0')


def write_manual_asr_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：识别指标，人工测试效果（准出）
    """
    xls.write_content(sheet_name, row, col + 1, curr_version['valid_asr_request_count'])
    xls.write_content(sheet_name, row, col + 2, curr_version['success_count'])
    xls.write_content(sheet_name, row, col + 3, curr_version['success_rate'])
    if len(base_version):
        xls.write_content_with_color(sheet_name, row + 1, col + 3, 
            get_substraction(curr_version['success_rate'], base_version['success_rate']))

    acu_start_col = col + 4
    accuracy_list = ['wer', 'ser']
    write_acu_result_data(row, acu_start_col, accuracy_list, curr_version, sheet_name, sample, vers, base_version)
    

def write_intersect_asr_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：识别指标，取交集，用户体验效果（准出）
    """
    xls.write_content(sheet_name, row, col + 1, curr_version['intersect_count'])
    xls.write_content(sheet_name, row, col + 2, curr_version['intersect_success_count'])
    xls.write_content(sheet_name, row, col + 3, curr_version['intersect_success_rate'])
    if len(base_version):
        xls.write_content_with_color(sheet_name, row + 1, col + 3, 
            get_substraction(curr_version['intersect_success_rate'], base_version['intersect_success_rate']))

    if 'mutiplyasr' in sample or 'pureasr' in sample:
        acu_start_col = col + 6
        xls.write_content(sheet_name, row, col + 4, curr_version['intersect_recall_count'])
        xls.write_content(sheet_name, row, col + 5, curr_version['intersect_recall_rate'])
        if len(base_version):
            xls.write_content_with_color(sheet_name, row + 1, col + 3, 
                get_substraction(curr_version['intersect_recall_rate'], base_version['intersect_recall_rate']))
    else:
        acu_start_col = col + 4
        
    accuracy_list = ['wer', 'ser']
    write_acu_result_data(row, acu_start_col, accuracy_list, curr_version, sheet_name, sample, vers, base_version)


def write_none_intersect_asr_data(
        row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：单版本识别指标，不取交集
    """
    idx = 0
    for key in ['asr_count', 'invalid_asr_request_count', 'valid_asr_request_count', 'success_count', 'success_rate']:
        xls.write_content(sheet_name, row, col + idx, curr_version[key])
        idx += 1
    if len(base_version):
        xls.write_content_with_color(sheet_name, row + 1, col + idx - 1, 
            get_substraction(curr_version['success_rate'], base_version['success_rate']))

    col += idx
    if 'mutiplyasr' in sample or 'pureasr' in sample:
        xls.write_content(sheet_name, row, col, curr_version['recall_count'])
        xls.write_content(sheet_name, row, col + 1, curr_version['recall_rate'])
        false_count = str(curr_version['false_asr_count']) + '/' + str(curr_version['false_interact_count'])
        if 'mutiplyasr' in sample:
            col += 2
            xls.write_content(sheet_name, row, col, curr_version['first_success_count'])
            if int(curr_version['wp_count']):
                xls.write_content(sheet_name, row, col + 1, 
                    '%.3f%%' % float(curr_version['first_success_count'] * 100 / curr_version['wp_count']))
        xls.write_content(sheet_name, row, col + 2, false_count)
        if len(base_version):
            xls.write_content_with_color(sheet_name, row + 1, col + 1, 
                get_substraction(curr_version['recall_rate'], base_version['recall_rate']))
            xls.write_content(sheet_name, row + 1, col + 2, 
                str(get_substraction(curr_version['false_asr_count'], base_version['false_asr_count'])) + '/' + \
                str(get_substraction(curr_version['false_interact_count'], base_version['false_interact_count'])))
        col += 2
    elif not 'split' in sample:
        col += 1
        xls.write_content(sheet_name, row, col, curr_version['false_asr_count'])

    accuracy_list = ['wer', 'ser']
    write_acu_result_data(row, col, accuracy_list, curr_version, sheet_name, sample, vers, base_version)


def write_contrast_asr_accuracy(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：取交集后的识别率指标，排除掉没有识别结果的情况，衡量服务端效果
    """
    accuracy_list = ['valid_intersect_count', 'intersect_wer', 'intersect_ser']
    write_acu_result_data(row, col, accuracy_list, curr_version, sheet_name, sample, vers, base_version)


def write_intersect_vp_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：声纹指标，取交集，用户体验效果（准出）
    """
    xls.write_content(sheet_name, row, col, curr_version['vp_total'])
    xls.write_content(sheet_name, row, col + 1, curr_version['vp_right'])
    xls.write_content(sheet_name, row, col + 3, curr_version['vp_wrong'])
    if curr_version['vp_total']:
        r_rate = '%.3f%%' % float(curr_version['vp_right'] * 100 / curr_version['vp_total'])
        w_rate = '%.3f%%' % float(curr_version['vp_wrong'] * 100 / curr_version['vp_total'])
        curr_version.update({'right_rate': r_rate, 'wrong_rate': w_rate})
    else:
        curr_version.update({'right_rate': '0.0%', 'wrong_rate': '0.0%'})
    xls.write_content(sheet_name, row, col + 2, curr_version['right_rate'])
    xls.write_content(sheet_name, row, col + 4, curr_version['wrong_rate'])
    if len(base_version):
        xls.write_content_with_color(sheet_name, row + 1, col + 2,
            get_substraction(curr_version['right_rate'], base_version['right_rate']))
        xls.write_content_with_color(sheet_name, row + 1, col + 4,
            get_substraction(curr_version['wrong_rate'], base_version['wrong_rate']), True)
    if 'voiceprint-jn' in sample:
        vp_unrecognized = curr_version['vp_total'] - curr_version['vp_right'] - curr_version['vp_wrong']
        xls.write_content(sheet_name, row, col + 5, vp_unrecognized)
        if curr_version['vp_total']:
            unrecg_rate = '%.3f%%' % float(vp_unrecognized * 100 / curr_version['vp_total'])
            curr_version.update({'unrecg_rate': unrecg_rate})
        else:
            curr_version.update({'unrecg_rate': '0.0%'})
        xls.write_content(sheet_name, row, col + 6, curr_version['unrecg_rate'])
        if len(base_version):
            xls.write_content_with_color(sheet_name, row + 1, col + 6,
                get_substraction(curr_version['unrecg_rate'], base_version['unrecg_rate']), True)
    if len(base_version):
        if 'voiceprint-jn' in sample:
            xls.set_fill(sheet_name, row + 1, [6, col + 6], 2)
        else:
            xls.set_fill(sheet_name, row + 1, [6, col + 4], 2)


def write_none_intersect_vp_data(row, col, curr_version, sheet_name, sample, vers, base_version={}):
    """写入数据：声纹指标，不取交集，用户体验效果（参考）
    """
    xls.write_content(sheet_name, row, col, curr_version['wp_query_count'])
    xls.write_content(sheet_name, row, col + 1, curr_version['wp_count'])
    xls.write_content(sheet_name, row, col + 2, curr_version['asr_query_count'])
    xls.write_content(sheet_name, row, col + 3, curr_version['valid_asr_request_count'])
    xls.write_content(sheet_name, row, col + 4, curr_version['success_count'])
    write_intersect_vp_data(row, col + 5, curr_version, sheet_name, sample, vers, base_version)


def write_none_intersect_role_data(row, col, curr_version, sheet_name, sample, vers, role_list, base_version={}):
    """写入数据：双人交互识别率指标，不取交集（准出）
    """
    idx = 0
    base_data = {}
    key_list = ['total_round', 'right_round', 'round_rate', 'role_wer', 'total_wer']
    for role in role_list:
        if role == 'sum':
            curr_data = curr_version['role_accuracy']['sum_info']
            if len(base_version):
                base_data = base_version['role_accuracy']['sum_info']
        elif role in curr_version['role_accuracy']['role_info']:
            curr_data = curr_version['role_accuracy']['role_info'][role]
            if len(base_version):
                base_data = base_version['role_accuracy']['role_info'][role]
        else:
            idx += len(key_list)
            continue
        for key in key_list:
            xls.write_content(sheet_name, row, col + idx, curr_data[key])
            if len(base_data) and ('rate' in key or 'wer' in key):
                xls.write_content_with_color(sheet_name, row + 1, col + idx, 
                    get_substraction(curr_data[key], base_data[key]))
            idx += 1
    if len(base_version):
        xls.set_fill(sheet_name, row + 1, [6, col + idx - 1], 2)


def cal_conclusion(sample, curr_conclusion, base_conclusion, curr_comparison, base_comparison):
    """
       计算性能指标差值
    Args:
        sample : [string] 测试场景信息
        curr_conclusion : [dict] 测试版本性能指标（取交集前）
        base_conclusion : [dict] 基线版本性能指标（取交集前）
        curr_comparison : [dict] 测试版本性能指标（取交集后）
        base_comparison : [dict] 基线版本性能指标（取交集后）
    Returns:
        param_json : [dict] 测试性能指标差值
    """
    param_json = {
        'WP_RATE': get_substraction(curr_conclusion['wp_rate'], base_conclusion['wp_rate'])
    }
    if base_conclusion['false_wp_count'] > 9:
        tmp = get_substraction(curr_conclusion['false_wp_count'], 
            base_conclusion['false_wp_count'])
        false_wp = "%.3f%%" % float(tmp / base_conclusion['false_wp_count'] * 100)
        param_json.update({'FALSE_WP': false_wp})
    if 'asr' in sample:
        if 'intersect_success_rate' in curr_comparison and \
            'intersect_success_rate' in base_comparison:
            param_json.update({
                'ASR_SUCCESS_RATE': get_substraction(
                    curr_comparison['intersect_success_rate'],
                    base_comparison['intersect_success_rate']),
                'ASR_WER': get_substraction(
                    curr_comparison['wer'],
                    base_comparison['wer']),
                'ASR_SER': get_substraction(
                    curr_comparison['ser'],
                    base_comparison['ser'])
            })
        if 'intersect_recall_rate' in curr_comparison and \
            'intersect_recall_rate' in base_comparison:
            param_json.update({
                'ASR_RECALL_RATE': get_substraction(
                    curr_comparison['intersect_recall_rate'],
                    base_comparison['intersect_recall_rate'])
            })
        if 'asr_error_code_info' in curr_conclusion:
            for k in curr_conclusion['asr_error_code_info']:
                if get_err_flag(k, get_audio_type(sample)):
                    param_json.update({
                        'ASR_ERROR_CODE': curr_conclusion['asr_error_code_info'][k]['diff']
                    })
    elif 'voiceprint' in sample:
        param_json.update({
            'VP_RIGHT_RATE': get_substraction(
                curr_comparison['right_rate'],
                base_comparison['right_rate']),
            'VP_WRONG_RATE': get_substraction(
                curr_comparison['wrong_rate'],
                base_comparison['wrong_rate'])
        })
        if 'voiceprint-jn' in sample:
            param_json.update({
                'VP_UNRECG_RATE': get_substraction(
                    curr_comparison['unrecg_rate'],
                    base_comparison['unrecg_rate'])
            })

    return param_json


class GetStreamReport(object):
    """
    Get stream data and write to report
    """
    def __init__(self):
        if os.path.exists(globalVal._RESULT_WORKBOOK_NAME):
            os.remove(globalVal._RESULT_WORKBOOK_NAME)
        self.__report_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        self.__result_workbook = Workbook()
        self.__result_workbook.remove(self.__result_workbook['Sheet']) 
        self.__result_worksheet_report = self.__result_workbook.create_sheet('声学测试结论')
        self.__result_worksheet_comp = self.__result_workbook.create_sheet('（单测试集、版本间对比）')
        self.__result_worksheet_detail = self.__result_workbook.create_sheet('（单测试集、单版本、无对比）')
        self.__result_worksheet_err = None
        self.__result_worksheet_speed = None
        self.__result_worksheet_acu = None
        self.__version_num = len(globalVal._VERSION_LIST)
        self.__title_col = {
            'wp': 11,
            'asr': 34, 'semanticasr': 34, 'splitasr': 34,
            'mutiplyasr': 40, 'pureasr': 46, 'roleasr': 21, 
            'voiceprint-jn': 25, 'voiceprint-jw': 21
        }
        self.read_stream_report()
        self.get_stream_report()
        self.record_stream_report()
    
    def read_stream_report(self):
        """
            [增量统计] 历史测试数据，从stream_data.db直接读取
        """
        if 'stream_data' in globalVal._DATABASE:
            stream_data = globalVal._DATABASE['stream_data']
            self.__vp_scene_result = stream_data['vp_scene_result']
            self.__speed_result = stream_data['speed_result']
            self.__asr_err_code = stream_data['asr_err_code']
            self.__true_wak_words = stream_data['true_wak_words']
            self.__false_wak_words = stream_data['false_wak_words']
            self.__wak_mode = stream_data['wak_mode']
        else:
            self.__vp_scene_result = {}
            self.__speed_result = {}
            self.__asr_err_code = {}
            self.__true_wak_words = []
            self.__false_wak_words = []
            self.__wak_mode = 0

    def record_stream_report(self):
        """
            [增量统计] 新增测试数据，存储进stream_data.db
        """
        globalVal._DATABASE['stream_data'] = {
            'vp_scene_result': self.__vp_scene_result,
            'speed_result': self.__speed_result,
            'asr_err_code': self.__asr_err_code,
            'true_wak_words': self.__true_wak_words,
            'false_wak_words': self.__false_wak_words,
            'wak_mode': self.__wak_mode
        }

    def get_stream_report(self):
        """
            [灌测报告] 对每个场景的数据进行汇总，生成"统计结果.xlsx"
        """
        # [0] 写入类型信息
        updateInfo(self.__version_num)
        logging.info(globalVal._VERSION_LIST)

        # [1] 读取每个版本的详细数据，并按照场景进行数据汇总
        start = time.time()
        self.merge_scene_data()
        end = time.time()
        print('merge_scene_data Execution Time: ', end - start)

        # [2] 将数据写入 统计结果.xlsx
        start = time.time()
        self.write_report()
        end = time.time()
        print('write_report Execution Time: ', end - start)
        self.__result_workbook.save(globalVal._RESULT_WORKBOOK_NAME)

        # [3] 将唤醒数据单独写入 统计结果.xlsx（唤醒详情）
        if self.__wak_mode:
            self.write_wp_report()

        # [5] 按测试场景汇总 各版本的声源定位准确率，并写入 统计结果.xlsx（声源定位准确率）
        self.getWpLocationReport()

        # [6] 写入指标统计说明
        xls.streamReportDirections(self.__result_workbook)

    def write_merge_scene_info(self, report_start, comp_start, report_row, comp_row, last_scene_info, curr_scene_info, mode):
        """
            [测试场景归类] 按语音类别、极客模式、测试模式、场景、米数进行数据归类，然后再写入表格
        """
        last_jike, last_speech, last_type, last_env = last_scene_info.split('_')[0:4]
        curr_jike, curr_speech, curr_type, curr_env = curr_scene_info.split('_')[0:4]
        if mode:
            add_c = 4
            if globalVal._PARAM['CONTRAST_WITH_ALL']:
                add_r = 6 + 2 * self.__version_num - 1
            else:
                add_r = 6 + 3 * self.__version_num - 3
        else:
            add_r = 0
            add_c = 0
        try:
            if mode or last_jike != curr_jike:
                write_col = conclusion.col_match(last_jike)
                xls.write_content(self.__result_worksheet_report, 
                    [report_start[0], report_row - 1], [1, 1], write_col)
                xls.write_content(self.__result_worksheet_comp, 
                    [comp_start[0], comp_row - 1], [1, 1], write_col)
                xls.write_content(self.__result_worksheet_speed, 
                    [report_start[0], report_row - 1], [1, 1], write_col)
                xls.write_content(self.__result_worksheet_acu, 
                    [report_start[0], report_row - 1], [1, 1], write_col)
                xls.write_content(self.__result_worksheet_err, 
                    [report_start[0], report_row - 1], [1, 1], write_col)
                report_start[0] = report_row + add_r
                comp_start[0] = comp_row + add_c
            if mode or last_type != curr_type:
                write_col = conclusion.col_match(last_type)
                xls.write_content(self.__result_worksheet_report, 
                    [report_start[1], report_row - 1], [2, 2], write_col)
                xls.write_content(self.__result_worksheet_comp, 
                    [comp_start[1], comp_row - 1], [2, 2], write_col)
                xls.write_content(self.__result_worksheet_speed, 
                    [report_start[1], report_row - 1], [2, 2], write_col)
                xls.write_content(self.__result_worksheet_acu, 
                    [report_start[1], report_row - 1], [2, 2], write_col)
                xls.write_content(self.__result_worksheet_err, 
                    [report_start[1], report_row - 1], [2, 2], write_col)
                report_start[1] = report_row + add_r
                comp_start[1] = comp_row + add_c
            if mode or last_env != curr_env:
                write_col = conclusion.col_match(last_env)
                xls.write_content(self.__result_worksheet_report, 
                    [report_start[2], report_row - 1], [3, 3], write_col)
                xls.write_content(self.__result_worksheet_comp, 
                    [comp_start[2], comp_row - 1], [3, 3], write_col)
                xls.write_content(self.__result_worksheet_speed, 
                    [report_start[2], report_row - 1], [3, 3], write_col)
                xls.write_content(self.__result_worksheet_acu, 
                    [report_start[2], report_row - 1], [3, 3], write_col)
                xls.write_content(self.__result_worksheet_err, 
                    [report_start[2], report_row - 1], [3, 3], write_col)
                report_start[2] = report_row + add_r
                comp_start[2] = comp_row + add_c
        except:
            globalVal.record_err(last_scene_info, 4003, '场景信息merge方式写入出错')
        return report_start, comp_start

    def write_conclusion(self, sheet_name, row, col, old_type, new_type, type_start_col_list):
        """
            [生成中文结论] 按既定规则，对不满足准出标准的指标进行汇总，并以文字结论形式写入表格下方
        """
        # 1、写入结论
        if '[' in conclusion._CONTENT:
            xls.write_content(sheet_name, row, col, conclusion._CONTENT, 6)
        elif self.__version_num > 1:
            xls.write_content(sheet_name, row, col, conclusion._CONTENT + ' 各场景基本持平', 6)
        # 2、清除结论
        conclusion.clear_conclusion()
        conclusion._CONTENT = '【' + conclusion._CONCLUSION_DESC['speech_type'][new_type] + '测试结论】\n'
        # 3、20210225 速度统计，额外增加全场景统计
        if 'roleasr' != old_type and not 'voiceprint' in old_type:
            row = self.write_report_data(self.__result_worksheet_speed, row, 
                type_start_col_list, self.__speed_result[old_type], old_type + '_汇总', 2)
        return row + 2

    def write_report(self):
        """
            [生成声学报告] 按语音类型，分别写入唤醒、单次识别、多次识别、纯识别、语义识别、声纹等声学报告
        """
        type_start_col_list = []
        report_row = 1
        comp_row = 1
        title_flag = {
            'wp': True,
            'asr': True,
            'mutiplyasr': True,
            'pureasr': True,
            'roleasr': True,
            'splitasr': True,
            'semanticasr': True,
            'voiceprint-jn': True,
            'voiceprint-jw': True
        }
        global_result = globalVal._DATABASE
        report_start = [5, 5, 5, 5]
        comp_start = [5, 5, 5, 5]
        last_scene_info = ""
        audio_type = ""
        detail_row = 1

        for scene_name in globalVal._SAMPLE_LIST:
            if 'wp_dci' in scene_name:
                continue
            audio_type = get_audio_type(scene_name)
            if len(scene_name.split('_')) > 4:
                distance = scene_name.split('_', 5)[4]
            else:
                distance = "all"
            # [1] 写表格title
            if audio_type in title_flag:
                if title_flag[audio_type]:
                    title_flag[audio_type] = False
                    # 类型更换时，写入第0、2、3列测试场景信息:极客模式、测试模式、场景
                    if last_scene_info == "": 
                        conclusion._CONTENT = \
                            '【' + conclusion._CONCLUSION_DESC['speech_type'][audio_type] + '测试结论】\n'
                    else:
                        report_start, comp_start = self.write_merge_scene_info(
                            report_start, comp_start, report_row, comp_row, last_scene_info, scene_name, True)
                        last_audio_type = get_audio_type(last_scene_info)
                        report_row = self.write_conclusion(self.__result_worksheet_report, 
                            report_row, [1, self.__title_col[last_audio_type]], 
                            last_audio_type, audio_type, type_start_col_list)
                    sp_report_row = report_row
                    report_row, type_start_col_list = self.write_report_title(
                        audio_type, self.__result_worksheet_report, report_row)
                    comp_row, type_start_col_list = self.write_report_title(
                        audio_type, self.__result_worksheet_comp, comp_row)
                    self.write_special_report_title(
                        audio_type, self.__result_worksheet_speed, sp_report_row, 'speed')
                    self.write_special_report_title(
                        audio_type, self.__result_worksheet_acu, sp_report_row, 'acu')
                    self.write_special_report_title(
                        audio_type, self.__result_worksheet_err, sp_report_row, 'err_code')
                else:
                    # 类型不变时，写入第0、2、3列测试场景信息:极客模式、测试模式、场景
                    report_start, comp_start = self.write_merge_scene_info(
                        report_start, comp_start, report_row, comp_row, last_scene_info, scene_name, False)
                report_start[3] = report_row
                comp_start[3] = comp_row
            else:
                continue
            # [2] 按测试集、测试场景写入数据
            scene_result = global_result
            for item in globalVal.get_scene_list(scene_name):
                scene_result = scene_result[item]
            # [3] scene_result:场景名称，audio_name:测试音频名称
            for audio_name in scene_result:
                if audio_name in ['sample_name', 'answer', 'excel_count']:
                    continue
                elif audio_name == 'summary':
                    report_row = self.write_report_data(self.__result_worksheet_report, report_row, 
                        type_start_col_list, scene_result[audio_name], scene_name + '_' + audio_name, 1)
                else:
                    comp_row = self.write_report_data(self.__result_worksheet_comp, comp_row, 
                        type_start_col_list, scene_result[audio_name], scene_name + '_' + audio_name)
                    detail_row = self.update_detail_excel(
                        self.__result_worksheet_detail, detail_row, 
                        scene_name + '/' + audio_name, scene_result[audio_name]
                    )
            # [4] 写入第4列测试场景信息
            try:
                xls.write_content(self.__result_worksheet_report, [report_start[3], report_row - 1], [4, 4], distance)
                xls.write_content(self.__result_worksheet_comp, [comp_start[3], comp_row - 1], [4, 4], distance)
                xls.write_content(self.__result_worksheet_speed, [report_start[3], report_row - 1], [4, 4], distance)
                xls.write_content(self.__result_worksheet_acu, [report_start[3], report_row - 1], [4, 4], distance)
                xls.write_content(self.__result_worksheet_err, [report_start[3], report_row - 1], [4, 4], distance)
            except:
                globalVal.record_err(scene_name, 4003, '距离merge方式写入出错')
            last_scene_info = scene_name
        # 特殊场景"1_2_3_4_5"，为实现最后一个场景信息merge写入
        if audio_type in title_flag:
            report_start, comp_start = self.write_merge_scene_info(
                report_start, comp_start, report_row, comp_row, last_scene_info, "1_2_3_4_5", False)
            report_row = self.write_conclusion(self.__result_worksheet_report, 
                report_row, [1, self.__title_col[audio_type]], audio_type, audio_type, type_start_col_list)

        # 20201105 声纹统计，额外增加测试环境汇总统计
        report_row += 5
        title_flag = {'voiceprint-jn': True, 'voiceprint-jw': True}
        last_audio_type = ''
        for environment in sorted(self.__vp_scene_result):
            audio_type = environment.split('_')[0]
            if last_audio_type == '': 
                conclusion._CONTENT = \
                    '【' + conclusion._CONCLUSION_DESC['speech_type'][audio_type] + '测试结论】\n'
            elif last_audio_type != audio_type:
                report_row = self.write_conclusion(self.__result_worksheet_report, 
                    report_row, [1, self.__title_col[last_audio_type]], last_audio_type, audio_type, type_start_col_list)
            last_audio_type = audio_type
            if title_flag[audio_type]:
                # 类型改变时，写入第0、2、3列测试场景信息:极客模式、测试模式、场景
                title_flag[audio_type] = False
                report_row, type_start_col_list = self.write_report_title(
                    audio_type, self.__result_worksheet_report, report_row)
            report_row_start = report_row
            report_row = self.write_report_data(self.__result_worksheet_report, report_row, 
                type_start_col_list, self.__vp_scene_result[environment], environment)
            xls.write_content(self.__result_worksheet_report, 
                [report_row_start, report_row - 1], [1, 4], environment.split('_')[1] + '(全距离、角度汇总)')
        if len(self.__vp_scene_result):
            report_row = self.write_conclusion(self.__result_worksheet_report, 
                report_row, [1, self.__title_col[audio_type]], audio_type, audio_type, type_start_col_list)

    def write_report_data(self, result_worksheet, row, count_list, result_summary, sample, flags=0):
        """
            写入report数据，按SDK版本：
            1、读入版本下的conclusion并记入Excel
            2、从comparison里找版本对应的数据并记入Excel
        """
        wp_start_col, asr_start_col, speed_start_col, none_intersect_start_col, contrast_start_col = count_list
        audio_type = get_audio_type(sample)
        # [0] 测试信息判定: 当所有版本都有sample集合的测试数据时，才进行汇总统计
        intersect_count = 0
        for version in globalVal._VERSION_LIST:
            if version in result_summary:
                intersect_count += 1
            else:
                globalVal.record_err(version + '/' + sample, 4102, '该场景数据不完整，不参与最终结果统计')
        if intersect_count != self.__version_num:
            return row + self.__version_num
        
        # [1] 通用信息写入: 如测试集名称、唤醒query总数、识别query总数等
        if self.__version_num > 1:
            if globalVal._PARAM['CONTRAST_WITH_ALL']:
                merge_row_count = row + 2 * self.__version_num - 2
            else:
                merge_row_count = row + 3 * self.__version_num - 4
        else:
            merge_row_count = row
        if flags == 2:
            xls.write_content(result_worksheet, [row, merge_row_count], [1, 5], sample)
        else:
            xls.write_content(result_worksheet, [row, merge_row_count], [5, 5], sample)
        if flags == 1:
            xls.write_content(self.__result_worksheet_speed, [row, merge_row_count], [5, 5], sample)
            xls.write_content(self.__result_worksheet_acu, [row, merge_row_count], [5, 5], sample)
            xls.write_content(self.__result_worksheet_err, [row, merge_row_count], [5, 5], sample)

        # [2] 按版本顺序，依次写入各项声学性能指标
        for version in range(self.__version_num):
            if globalVal._PARAM['CONTRAST_WITH_ALL']:
                contrast_version_list = [version]
            else:
                if version == 0:
                    continue
                else:
                    contrast_version_list = [0, version]
            base_conclusion = {}
            base_comparison = {}
            for vers in contrast_version_list:
                xls.write_content(result_worksheet, row, 6, globalVal._VERSION_LIST[vers])
                if vers != 0:
                    xls.write_content(result_worksheet, row + 1, 6, '差值(测试-基线)')
                    xls.set_fill(result_worksheet, row + 1, [6, self.__title_col[get_audio_type(sample)]], 2)
                if flags == 1:
                    xls.write_content(self.__result_worksheet_speed, row, 6, globalVal._VERSION_LIST[vers])
                    xls.write_content(self.__result_worksheet_acu, row, 6, globalVal._VERSION_LIST[vers])
                    xls.write_content(self.__result_worksheet_err, row, 6, globalVal._VERSION_LIST[vers])
                elif flags == 2:
                    curr_conclusion = result_summary[globalVal._VERSION_LIST[vers]]
                    if vers != 0:
                        base_conclusion = result_summary[globalVal._VERSION_LIST[0]]
                    write_speed_analyse_data(row, 7, curr_conclusion, 
                        result_worksheet, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                    row += 1
                    if vers != 0:
                        row += 1
                    continue
                # 得到当前版本&基线版本的conclusion数据（不取交集）
                curr_conclusion = result_summary[globalVal._VERSION_LIST[vers]]['conclusion']
                if vers != 0:
                    base_conclusion = result_summary[globalVal._VERSION_LIST[0]]['conclusion']
                # 得到当前版本&基线版本的comparison数据（取交集）
                if asr_start_col and not 'roleasr' in sample and not globalVal._MANUAL_TEST_MODE:
                    if vers != 0 or globalVal._PARAM['CONTRAST_WITH_ALL'] or not 'mutiplyasr' in sample:
                        curr_comparison = result_summary['comparison'][globalVal._VERSION_LIST[vers]]
                    else:
                        result_idx = int(contrast_version_list[1]) - 1
                        curr_comparison = result_summary['comparison'][globalVal._VERSION_LIST[0]][result_idx]
                    if vers != 0:
                        if globalVal._PARAM['CONTRAST_WITH_ALL'] or not 'mutiplyasr' in sample:
                            base_comparison = result_summary['comparison'][globalVal._VERSION_LIST[0]]
                        else:
                            base_comparison = result_summary['comparison'][globalVal._VERSION_LIST[0]][vers - 1]
                else:
                    curr_comparison = {}
                if 'voiceprint' in sample:
                    # 写入数据：声纹指标（正确识别率、错误识别率、未识别率）
                    write_intersect_vp_data(row, asr_start_col, curr_comparison, 
                        result_worksheet, sample, globalVal._VERSION_LIST[vers], base_comparison)
                    write_none_intersect_vp_data(row, speed_start_col, curr_conclusion, 
                        result_worksheet, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                elif 'roleasr' in sample:
                    # 写入数据：双人交互指标（角色判定准确率、识别准确率）
                    write_none_intersect_role_data(row, asr_start_col, curr_conclusion, 
                        result_worksheet, sample, globalVal._VERSION_LIST[vers], self.role_list, base_conclusion)
                else:
                    # 写入数据：唤醒指标（唤醒率、误唤醒次数、唤醒速度）
                    write_wp_result_data(row, wp_start_col, curr_conclusion, 
                        result_worksheet, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                    # 写入数据：识别指标（识别成功率、召回率、字句准、识别首包响应时间、硬延迟、TTS首包时间、VAD检测延迟时间）
                    if 'asr' in sample:
                        xls.write_content(result_worksheet, row, asr_start_col, curr_conclusion['asr_query_count'])
                        if globalVal._MANUAL_TEST_MODE:
                            write_manual_asr_data(row, asr_start_col, curr_conclusion, 
                                result_worksheet, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                        else:
                            write_intersect_asr_data(row, asr_start_col, curr_comparison, 
                                result_worksheet, sample, globalVal._VERSION_LIST[vers], base_comparison)
                            write_asr_speed_data(row, speed_start_col, curr_conclusion, 
                                result_worksheet, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                            write_none_intersect_asr_data(row, none_intersect_start_col, curr_conclusion, 
                                result_worksheet, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                            write_contrast_asr_accuracy(row, contrast_start_col, curr_comparison, 
                                result_worksheet, sample, globalVal._VERSION_LIST[vers], base_comparison)
                        # 写入识别准确率专项报告（字准、句准、插入错误、删除错误、替换错误）
                        if flags == 1:
                            if vers != 0:
                                xls.write_content(self.__result_worksheet_acu, row + 1, 6, '差值(测试-基线)')
                                xls.write_content(self.__result_worksheet_err, row + 1, 6, '差值(测试-基线)')
                            write_acu_analyse_data(row, 7, curr_conclusion, 
                                self.__result_worksheet_acu, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                            write_err_analyse_data(row, 7, curr_conclusion, 
                                self.__result_worksheet_err, sample, globalVal._VERSION_LIST[vers], 
                                self.__asr_err_code[audio_type], base_conclusion)
                    # 写入速度专项报告（唤醒、识别速度；各项速度指标的中位值、90分位值、最大值、、最小值、方差）
                    if flags == 1:
                        if vers != 0:
                            xls.write_content(self.__result_worksheet_speed, row + 1, 6, '差值(测试-基线)')
                        write_speed_analyse_data(row, 7, curr_conclusion, 
                            self.__result_worksheet_speed, sample, globalVal._VERSION_LIST[vers], base_conclusion)
                # 声学结论以文字形式记录
                if vers != 0 and 'summary' in sample:
                    param_json = cal_conclusion(sample, curr_conclusion, base_conclusion, 
                        curr_comparison, base_comparison)
                    conclusion.get_conclusion(sample, globalVal._VERSION_LIST[vers], param_json)
                row += 1
                if vers != 0:
                    row += 1
        return row

    def write_title(self, audio_type, t_type, result_worksheet, row, title_merge_col={}):
        """通用title - 包括模式、场景、米数/角度、SDK版本等
        """
        if len(title_merge_col) == 0:
            title_merge_col = self.__title_col
            if globalVal._MANUAL_TEST_MODE:
                self.__title_col['asr'] = 16
        xls.write_content(result_worksheet, [row, row], [1, title_merge_col[t_type]], 
            conclusion._CONCLUSION_DESC['speech_type'][audio_type], 3)
        xls.write_content(result_worksheet, [row + 1, row + 1], [1, title_merge_col[t_type]], 
            '报告时间:' + str(self.__report_time) , 3)
        xls.write_title(result_worksheet, globalVal._R_TITLE_LIST, row + 2, 1, 1)

    def write_report_title(self, audio_type, result_worksheet, row):
        """写入统计结果.xlsx(声学结论)的title
        """
        # 通用title
        self.write_title(audio_type, audio_type, result_worksheet, row)
        row += 2

        # 声学指标title（准出）
        if 'voiceprint' in audio_type:
            # 声纹指标title（准出）
            vp_start_col = 7
            vp_title_list = globalVal._R_INTERSECT_VP_LIST[audio_type]
            none_intersect_start_col = vp_start_col + len(vp_title_list)
            none_intersect_end_col = none_intersect_start_col + \
                len(globalVal._R_COMMON_VP_LIST) + len(vp_title_list) - 1
            xls.write_content(result_worksheet, 
                [row, row], [vp_start_col, none_intersect_start_col - 1],
                '声纹效果（准出）- 排除掉未唤醒、识别错误、无tts后，取交集', 8)
            xls.write_content(result_worksheet, 
                [row, row], [none_intersect_start_col, none_intersect_end_col],
                '声纹效果（参考）- 排除掉未唤醒、识别错误、无tts后，不取交集，仅关注单版本效果', 9)
            row += 1
            xls.write_title(result_worksheet, vp_title_list, 
                row, vp_start_col, 0, 0, 8)
            xls.write_title(result_worksheet, globalVal._R_COMMON_VP_LIST + vp_title_list, 
                row, none_intersect_start_col, 0, 0, 9)
            return row + 1, [vp_start_col, vp_start_col, none_intersect_start_col, 0, 0]
        else:
            if 'pureasr' == audio_type or 'splitasr' == audio_type or 'roleasr' == audio_type:
                wp_start_col = 7
                asr_start_col = 7
            else:
                wp_start_col = 7
                asr_start_col = wp_start_col + len(globalVal._R_WAKEUP_LIST)
                # 唤醒指标title（准出）
                xls.write_content(result_worksheet, [row, row], 
                    [wp_start_col, asr_start_col - 1], '唤醒效果（准出）', 7)
                xls.write_title(result_worksheet, globalVal._R_WAKEUP_LIST, row + 1, wp_start_col, 0, 0, 7)
            if 'roleasr' in audio_type:
                # 双人交互指标title（准出）
                length = len(globalVal._R_NONE_INTERSECT_ASR_LIST[audio_type])
                col = asr_start_col
                for i in range(len(self.role_list)):
                    role = self.role_list[i]
                    if role == 'sum':
                        xls.write_content(result_worksheet, [row, row],
                            [col, col + length -1], '汇总', 20 + i)
                    else:
                        xls.write_content(result_worksheet, [row, row],
                            [col, col + length -1], '角色' + str(role), 20 + i)
                    xls.write_title(result_worksheet, globalVal._R_NONE_INTERSECT_ASR_LIST[audio_type], 
                        row + 1, col, 0, 0, 20 + i)
                    col += length
                return row + 2, [wp_start_col, asr_start_col, 0, 0, 0]
            elif 'asr' in audio_type:
                if self.__result_worksheet_err is None:
                    self.__result_worksheet_err = self.__result_workbook.create_sheet('识别错误码')
                    self.__result_worksheet_acu = self.__result_workbook.create_sheet('识别率专项报告')
                    if not globalVal._MANUAL_TEST_MODE:
                        self.__result_worksheet_speed = self.__result_workbook.create_sheet('速度专项报告')
                if globalVal._MANUAL_TEST_MODE:
                    xls.write_content(result_worksheet, [row, row],
                        [asr_start_col, asr_start_col + len(globalVal._R_INTERSECT_ASR_LIST[audio_type]) - 1],
                        '识别效果（准出）- 排除掉未唤醒、错误码后', 8)
                    xls.write_title(result_worksheet, globalVal._R_INTERSECT_ASR_LIST[audio_type], 
                        row + 1, asr_start_col, 0, 0, 8)
                    return row + 2, [wp_start_col, asr_start_col, 0, 0, 0]
                # 识别指标title（准出：取交集）
                if 'splitasr' == audio_type:
                    none_intersect_start_col = asr_start_col + len(globalVal._R_INTERSECT_ASR_LIST[audio_type])
                    contrast_start_col = none_intersect_start_col \
                                         + len(globalVal._R_NONE_INTERSECT_ASR_LIST[audio_type])
                    speed_start_col = contrast_start_col + len(globalVal._R_CONTRAST_ASR_LIST)
                    xls.write_content(result_worksheet, [row, row],
                                      [asr_start_col, none_intersect_start_col - 1],
                                      '识别效果（准出）- 排除掉未唤醒、网络错误后，取交集', 8)
                    xls.write_content(result_worksheet, [row, row],
                                      [none_intersect_start_col, contrast_start_col - 1],
                                      '端识别效果（参考）- 排除掉未唤醒、网络错误后，不取交集，仅关注单版本效果', 10)
                    xls.write_content(result_worksheet, [row, row],
                                      [contrast_start_col,
                                       speed_start_col - 1],
                                      '云端识别准确率（参考）- 去掉空识别结果，取交集', 11)
                    xls.write_content(result_worksheet, [row, row],
                                      [speed_start_col, speed_start_col + len(globalVal._R_SPEED_LIST[audio_type]) - 1],
                                      '识别速度效果（准出）- 仅统计回调完整的query，不取交集', 9)
                else:
                    speed_start_col = asr_start_col + len(globalVal._R_INTERSECT_ASR_LIST[audio_type])
                    none_intersect_start_col = speed_start_col + len(globalVal._R_SPEED_LIST[audio_type])
                    contrast_start_col = none_intersect_start_col \
                                         + len(globalVal._R_NONE_INTERSECT_ASR_LIST[audio_type])
                    xls.write_content(result_worksheet, [row, row],
                                      [asr_start_col, speed_start_col - 1],
                                      '识别效果（准出）- 排除掉未唤醒、网络错误后，取交集', 8)
                    xls.write_content(result_worksheet, [row, row],
                                      [speed_start_col, none_intersect_start_col - 1],
                                      '识别速度效果（准出）- 仅统计回调完整的query，不取交集', 9)
                    xls.write_content(result_worksheet, [row, row],
                                      [none_intersect_start_col, contrast_start_col - 1],
                                      '端识别效果（参考）- 排除掉未唤醒、网络错误后，不取交集，仅关注单版本效果', 10)
                    xls.write_content(result_worksheet, [row, row],
                                      [contrast_start_col,
                                       contrast_start_col + len(globalVal._R_CONTRAST_ASR_LIST) - 1],
                                      '云端识别准确率（参考）- 去掉空识别结果，取交集', 11)
                row += 1
                xls.write_title(result_worksheet, globalVal._R_INTERSECT_ASR_LIST[audio_type], 
                    row, asr_start_col, 0, 0, 8)
                xls.write_title(result_worksheet, globalVal._R_SPEED_LIST[audio_type], 
                    row, speed_start_col, 0, 0, 9)
                xls.write_title(result_worksheet, globalVal._R_NONE_INTERSECT_ASR_LIST[audio_type], 
                    row, none_intersect_start_col, 0, 0, 10)
                if 'semanticasr' == audio_type:
                    xls.write_title(result_worksheet, globalVal._R_CONTRAST_SEMANTICASR_LIST, 
                        row, contrast_start_col, 0, 0, 11)
                else:
                    xls.write_title(result_worksheet, globalVal._R_CONTRAST_ASR_LIST, 
                        row, contrast_start_col, 0, 0, 11)
                return row + 1, [wp_start_col, asr_start_col, \
                    speed_start_col, none_intersect_start_col, contrast_start_col]
            else:
                return row + 2, [wp_start_col, 0, 0, 0, 0]

    def write_special_report_title(self, audio_type, result_worksheet, row, sp_type):
        """写入统计结果.xlsx(专项报告)的title
        """
        if 'roleasr' == audio_type:
            return
        title_merge_col = {}
        row += 2
        # 专项报告title
        col = 7
        if sp_type == 'speed':
            report_title = globalVal._SPEED_ANALYSE_LIST
            title_list, speed_list = get_speed_list(audio_type)
            add_col = len(report_title)
            xls.write_content(result_worksheet, [row, row], [col, col + len(title_list) - 1], '速度效果（准出）', 7)
            xls.write_title(result_worksheet, title_list, row + 1, col, 0, 0, 7)
            title_merge_col.update({'speed': 6 + len(title_list) + len(speed_list) * add_col})
            idx = 0
            col += len(title_list)
            for key in speed_list:
                if 'wp_speed_list' == key:
                    xls.write_content(result_worksheet, 
                        [row, row], [col, col + add_col - 1], globalVal._SPEED_TYPE['wp'][key], 8 + idx)
                else:
                    xls.write_content(result_worksheet, 
                        [row, row], [col, col + add_col - 1], globalVal._SPEED_TYPE['asr'][key], 8 + idx)
                xls.write_title(result_worksheet, report_title, row + 1, col, 0, 0, 8 + idx)
                col += add_col
                idx += 1
        elif sp_type == 'acu':
            report_title = globalVal._ACU_ANALYSE_LIST
            title_merge_col.update({'acu': 6 + len(report_title)})
            xls.write_content(result_worksheet, 
                [row, row], [col, col + len(report_title) - 1], '识别率效果（准出）', 8)
            xls.write_title(result_worksheet, report_title, row + 1, col, 0, 0, 8)
        elif sp_type == 'err_code':
            i = 0
            idx = 0
            for k in self.__asr_err_code[audio_type]:
                if get_err_flag(k, audio_type):
                    suffix = '(准出)'
                else:
                    suffix = '(参考)'
                start_col = col + i
                for err_code in self.__asr_err_code[audio_type][k]:
                    i += 1
                    xls.write_content(result_worksheet, row + 1, col + i, str(err_code) + ' 个数', 8 + idx)
                xls.write_content(result_worksheet, row + 1, start_col, '识别请求总条数', 8 + idx)
                xls.write_content(result_worksheet, row, [start_col, col + i], k + suffix, 8 + idx)
                i += 1
                idx += 1
            title_merge_col.update({'err_code': col + i - 1})
            
        # 通用title
        self.write_title(audio_type, sp_type, result_worksheet, row - 2, title_merge_col)

    def get_compare_result(self, sample_name, flag):
        """
        计算多版本交集结果(测试场景数据汇总：各个测试集txt全部cat到一起，调用wer工具计算准确率；其他指标根据comparison计算)
        """
        result = {}
        if 'voiceprint' in sample_name:
            contrast_class = get_contrast.VoiceprintContrast(sample_name)
            result = contrast_class.get_intersect_vp(flag)
        elif 'asr' in sample_name:
            contrast_class = get_contrast.ContrastAccuracy(sample_name)
            result = contrast_class.get_contrast_accuracy(flag)
        try:
            result = json.loads(str(result).replace("'", '"'))
            logging.debug('get_compare_result: ' + sample_name)
            logging.debug(result)
        except TypeError:
            globalVal.record_err(sample_name, 4201, '交集统计结果不是json格式')
        return result
       
    def get_summary_result(self, scene_name):
        """
        单个测试场景:汇总各版本的性能指标
        """
        scene_data = globalVal.get_scene_data(scene_name)
        audio_type = get_audio_type(scene_name)
        if not audio_type in self.__asr_err_code:
            self.__asr_err_code.update({audio_type: {}})
        result_json = {}
        # [1] 相同场景、按测试集进行数据汇总
        for audio_name in scene_data:
            if audio_name == 'summary':
                continue
            intersect_count = 0
            for version in globalVal._VERSION_LIST:
                if version in scene_data[audio_name]:
                    intersect_count += 1
                else:
                    globalVal.record_err(version + '/' + audio_name, 4102, '该场景数据不完整，不参与最终结果统计')
            if intersect_count == self.__version_num:
                if globalVal._MANUAL_TEST_MODE:
                    answer_path = os.path.join(globalVal.get_save_path(scene_name, audio_name), 
                        scene_name.split('/')[0] + '_' + audio_name + '_answer_result.txt')
                    get_contrast.update_result_file(scene_name.split('/')[0], audio_name, answer_path)
                for version in globalVal._VERSION_LIST:
                    if globalVal._MANUAL_TEST_MODE:
                        result_path = os.path.join(globalVal.get_save_path(scene_name, audio_name), 
                            version + '_' + scene_name.split('/')[0] + '_' + audio_name + '_recognize_result.txt')
                        get_contrast.update_result_file(scene_name.split('/')[0], audio_name, result_path, version)
                    conclusion_data = scene_data[audio_name][version]['conclusion']
                    if version in result_json:
                        version_data = result_json[version]['conclusion']
                        for key in conclusion_data:
                            if key == 'role_accuracy':
                                speech_parse._update_role_info(scene_name, audio_name, version, 
                                    version_data['role_accuracy'], conclusion_data['role_accuracy'])
                            elif key == 'wp_mode':
                                if conclusion_data[key] == 3:
                                    wak_dict = globalVal._BASE_WHISPER
                                else:
                                    wak_dict = globalVal._WP_DICT
                                for wpWords in conclusion_data['wp_status']:
                                    if not wpWords in version_data['wp_status']:
                                        version_data['wp_status'].update({wpWords: copy.deepcopy(wak_dict)})
                                    curr_result = conclusion_data['wp_status'][wpWords]
                                    for k in wak_dict:
                                        if k != 'wpAvgTime' and k != 'wpMaxTime':
                                            version_data['wp_status'][wpWords][k] += curr_result[k]
                                    if version_data['wp_status'][wpWords]['wpMaxTime'] < curr_result['wpMaxTime']:
                                        version_data['wp_status'][wpWords]['wpMaxTime'] = curr_result['wpMaxTime']
                                    if 'whisperInfo' in curr_result:
                                        speech_parse.update_whisper_info(
                                            version_data['wp_status'][wpWords]['whisperInfo'], 
                                            curr_result['whisperInfo']
                                        )
                                for wpWords in conclusion_data['wp_false']:
                                    if not wpWords in version_data['wp_false']:
                                        version_data['wp_false'].update({wpWords: conclusion_data['wp_false'][wpWords]})
                                    else:
                                        version_data['wp_false'][wpWords] += conclusion_data['wp_false'][wpWords]
                            elif key == 'wp_time':
                                version_data[key] += conclusion_data[key] * conclusion_data['wp_count']
                            elif 'time' in key:
                                try:
                                    version_data[key] += conclusion_data[key] * conclusion_data['speed_query_num']
                                except KeyError:
                                    logging.debug("找不到key: " + key + ' in ' + scene_name)
                            elif 'count' in key or 'vp' in key:
                                version_data[key] += conclusion_data[key]
                            elif key == 'asr_speed_list':
                                for k in version_data[key]:
                                    if k in conclusion_data[key]:
                                        version_data[key][k] += conclusion_data[key][k]
                            elif key == 'asr_error_code_info':
                                for k in conclusion_data[key]:
                                    version_data[key][k]['denominator'] += conclusion_data[key][k]['denominator']
                                    err_info = conclusion_data[key][k]['info']
                                    for err_code in err_info:
                                        if err_code in version_data[key][k]['info']:
                                            version_data[key][k]['info'][err_code] += err_info[err_code]
                                        else:
                                            version_data[key][k]['info'].update({err_code: err_info[err_code]})
                        version_data['speed_query_num'] += conclusion_data['speed_query_num']
                    else:
                        result_json.update({version: {'conclusion': copy.deepcopy(conclusion_data)}})
                        version_data = result_json[version]['conclusion']
                        for key in conclusion_data:
                            if key == 'role_accuracy':
                                speech_parse._update_role_info(scene_name, audio_name, version, 
                                    version_data['role_accuracy'], [])
                            elif key == 'wp_time':
                                version_data[key] *= version_data['wp_count']
                            elif 'time' in key:
                                version_data[key] *= version_data['speed_query_num']
        # [2] 相同场景、计算汇总后的唤醒、识别准确率以及速度指标          
        for version in globalVal._VERSION_LIST:
            if len(result_json) == 0 or not 'conclusion' in result_json[version]:
                continue
            version_data = result_json[version]['conclusion']
            for key in version_data:
                if key == 'wp_time' and version_data['wp_count']:
                    version_data[key] /= version_data['wp_count']
                elif 'time' in key:
                    if version_data['speed_query_num'] != 0:
                        version_data[key] /= version_data['speed_query_num']
                elif 'asr_error_code_info' == key:
                    for k in version_data[key]:
                        err_info = version_data[key][k]['info']
                        if not k in self.__asr_err_code[audio_type]:
                            self.__asr_err_code[audio_type].update({k: []})
                        for err_code in err_info:
                            if not err_code in self.__asr_err_code[audio_type][k]:
                                self.__asr_err_code[audio_type][k].append(err_code)
            if 'wp_mode' in version_data:
                if self.__wak_mode == 0:
                    self.__wak_mode = version_data['wp_mode']
                for wpWords in version_data['wp_status']:
                    speech_parse.cal_wp_status(version_data['wp_status'][wpWords])
                    if wpWords != 'summary' and not wpWords in self.__true_wak_words \
                        and not 'kws' in scene_name:
                        self.__true_wak_words.append(wpWords)
                for wpWords in conclusion_data['wp_false']:
                    if wpWords != 'summary' and not wpWords in self.__false_wak_words:
                        self.__false_wak_words.append(wpWords)
            if version_data['wp_query_count'] != 0:
                version_data['wp_rate'] = '%.3f%%' % \
                    (float(version_data['wp_count']) * 100 / float(version_data['wp_query_count']))
            if version_data['valid_asr_request_count'] != 0:
                version_data['success_rate'] = '%.3f%%' % \
                    (float(version_data['success_count']) * 100 / float(version_data['valid_asr_request_count']))
                version_data['recall_rate'] = '%.3f%%' % \
                    (float(version_data['recall_count']) * 100 / float(version_data['valid_asr_request_count']))
            else:
                version_data.update({'success_rate': '0.0%', 'recall_rate': '0.0%'})
            if 'roleasr' in scene_name:
                speech_parse._update_role_perf(scene_name, version, version_data['role_accuracy'])
                self.role_list = []
                for role in list(version_data['role_accuracy']['role_info'].keys()) + ['sum']:
                    self.role_list.append(role)
            elif 'asr' in scene_name:
                answer_path = os.path.join(globalVal.get_save_path(scene_name), 
                    scene_name.split('/')[0] + '_answer_result.txt')
                result_path = os.path.join(globalVal.get_save_path(scene_name), 
                    version + '_' + scene_name.split('/')[0] + '_recognize_result.txt')
                if 'semanticasr' in scene_name:
                    wer, ser = speech_parse.cal_asr_wer(answer_path, result_path, version_data['asr_count'])
                    err = []
                else:
                    wer, ser, err = speech_parse.cal_asr_wer(answer_path, result_path, -1)
                version_data['wer'] = wer
                version_data['ser'] = ser
                # 0226 当前计算的是取交集前各版本的结果，不是取交集后用于准出的！
                version_data['err'] = err
                if len(err):
                    total_e, sub_e, del_e, in_e, words = err
                    if int(words) > 0:
                        version_data.update({'sub_e': "%.3f%%" % (float(sub_e) * 100 / float(words))})
                        version_data.update({'del_e': "%.3f%%" % (float(del_e) * 100 / float(words))})
                        version_data.update({'in_e': "%.3f%%" % (float(in_e) * 100 / float(words))})
                    else:
                        version_data.update({'sub_e': '0.0%', 'del_e': '0.0%', 'in_e': '0.0%'})
            # 全场景速度指标汇总
            tmp = {
                'wp_speed_list': version_data['wp_status']['summary']['wp_speed_list'],
                'asr_speed_list': version_data['asr_speed_list']
            }
            if audio_type in self.__speed_result:
                if version in self.__speed_result[audio_type]:
                    self.__speed_result[audio_type][version]['wp_speed_list'] += tmp['wp_speed_list']
                    for k in tmp['asr_speed_list']:
                        self.__speed_result[audio_type][version]['asr_speed_list'][k] += tmp['asr_speed_list'][k]
                else:
                    self.__speed_result[audio_type].update({version: copy.deepcopy(tmp)})
            else:
                self.__speed_result.update({audio_type: {version: copy.deepcopy(tmp)}})
        globalVal._AIEAP_JSON.update({scene_name.split('/')[0]: result_json})

        return result_json

    def update_scene_excel(self, work_sheet, row, sampleset_name, scene_data):
        """
        更新"场景.xlsx"表格里，多个版本对比的识别结果展示
        0:底色飘黄，1:底色飘黄、框体加粗，2:底色飘红、框体加粗，3:底色飘绿、框体加粗，4:字体飘红，5:字体飘绿，6:字体飘蓝
        """
        version_index = 0
        if 'roleasr' in sampleset_name:
            write_flag = 0
        elif 'semanticasr' in sampleset_name:
            write_flag = 1
        else:
            write_flag = 2
        xls.write_content(work_sheet, row, 1, str(sampleset_name), 3)
        xls.write_content(work_sheet, row + 1, 2, '标注答案', 0) 
        for version in globalVal._VERSION_LIST:
            version_index += 1
            col = version_index * 3
            xls.write_content(work_sheet, row + 1, col + 1, '识别SN', 0) 
            xls.write_content(work_sheet, row + 1, col + 2, 'corpus_id', 0) 
            xls.write_content(work_sheet, row, col, str(version), 3)
            xls.write_content(work_sheet, row + 1, col, '识别结果', 0)
            add_row_count = 2
            query_answer = scene_data['answer']
            query_result = scene_data[version]['query_result']
            for query_name in query_result:
                if query_name in query_answer:
                    asr_answer = query_answer[query_name]['content']
                else:
                    continue
                asr_desc = query_result[query_name]['desc']
                asr_sn = query_result[query_name]['sn']
                asr_corpus_id = query_result[query_name]['corpus_id']
                if asr_desc == '':
                    asr_result = query_result[query_name]['recg_result']
                else:
                    if 'no asr' in asr_desc:
                        asr_result = '(' + asr_desc + ')'
                    else:
                        asr_result = query_result[query_name]['recg_result']
                        if asr_result != '':
                            asr_result += '(' + asr_desc + ')'
                        else:
                            asr_result = asr_desc
                if 'unsupported' in asr_result or 'first asr' in asr_result or 'net_err' in asr_result \
                    or 'no wakeup' in asr_result:
                    _font = 0
                elif 'no asr' in asr_result or 'err:' in asr_result:
                    _font = 4
                else:
                    _font = -1
                    if write_flag:
                        if write_flag == 1:
                            if asr_result.split('(')[0] != asr_answer.split('#', 1)[1]:
                                _font = 6
                        elif asr_result.split('(')[0] != asr_answer:
                            _font = 6
                if version_index == 1:
                    if 'split' in sampleset_name:
                        xls.write_content(work_sheet, row + add_row_count, 1, query_name.split('_')[-1])
                    xls.write_content(work_sheet, row + add_row_count, 2, asr_answer)
                xls.write_content(work_sheet, row + add_row_count, col, asr_result, _font)
                if asr_sn is not None and globalVal._PARAM['ASR_SN_WRITE_FLAG']:
                    xls.write_content(work_sheet, row + add_row_count, col + 1, asr_sn)
                    if asr_corpus_id != "-1":
                        # 车机离线，无corpus_id,默认为-1，则不显示
                        xls.write_content(work_sheet, row + add_row_count, col + 2, asr_corpus_id)
                add_row_count += 1
            if not 'asr' in sampleset_name:
                continue
            if version_index == 1:
                if write_flag == 1:
                    xls.write_content(work_sheet, row + add_row_count + 1, 2, '识别准确率')
                    xls.write_content(work_sheet, row + add_row_count + 2, 2, '识别错误条数')
                else:
                    xls.write_content(work_sheet, row + add_row_count + 1, 2, '字准')
                    xls.write_content(work_sheet, row + add_row_count + 2, 2, '句准')
            xls.write_content(work_sheet, row + add_row_count + 1, 
                col, scene_data[version]['conclusion']['wer'])
            xls.write_content(work_sheet, row + add_row_count + 2, 
                col, scene_data[version]['conclusion']['ser'])
        return row + add_row_count + 12

    def update_detail_excel(self, work_sheet, row, sampleset_name, scene_data):
        """
        更新"统计结果.xlsx"的sheet"(单版本、单测试集、无对比)"的识别结果展示
        """
        if 'semanticasr' in sampleset_name:
            xls.write_title(self.__result_worksheet_detail, 
                globalVal._WORK_TITLE_LIST[0:14] + ['识别准确率', '识别错误条数'] + globalVal._WORK_TITLE_LIST[16:21], 
                row)
        else:
            xls.write_title(self.__result_worksheet_detail, 
                globalVal._WORK_TITLE_LIST[0:21] + ['首次识别请求成功次数'], row)
        if 'voiceprint' in sampleset_name:
            xls.write_title(self.__result_worksheet_detail, 
                ['声纹判定总条数', '正确判定次数', '错误判定次数'], row, 23)
        for version in scene_data:
            # 此处为DCI专门设计，DCI的版本为"SDK版本_距离"，不在globalVal._VERSION_LIST里
            right_version_flag = False
            for sdk_vers in globalVal._VERSION_LIST:
                if sdk_vers in version:
                    right_version_flag = True
            if not right_version_flag:
                continue
            # 开始写『单版本、单集合、不对比』sheet
            row += 1
            version_data = scene_data[version]['conclusion']
            xls.write_content(work_sheet, row, 1, version)
            xls.write_content(work_sheet, row, 2, sampleset_name)
            col = 3
            if 'mutiplyasr' in sampleset_name or 'pureasr' in sampleset_name:
                for key in globalVal._REPORT_KEY_LIST:
                    if key == 'invalid_asr_request_count':
                        xls.write_content(work_sheet, row, col, 
                            str(version_data[key]) + '/' + str(version_data['valid_asr_request_count']))
                    elif key == 'success_count':
                        xls.write_content(work_sheet, row, col, 
                            str(version_data['recall_count'])  + '/' + str(version_data[key]))
                    elif key == 'success_rate':
                        xls.write_content(work_sheet, row, col, 
                            version_data['recall_rate']  + '/' + version_data[key])
                    elif key == 'false_asr_count':
                        xls.write_content(work_sheet, row, col, 
                            str(version_data[key]) + '/' + str(version_data['false_interact_count']))
                    elif key in version_data:
                        xls.write_content(work_sheet, row, col, version_data[key])
                    col += 1
            else:
                for key in globalVal._REPORT_KEY_LIST:
                    if key == 'invalid_asr_request_count':
                        xls.write_content(work_sheet, row, col, 
                            str(version_data[key]) + '/' + str(version_data['valid_asr_request_count']))
                    elif key in version_data:
                        xls.write_content(work_sheet, row, col, version_data[key])
                    col += 1
                if 'voiceprint' in sampleset_name:
                    for key in ['vp_total', 'vp_right', 'vp_wrong']:
                        xls.write_content(work_sheet, row, col, version_data[key])
                        col += 1
        return row + 2

    def merge_scene_data(self):
        """统计结果汇总
        对齐后的详细数据按测试场景汇总到一起，便于人工比对
        """
        global_result = globalVal._DATABASE
        # [1] 按 "类型_场景_米数" 进行数据汇总
        for scene_name in globalVal._NEW_DICT:
            scene_result = global_result
            for item in globalVal.get_scene_list(scene_name):
                scene_result = scene_result[item]
            # 测试场景.xls（summary），更新单个测试集、多个版本识别结果的对比展示
            if 'asr' in scene_name:
                wb_path = os.path.join(globalVal._WORK_DIRECTORY, scene_name + '.xlsx')
                scene_workbook = globalVal._WORKBOOK_DICT[wb_path]
                start_row = 1
                if 'summary' in scene_workbook.sheetnames:
                    scene_worksheet = scene_workbook['summary']
                else:
                    scene_worksheet = scene_workbook.create_sheet('summary', 0)
            # 单个测试集，数据更新&汇总
            for audio_name in globalVal._NEW_DICT[scene_name]:
                intersect_count = 0
                for version in globalVal._VERSION_LIST:
                    if version in scene_result[audio_name]:
                        intersect_count += 1
                    elif not 'wp_dci' in scene_name:
                        globalVal.record_err(version + '/' + audio_name, 4102, '该场景数据不完整，不参与最终结果统计')
                if intersect_count == self.__version_num:
                    # update_scene_excel: 测试场景.xls（summary）
                    # 汇总取交集后的指标,需要满足所有版本都存在的条件
                    if 'asr' in scene_name:
                        if not 'roleasr' in scene_name and not globalVal._MANUAL_TEST_MODE:
                            scene_result[audio_name].update({
                                'comparison': self.get_compare_result(scene_name + '/' + audio_name, True)
                            })
                        if 'excel_count' in scene_result[audio_name]:
                            start_row = scene_result[audio_name]['excel_count']
                        else:
                            if scene_worksheet.max_row > 2:
                                start_row = scene_worksheet.max_row + 10
                        start_row = self.update_scene_excel(
                            scene_worksheet, start_row, scene_name + '/' + audio_name, scene_result[audio_name]
                        )
                    elif 'voiceprint' in scene_name:
                        scene_result[audio_name].update({
                            'comparison': self.get_compare_result(scene_name + '/' + audio_name, True)
                        })
                elif 'wp_dci' in scene_name:
                    # 提取 dci data
                    dci_thread = speech_parse.WpDCIdataExtract(
                        scene_name + '/' + audio_name, globalVal._NEW_DICT[scene_name][audio_name]
                    )
                    scene_result[audio_name].update({'dciinfo': dci_thread.get_result_json()})
            # 单个测试场景，数据更新&汇总
            if not 'wp_dci' in scene_name:
                scene_result.update({'summary': self.get_summary_result(scene_name + '/')})
            if 'voiceprint' in scene_name or \
                ('asr' in scene_name and not 'roleasr' in scene_name and not globalVal._MANUAL_TEST_MODE):
                scene_result['summary'].update({'comparison': self.get_compare_result(scene_name + '/', False)})
            # 20201105 声纹统计，额外增加按场景统计
            if 'voiceprint' in scene_name:
                environment = scene_name.split('_')[1] + '_' + scene_name.split('_')[2]
                if environment in self.__vp_scene_result:
                    for name in self.__vp_scene_result[environment]:
                        if name == 'comparison':
                            for version in self.__vp_scene_result[environment][name]:
                                for key in self.__vp_scene_result[environment][name][version]:
                                    self.__vp_scene_result[environment][name][version][key] += \
                                        scene_result['summary']['comparison'][version][key]
                        else:
                            for key in self.__vp_scene_result[environment][name]['conclusion']:
                                if 'count' in key or 'vp' in key:
                                    self.__vp_scene_result[environment][name]['conclusion'][key] += \
                                        scene_result['summary'][name]['conclusion'][key]
                else:
                    self.__vp_scene_result.update({environment: copy.deepcopy(scene_result['summary'])})
        # [2] 特殊指标，如速度等，需要按测试类型进行全场景数据汇总
        self.update_type_data()

    def update_type_data(self):
        """按测试类型进行全场景数据汇总
        """
        for audio_type in self.__speed_result:
            for version in globalVal._VERSION_LIST:
                version_data = self.__speed_result[audio_type][version]
                version_data.update({
                    'wp_count': len(version_data['wp_speed_list']),
                    'wp_time': speech_parse._get_mean(version_data['wp_speed_list']),
                    'speed_query_num': len(version_data['asr_speed_list']['finish_speed'])
                })
                for k in version_data['asr_speed_list']:
                    version_data.update({
                        k.replace('speed', 'time'): speech_parse._get_mean(version_data['asr_speed_list'][k])
                    })

    def getWpLocationReport(self):
        """
        按场景和版本汇总唤醒声源定位结果，生成report
        """
        if not os.path.exists(globalVal._LOCATION_RECORD_FILE_NAME):
            return
        try:
            result_worksheet = self.__result_workbook.create_sheet('（声源定位准确率）')
        except:
            logging.error('（声源定位准确率）sheet创建失败')
            return

        # 从_LOCATION_RECORD_FILE_NAME文件中读取信息
        title_list = ['测试集名称', '场景', '距离', 'SDK版本']
        count_list = ['音频条数', '唤醒次数', '定位正确次数', '相邻灯位次数']
        wp_location_list = ['total_count', 'wp_count', 'right_count', 'in_count']
        EMPTY_LIST = {}.fromkeys(wp_location_list, 0)
        _LOCATION_RECORD_LIST = {'汇总': {}}
        location_list = []
        scene_count = 0
        with open(globalVal._LOCATION_RECORD_FILE_NAME, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                if line == '':
                    continue
                if not '[' in line:
                    sdk_version, audio_scene, environment, location, result = line.split(':')
                    environment, distance = environment.rsplit('_', 1)
                    right_count, in_count, wp_count, total_count = result.split(',')
                    RESULT_LIST = {
                        'total_count': int(total_count),
                        'wp_count': int(wp_count),
                        'right_count': int(right_count),
                        'in_count': int(in_count) + int(right_count)
                    }
                    if not audio_scene in _LOCATION_RECORD_LIST:
                        _LOCATION_RECORD_LIST.update({audio_scene: {environment: {distance: {sdk_version: \
                            {location: RESULT_LIST, 'sum': copy.deepcopy(EMPTY_LIST)}}}}})
                    elif not environment in _LOCATION_RECORD_LIST[audio_scene]:
                        _LOCATION_RECORD_LIST[audio_scene].update({environment: \
                            {distance: {sdk_version: {location: RESULT_LIST, 'sum': copy.deepcopy(EMPTY_LIST)}}}})
                    elif not distance in _LOCATION_RECORD_LIST[audio_scene][environment]:
                        _LOCATION_RECORD_LIST[audio_scene][environment].update( \
                            {distance: {sdk_version: {location: RESULT_LIST}}})
                    elif not sdk_version in _LOCATION_RECORD_LIST[audio_scene][environment][distance]:
                        _LOCATION_RECORD_LIST[audio_scene][environment][distance].update( \
                            {sdk_version: {location: RESULT_LIST}})
                    elif not location in _LOCATION_RECORD_LIST[audio_scene][environment][distance][sdk_version]:
                        _LOCATION_RECORD_LIST[audio_scene][environment][distance][sdk_version].update( \
                            {location: RESULT_LIST})
                    else:
                        logging.error('重复出现:' + line)
                        globalVal.record_err(line, 4103, '重复出现')

                    if not location in location_list:
                        location_list.append(location)
                    loc_data = _LOCATION_RECORD_LIST[audio_scene][environment][distance][sdk_version]
                    _LOCATION_RECORD_SUMMARY = loc_data[location]
                    if not environment in _LOCATION_RECORD_LIST['汇总']:
                        _LOCATION_RECORD_LIST['汇总'].update({environment: {distance: {sdk_version: \
                            {location: copy.deepcopy(EMPTY_LIST), 'sum': copy.deepcopy(EMPTY_LIST)}}}})
                    elif not distance in _LOCATION_RECORD_LIST['汇总'][environment]:
                        _LOCATION_RECORD_LIST['汇总'][environment].update({distance: {sdk_version: \
                            {location: copy.deepcopy(EMPTY_LIST), 'sum': copy.deepcopy(EMPTY_LIST)}}})
                    elif not sdk_version in _LOCATION_RECORD_LIST['汇总'][environment][distance]:
                        _LOCATION_RECORD_LIST['汇总'][environment][distance].update({sdk_version: \
                            {location: copy.deepcopy(EMPTY_LIST), 'sum': copy.deepcopy(EMPTY_LIST)}})
                    elif not location in _LOCATION_RECORD_LIST['汇总'][environment][distance][sdk_version]:
                        _LOCATION_RECORD_LIST['汇总'][environment][distance][sdk_version].update( \
                            {location: copy.deepcopy(EMPTY_LIST)})
                    if not 'sum' in loc_data:
                        loc_data.update({'sum': copy.deepcopy(EMPTY_LIST)})
                    for key in wp_location_list:
                        tmp_data = int(_LOCATION_RECORD_SUMMARY[key])
                        loc_data['sum'][key] += tmp_data
                        _LOCATION_RECORD_LIST['汇总'][environment][distance][sdk_version][location][key] += tmp_data
                        _LOCATION_RECORD_LIST['汇总'][environment][distance][sdk_version]['sum'][key] += tmp_data

        for audio_scene in _LOCATION_RECORD_LIST:
            if audio_scene == '汇总':
                continue
            for environment in _LOCATION_RECORD_LIST[audio_scene]:
                for distance in _LOCATION_RECORD_LIST[audio_scene][environment]:
                    scene_count += 1
        # 根据audio_scene, 汇总定位准确率
        start_col = 5
        start_row = scene_count * self.__version_num + 6
        xls.write_title(result_worksheet, title_list, start_row, 1, 1)
        row = start_row + 2
        for audio_scene in _LOCATION_RECORD_LIST:
            audio_scene_start_row = row
            for environment in _LOCATION_RECORD_LIST[audio_scene]:
                environ_start_row = row
                for distance in _LOCATION_RECORD_LIST[audio_scene][environment]:
                    distance_start_row = row
                    for sdk_version in _LOCATION_RECORD_LIST[audio_scene][environment][distance]:
                        col_idx = 0
                        xls.write_content(result_worksheet, row, 4, sdk_version)
                        location_idx = 0
                        _LOCATION_DATA = _LOCATION_RECORD_LIST[audio_scene][environment][distance][sdk_version]
                        # 单个方位数据
                        for location in location_list:
                            xls.write_content(result_worksheet, [start_row, start_row], 
                                [start_col + location_idx * 4, start_col + location_idx * 4 + 3],
                                location, 1)
                            xls.write_title(result_worksheet, count_list, start_row + 1, start_col + col_idx)
                            if str(location) in _LOCATION_DATA:
                                _LOCATION_RECORD_SCENE_SUMMARY = _LOCATION_DATA[location]
                                for key in wp_location_list:
                                    xls.write_content(result_worksheet, 
                                        row, start_col + col_idx, 
                                        _LOCATION_DATA[location][key])
                                    col_idx += 1
                            else:
                                col_idx += 4
                            location_idx += 1
                        # 多个方位汇总数据
                        xls.write_content(result_worksheet, [start_row, start_row], 
                            [start_col + location_idx * 4, start_col + location_idx * 4 + 3], 
                            '汇总', 1)
                        xls.write_title(result_worksheet, count_list, start_row + 1, start_col + col_idx)
                        for key in wp_location_list:
                            xls.write_content(result_worksheet, row, start_col + col_idx, 
                                _LOCATION_DATA['sum'][key])
                            col_idx += 1
                        wp_rate = '0.0%'
                        right_rate = '0.0%'
                        in_rate = '0.0%'
                        if _LOCATION_DATA['sum']['total_count']:
                            wp_rate = '%.3f%%' % (float(_LOCATION_DATA['sum']['wp_count']) \
                                * 100 / float(_LOCATION_DATA['sum']['total_count']))
                            if _LOCATION_DATA['sum']['wp_count']:
                                right_rate = '%.3f%%' % (float(_LOCATION_DATA['sum']['right_count']) \
                                    * 100 / float(_LOCATION_DATA['sum']['wp_count']))
                                in_rate = '%.3f%%' % (float(_LOCATION_DATA['sum']['in_count']) \
                                    * 100 / float(_LOCATION_DATA['sum']['wp_count']))
                        _LOCATION_DATA['sum'].update({
                            'wp_rate': wp_rate,
                            'right_rate': right_rate,
                            'in_rate': in_rate
                        })
                        row += 1
                    xls.write_content(result_worksheet, [distance_start_row, row - 1], [3, 3], distance)
                xls.write_content(result_worksheet, [environ_start_row, row - 1], [2, 2], environment)
            xls.write_content(result_worksheet, [audio_scene_start_row, row - 1], [1, 1], audio_scene)

        # 根据audio_scene, 汇总定位准确率
        count_list = ['唤醒音频总条数', '唤醒对齐次数', '唤醒率', '唤醒率DIFF', 
            '定位正确次数(标记位置±15°)', '定位正确率', '定位正确率DIFF', 
            '相邻灯位次数(标记位置±45°)', '相邻灯位正确率', '相邻灯位正确率DIFF']
        wp_location_list = ['total_count', 'wp_count', 'wp_rate', '',
            'right_count', 'right_rate', '', 'in_count', 'in_rate']
        xls.write_title(result_worksheet, title_list, 1, 1, 1)
        xls.write_title(result_worksheet, count_list, 1, start_col, 1)
        row = 3
        for environment in _LOCATION_RECORD_LIST['汇总']:
            environ_start_row = row
            for distance in _LOCATION_RECORD_LIST['汇总'][environment]:
                distance_start_row = row
                ver_idx = 0
                scene_data = _LOCATION_RECORD_LIST['汇总'][environment][distance]
                for sdk_version in globalVal._VERSION_LIST:
                    col_idx = 0
                    location_idx = 0
                    xls.write_content(result_worksheet, row, 4, sdk_version)
                    base_version = scene_data[globalVal._VERSION_LIST[0]]['sum']
                    _LOCATION_DATA = scene_data[sdk_version]['sum']
                    for key in wp_location_list:
                        if key != '':
                            xls.write_content(result_worksheet, row, start_col + col_idx, 
                                _LOCATION_DATA[key])
                        col_idx += 1
                    if ver_idx > 0:
                        xls.write_content_with_color(result_worksheet, row, start_col + 3, 
                            get_substraction(_LOCATION_DATA['wp_rate'], base_version['wp_rate']))
                        xls.write_content_with_color(result_worksheet, row, start_col + 6, 
                            get_substraction(_LOCATION_DATA['right_rate'], base_version['right_rate']))
                        xls.write_content_with_color(result_worksheet, row, start_col + 9, 
                            get_substraction(_LOCATION_DATA['in_rate'], base_version['in_rate']))
                    ver_idx += 1
                    row += 1
                xls.write_content(result_worksheet, [distance_start_row, row - 1], [3, 3], distance)
            xls.write_content(result_worksheet, [environ_start_row, row - 1], [2, 2], environment)
        xls.write_content(result_worksheet, [3, row - 1], [1, 1], '汇总')

        self.__result_workbook.save(globalVal._RESULT_WORKBOOK_NAME)

    def write_wp_report(self):
        """写入统计结果.xlsx(唤醒专项)
        """
        title_flag = {'wp': True, 'asr': True}
        global_result = globalVal._DATABASE
        _wp_report = self.__result_workbook.create_sheet('唤醒专项报告')
        if self.__wak_mode == 1:
            wak_list = ['queryCount', 'detectCount', 'rightCount', 'wrongWordCount', 'wrongIndexCount', 
                'wpAvgTime', 'wpMaxTime', 'rightRate']
        elif self.__wak_mode == 2:
            wak_list = ['queryCount', 'rightWordCount', 'wrongWordCount', 'wrongIndexCount']
        elif self.__wak_mode == 3:
            wak_list = ['queryCount', 'detectCount', 'normalCount', 
                'whisperCount', 'wpRate', 'rightRate', 'wpAvgTime']
        else:
            wak_list = []
        
        if globalVal._DISPLAY_BY_COL:
            self.write_codriver_wp_report(title_flag, global_result, _wp_report, wak_list)
        else:
            row = 1
            col = 1
            for scene_name in globalVal._SAMPLE_LIST:
                if (not 'wp' in scene_name and not 'asr' in scene_name) or 'wp_dci' in scene_name:
                    continue
                audio_type = get_audio_type(scene_name)
                # [1] 写表格title
                if audio_type in title_flag:
                    if title_flag[audio_type]:
                        title_flag[audio_type] = False
                        wak_info_col = write_wp_report_title_by_row(audio_type, _wp_report, row, 
                            [self.__true_wak_words, self.__false_wak_words], self.__wak_mode)
                        row += 3
                else:
                    continue
                scene_start_row = row
                # [2] 按测试场景写入数据
                scene_result = global_result
                for item in globalVal.get_scene_list(scene_name):
                    scene_result = scene_result[item]
                result_summary = scene_result['summary']
                for version in globalVal._VERSION_LIST:
                    version_start_row = row
                    # 得到当前版本的conclusion数据
                    curr_conclusion = result_summary[version]['conclusion']
                    # 写入唤醒详细信息
                    write_wak_detail_info_by_row(_wp_report, row, col + 2, wak_info_col, 
                        [self.__true_wak_words, self.__false_wak_words], wak_list, curr_conclusion, self.__wak_mode)
                    row += 1
                    if self.__wak_mode == 3:
                        row += len(curr_conclusion['wp_status']['summary']['whisperInfo']) - 1
                    xls.write_content(_wp_report, [version_start_row, row - 1], col + 1, version)
                # [3] 第1列写入测试场景信息
                xls.write_content(_wp_report, [scene_start_row, row - 1], 1, scene_name)

            self.__result_workbook.save(globalVal._RESULT_WORKBOOK_NAME)

    def write_codriver_wp_report(self, title_flag, global_result, _wp_report, wak_list):
        """写入统计结果.xlsx(唤醒专项)
        """
        wak_tile = globalVal._R_CODRIVER_WAK_LIST
        row = (self.__version_num + 1) * (len(self.__true_wak_words) + 1) + 5
        wak_num = write_wp_report_title_by_col(_wp_report, self.__true_wak_words)
        
        ver_idx = 0
        add_row = 0
        for version in globalVal._VERSION_LIST:
            col = 1
            col_1 = 3
            col_2 = col_1 + wak_num[0] + 2
            scene_start_row = row
            for scene_name in globalVal._SAMPLE_LIST:
                if (not 'wp' in scene_name and not 'asr' in scene_name) or 'wp_dci' in scene_name:
                    continue
                audio_type = get_audio_type(scene_name)
                # [1] 写表格title
                if audio_type in title_flag:
                    xls.write_content(_wp_report, row, [col + 1, col + len(wak_tile)], version, 9)
                    xls.write_content(_wp_report, row + 1, [col + 1, col + len(wak_tile)], scene_name, 9)
                    row += 2
                else:
                    continue
                # [2] 按测试场景写入数据
                scene_result = global_result
                for item in globalVal.get_scene_list(scene_name):
                    scene_result = scene_result[item]
                result_summary = scene_result['summary']
                # 得到当前版本的conclusion数据
                curr_conclusion = result_summary[version]['conclusion']
                # 写入唤醒详细信息
                xls.write_title(_wp_report, wak_tile, row, col + 1, 0, 0, 9)
                row = write_wak_detail_info_by_col(_wp_report, row + 1, col + 1, wak_list, curr_conclusion)
                # 写入唤醒汇总信息
                if 'kws' in scene_name:
                    xls.write_content(_wp_report, 3 + ver_idx * len(self.__true_wak_words), 
                        col_2, curr_conclusion['wp_status']['summary']['rightRate'])
                    col_2 += 1
                else:
                    i = 0
                    for wp_word in self.__true_wak_words:
                        if wp_word in curr_conclusion['wp_status']:
                            xls.write_content(_wp_report, 3 + ver_idx * len(self.__true_wak_words) + i, 
                                col_1, curr_conclusion['wp_status'][wp_word]['rightRate'])
                        i += 1
                    col_1 += 1
                if add_row < row - scene_start_row:
                    add_row = row - scene_start_row
                row = scene_start_row
                col += len(globalVal._R_CODRIVER_WAK_LIST) + 1
            row += add_row + 3
            ver_idx += 1

        self.__result_workbook.save(globalVal._RESULT_WORKBOOK_NAME)
        
