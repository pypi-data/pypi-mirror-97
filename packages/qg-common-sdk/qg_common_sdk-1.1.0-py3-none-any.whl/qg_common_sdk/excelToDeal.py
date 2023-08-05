from openpyxl import *
import openpyxl
from io import BytesIO,StringIO
import xlrd
from qg_common_sdk.catchException import SysCheckException

# 实例化

# 激活 worksheet

class ExcelToDeal:

    #导入
    def excelToList(self,file,columnList=[]):
        data = xlrd.open_workbook(file_contents=file)
        table = data.sheets()[0] 
        rowMax = table.nrows
        colMax = table.ncols
        if not columnList:
            columnList = table.row_values(0)
        print (len(columnList))
        print (colMax)
        if colMax!= len(columnList):
            raise SysCheckException(data='',msg="表头长度和接口不对应"+str(len(columnList))+':'+str(colMax))
        dataList = []
        
        for col_index in range(rowMax-1):
            data = {}
            rowOneList = table.row_values(col_index+1)
            rowOneLen = len(rowOneList)
            for cellindex in range(rowOneLen):
                data[columnList[cellindex]] = rowOneList[cellindex]
            dataList.append(data)
        return dataList


    # 导出
    def listToExcel(self,titleList,dataList):
        wb = Workbook()
        sheet = wb.create_sheet(index=0)
        for i in range(len(titleList)):
            sheet.cell(1,i+1,titleList[i])
        for j in range(len(dataList)):
            for i in range(len(dataList[j])):
                value = dataList[j][i]
                print (type(value))
                if isinstance(value,list):
                    value =  str(value)
                if isinstance(value,dict):
                    value =  str(value)
                sheet.cell(j+2,i+1,value)
        sio=BytesIO()
        wb.save(sio)
        return sio 




class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data